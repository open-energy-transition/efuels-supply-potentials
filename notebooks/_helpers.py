# SPDX-FileCopyrightText:  Open Energy Transition gGmbH
#
# SPDX-License-Identifier: AGPL-3.0-or-later

import warnings
from typing import Dict, List, Optional, Tuple
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from matplotlib.lines import Line2D
import matplotlib.patches as mpatches
from shapely.geometry import LineString
from collections import OrderedDict
import plotly.graph_objects as go
import plotly.express as px
from collections import defaultdict
from IPython.display import display, HTML
from matplotlib.font_manager import FontProperties
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
import matplotlib.dates as mdates
from matplotlib.legend_handler import HandlerPatch
import matplotlib.lines as mlines
import matplotlib.transforms as mtransforms
from matplotlib.patches import Patch
import matplotlib.path as mpath
from matplotlib.patches import Wedge
from matplotlib.offsetbox import AnnotationBbox, AuxTransformBox
from shapely.geometry import box
from shapely.geometry import Point
import geopandas as gpd
import cartopy.feature as cfeature
import cartopy.crs as ccrs  # For plotting maps
import xarray as xr
from pathlib import Path
from pypsa.plot import add_legend_circles, add_legend_lines, add_legend_patches
import math
import pypsa
import pycountry
import re
import seaborn as sns
from IPython.display import Image, display
import matplotlib.pyplot as plt
import os
import pandas as pd
import numpy as np
import matplotlib

matplotlib.use("Agg")


warnings.filterwarnings("ignore")


def showfig(name="_tmp.png", dpi=120):
    fig = plt.gcf()
    fig.savefig(name, dpi=dpi, bbox_inches="tight")
    display(Image(name))
    plt.close(fig)


def attach_grid_region_to_buses(network, path_shapes, distance_crs="EPSG:4326"):
    """
    Attach grid region to buses
    """
    # Read the shapefile using geopandas
    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes.rename(columns={"GRID_REGIO": "Grid Region"}, inplace=True)

    ac_dc_carriers = ["AC", "DC"]
    location_mapping = network.buses.query("carrier in @ac_dc_carriers")[["x", "y"]]

    network.buses["x"] = network.buses["location"].map(location_mapping["x"]).fillna(0)
    network.buses["y"] = network.buses["location"].map(location_mapping["y"]).fillna(0)

    pypsa_gpd = gpd.GeoDataFrame(
        network.buses,
        geometry=gpd.points_from_xy(network.buses.x, network.buses.y),
        crs=4326,
    )

    bus_cols = network.buses.columns
    bus_cols = list(bus_cols) + ["grid_region"]

    st_buses = gpd.sjoin_nearest(shapes, pypsa_gpd, how="right")

    network.buses.rename(columns={"region": "emm_region"}, inplace=True)
    network.buses["grid_region"] = st_buses["Grid Region"]


def attach_state_to_buses(network, path_shapes, distance_crs="EPSG:4326"):
    """
    Attach state to buses
    """
    # Read the shapefile using geopandas
    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes["ISO_1"] = shapes["ISO_1"].apply(lambda x: x.split("-")[1])
    shapes.rename(columns={"ISO_1": "State"}, inplace=True)

    ac_dc_carriers = ["AC", "DC"]
    location_mapping = network.buses.query("carrier in @ac_dc_carriers")[["x", "y"]]

    network.buses["x"] = network.buses["location"].map(location_mapping["x"]).fillna(0)
    network.buses["y"] = network.buses["location"].map(location_mapping["y"]).fillna(0)

    pypsa_gpd = gpd.GeoDataFrame(
        network.buses,
        geometry=gpd.points_from_xy(network.buses.x, network.buses.y),
        crs=4326,
    )

    bus_cols = network.buses.columns
    bus_cols = list(bus_cols) + ["State"]

    st_buses = gpd.sjoin_nearest(shapes, pypsa_gpd, how="right")

    network.buses["state"] = st_buses["State"]


def compute_demand(network):
    """
    Compute total demand by region and by state
    """
    static_load_carriers = [
        "rail transport electricity",
        "agriculture electricity",
        "industry electricity",
    ]
    dynamic_load_carriers = ["AC", "services electricity", "land transport EV"]

    ac_loads = network.loads.query("carrier in 'AC'").index
    ac_profile = (
        network.loads_t.p_set[ac_loads]
        .multiply(network.snapshot_weightings.objective, axis=0)
        .sum()
        / 1e6
    )
    ac_load_bus = (
        ac_profile.to_frame()
        .reset_index()
        .rename(columns={0: "load", "Load": "region"})
    )
    ac_load_bus["carrier"] = "AC"

    nhours = network.snapshot_weightings.objective.sum()
    static_load = (
        network.loads.groupby(["bus", "carrier"]).sum()[["p_set"]].reset_index()
    )
    static_load_bus = static_load.query("carrier in @static_load_carriers").reset_index(
        drop=True
    )
    static_load_bus["p_set"] = static_load_bus.p_set * nhours / 1e6

    services_profile = network.loads_t.p_set.filter(like="services electricity") / 1e6
    services_load = (
        services_profile.multiply(network.snapshot_weightings.objective, axis=0)
        .sum()
        .to_frame()
        .reset_index()
        .rename(columns={0: "services electricity load", "Load": "bus"})
    )
    services_load["region"] = services_load["bus"].str.extract(r"(US\d{1} \d{1,2})")
    services_load.rename(columns={"services electricity load": "load"}, inplace=True)
    services_load["carrier"] = "services electricity"

    static_load_bus["region"] = static_load_bus["bus"].str.extract(r"(US\d{1} \d{1,2})")
    agriculture_electricity_load = static_load_bus.query(
        "carrier == 'agriculture electricity'"
    )
    agriculture_electricity_load.rename(columns={"p_set": "load"}, inplace=True)

    industry_electricity_load = static_load_bus.query(
        "carrier == 'industry electricity'"
    )
    industry_electricity_load.rename(columns={"p_set": "load"}, inplace=True)

    rail_transport_electricity_load = static_load_bus.query(
        "carrier == 'rail transport electricity'"
    )
    rail_transport_electricity_load.rename(columns={"p_set": "load"}, inplace=True)

    ev_profile = network.loads_t.p_set.filter(like="land transport EV")
    ev_load = (
        (ev_profile.multiply(network.snapshot_weightings.objective, axis=0).sum() / 1e6)
        .to_frame()
        .reset_index()
        .rename(columns={0: "load", "Load": "bus"})
    )
    ev_load["region"] = ev_load["bus"].str.extract(r"(US\d{1} \d{1,2})")
    ev_load["carrier"] = "land transport EV"

    all_loads = pd.concat(
        [
            ac_load_bus,
            ev_load,
            services_load,
            agriculture_electricity_load,
            industry_electricity_load,
            rail_transport_electricity_load,
        ],
        axis=0,
    )

    all_loads_df_grid_region = (
        all_loads.pivot(index="region", columns="carrier", values="load")
        .fillna(0)
        .round(2)
    )
    all_loads_df_grid_region.index = all_loads_df_grid_region.index.map(
        network.buses.grid_region
    )
    all_loads_df_grid_region_sum = all_loads_df_grid_region.groupby("region").sum()

    all_loads_df_state = (
        all_loads.pivot(index="region", columns="carrier", values="load")
        .fillna(0)
        .round(2)
    )
    all_loads_df_state.index = all_loads_df_state.index.map(network.buses.state)
    all_loads_df_state_sum = all_loads_df_state.groupby("region").sum()

    return all_loads_df_grid_region_sum, all_loads_df_state_sum


def compute_data_center_load(network):
    """
    Compute data center load by grid region and by state
    """

    data_center_loads = network.loads.query("carrier in 'data center'")

    data_center_loads["grid_region"] = data_center_loads.bus.map(
        network.buses.grid_region
    )
    data_center_loads["state"] = data_center_loads.bus.map(network.buses.state)

    return data_center_loads


def compute_carrier_costs(network, rename_tech):
    """Compute total carrier costs by region and by state"""
    cost_df = network.statistics()[["Capital Expenditure", "Operational Expenditure"]]
    carrier_cost_df = (
        cost_df.reset_index(level=0, drop=True)
        .sum(axis=1)
        .reset_index()
        .rename(columns={0: "cost"})
    )
    carrier_cost_df.carrier = carrier_cost_df.carrier.map(rename_tech)
    grouped_carrier_cost_df = carrier_cost_df.groupby(["carrier"])[["cost"]].sum()

    return grouped_carrier_cost_df


def update_ac_dc_bus_coordinates(network):
    """
    For all buses with carrier 'AC' or 'DC', update their 'x' and 'y' coordinates
    based on their 'location' field and the mapping from existing AC/DC buses.
    """
    ac_dc_carriers = ["AC", "DC"]
    location_mapping = network.buses.query("carrier in @ac_dc_carriers")[["x", "y"]]
    network.buses["x"] = network.buses["location"].map(location_mapping["x"]).fillna(0)
    network.buses["y"] = network.buses["location"].map(location_mapping["y"]).fillna(0)
    return network


def fill_missing_nice_names(n, nice_names):
    """
    Fill missing nice_name values in n.carriers using the provided nice_names dict.
    Prints carriers that were missing and their new nice_name.
    """
    missing = n.carriers[n.carriers.nice_name == ""].index
    for idx in missing:
        if idx in nice_names:
            n.carriers.nice_name[idx] = nice_names[idx]
        else:
            print(f"No nice_name found for '{idx}' in nice_names dictionary.")


def fill_missing_color(n, color):
    """
    Fill missing color values in n.carriers using the provided color dict.
    Prints carriers that were missing and their new color.
    """
    missing = n.carriers[n.carriers.color == ""].index
    for idx in missing:
        if idx in color:
            n.carriers.color[idx] = color[idx]
        else:
            print(f"No color found for '{idx}' in color dictionary.")


def assign_location(n):
    for c in n.iterate_components(n.one_port_components | n.branch_components):
        ifind = pd.Series(c.df.index.str.find(" ", start=4), c.df.index)

        for i in ifind.value_counts().index:
            # these have already been assigned defaults
            if i == -1:
                continue

            names = ifind.index[ifind == i]

            c.df.loc[names, "location"] = names.str[:i]


def compute_h2_capacities(network):
    """
    Compute the total capacities (in MW) of hydrogen-related components in the network.
    Returns a DataFrame with the total capacities for each hydrogen-related component.
    """
    h2_carriers_buses = [
        "Alkaline electrolyzer large",
        "Alkaline electrolyzer medium",
        "Alkaline electrolyzer small",
        "PEM electrolyzer",
        "SOEC",
    ]

    # Filter hydrogen-related links
    hydrogen_links = network.links.query("carrier in @h2_carriers_buses").copy()

    # Merge with bus metadata (state and grid_region)
    capacity_data = hydrogen_links.merge(
        network.buses[["state", "grid_region"]],
        left_on="bus0",
        right_index=True,
        how="left",
    )

    # Use p_nom_opt directly (already in MW)
    capacity_data["p_nom_mw"] = capacity_data["p_nom_opt"]

    # Pivot table to aggregate capacity by carrier and bus
    h2_capacity_data = capacity_data.pivot_table(
        index="bus0", columns="carrier", values="p_nom_mw", aggfunc="sum", fill_value=0
    )

    # Add state and grid_region information
    h2_capacity_data["state"] = h2_capacity_data.index.map(network.buses.state)
    h2_capacity_data["grid_region"] = h2_capacity_data.index.map(
        network.buses.grid_region
    )

    return h2_capacity_data


def plot_h2_capacities_by_state(
    grouped, title, ymax, max_n_states, bar_width=0.5, height=5
):
    """
    Plot with fixed x-axis limits to preserve bar width across different state counts.
    """
    if grouped.empty:
        print(f"Skipping plot for {title} (no data)")
        return

    state = grouped.index.tolist()
    n = len(state)
    x = np.arange(n)

    techs = grouped.columns.tolist()
    bottoms = np.zeros(n)

    fig, ax = plt.subplots(figsize=(max_n_states * bar_width * 3.5, height))

    for tech in techs:
        ax.bar(x, grouped[tech].values, bar_width, bottom=bottoms, label=tech)
        bottoms += grouped[tech].values

    # Add value labels on top of each stacked bar
    for i, total in enumerate(bottoms):
        if total > 0:
            ax.text(
                x[i],
                total + 0.01 * ymax,
                f"{total:.0f}",
                ha="center",
                va="bottom",
                fontsize=8,
            )

    ax.set_xticks(x)
    ax.set_xticklabels(state, rotation=30, ha="center")
    ax.set_xlabel("State")
    ax.set_ylabel("Capacity (MW input electricity)")

    if ymax > 0:
        ax.set_ylim(0, ymax * 1.05)
    else:
        ax.set_ylim(0, 1)

    ax.set_xlim(-0.5, max_n_states - 0.5)

    ax.set_title(
        f"\nHydrogen electrolyzer capacity by State and technology - {title}\n"
    )
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1), frameon=False)

    plt.tight_layout()
    showfig()


def plot_h2_capacities_by_grid_region(
    grouped, title, ymax, max_n_grid_regions, bar_width=0.5, height=5
):
    """
    Plot with fixed x-axis limits to preserve bar width across different grid regions counts.
    """
    if grouped.empty:
        print(f"Skipping plot for {title} (no data)")
        return

    grid_region = grouped.index.tolist()
    n = len(grid_region)
    x = np.arange(n)

    techs = grouped.columns.tolist()
    bottoms = np.zeros(n)

    fig, ax = plt.subplots(figsize=(max_n_grid_regions * bar_width * 5.5, height))

    for tech in techs:
        ax.bar(x, grouped[tech].values, bar_width, bottom=bottoms, label=tech)
        bottoms += grouped[tech].values

    # Add text on top of each stacked bar
    for i in range(n):
        total = grouped.iloc[i].sum()
        ax.text(
            x[i],
            total + 0.01 * ymax,
            f"{total:.0f}",
            ha="center",
            va="bottom",
            fontsize=8,
        )

    ax.set_xticks(x)
    ax.set_xticklabels(grid_region, rotation=30, ha="center")
    ax.set_xlabel("Grid Region")
    ax.set_ylabel("Capacity (MW input electricity)")

    if ymax > 0:
        ax.set_ylim(0, ymax * 1.05)
    else:
        ax.set_ylim(0, 1)

    ax.set_xlim(-0.5, max_n_grid_regions - 0.5)

    ax.set_title(f"\nElectrolyzer capacity by Grid Region and technology - {title}\n")
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1), frameon=False)

    plt.tight_layout()
    showfig()


def create_hydrogen_capacity_map(
    network, path_shapes, distance_crs=4326, min_capacity_mw=10
):
    """
    Create a map with pie charts showing electrolyzer capacity breakdown by type for each state
    """
    if hasattr(network, "links") and len(network.links) > 0:
        # Filter for hydrogen-related links (electrolyzers)
        hydrogen_links = network.links[
            network.links["carrier"].str.contains(
                "electrolyzer|SOEC", case=False, na=False
            )
            | network.links.index.str.contains(
                "electrolyzer|SOEC", case=False, na=False
            )
        ].copy()

    capacity_data = hydrogen_links.merge(
        network.buses[["state"]],
        left_on="bus0",  # Assuming bus0 is the electrical connection
        right_index=True,
        how="left",
    )

    # capacity_data = links_with_state

    # Convert MW to MW (keep as MW for hydrogen as capacities are typically smaller)
    capacity_data["p_nom_mw"] = capacity_data["p_nom_opt"]

    print(f"Found hydrogen capacity data for {capacity_data['state'].nunique()} states")
    print("Electrolyzer types found:", capacity_data["carrier"].unique().tolist())

    # Step 2: Read and prepare shapefile
    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes["ISO_1"] = shapes["ISO_1"].apply(lambda x: x.split("-")[1])
    shapes.rename(columns={"ISO_1": "State"}, inplace=True)

    # Get state centroids for pie chart placement
    shapes_centroid = shapes.copy()
    shapes_centroid["centroid"] = shapes_centroid.geometry.centroid
    shapes_centroid["cent_x"] = shapes_centroid.centroid.x
    shapes_centroid["cent_y"] = shapes_centroid.centroid.y

    # Define colors for electrolyzer types
    unique_carriers = capacity_data["carrier"].unique()
    colors = plt.cm.Set2(np.linspace(0, 1, len(unique_carriers)))
    carrier_colors = dict(zip(unique_carriers, colors))

    # Customize colors for common electrolyzer types
    custom_colors = {
        "H2 Electrolysis": "#1f77b4",  # Blue
        "alkaline": "#ff7f0e",  # Orange
        "PEM": "#2ca02c",  # Green
        "SOEC": "#d62728",  # Red
        "AEL": "#9467bd",  # Purple
        "electrolyzer": "#8c564b",  # Brown
        "hydrogen": "#e377c2",  # Pink
        "H2": "#7f7f7f",  # Gray
    }

    # Update carrier_colors with custom colors
    for carrier, color in custom_colors.items():
        if carrier in carrier_colors:
            carrier_colors[carrier] = color

    # Create the plot
    fig, ax = plt.subplots(figsize=(30, 20))

    # Plot the base map
    shapes.plot(ax=ax, color="lightgray", edgecolor="black", alpha=0.3)

    # Group capacity data by state
    state_capacity = (
        capacity_data.groupby("state").agg({"p_nom_mw": "sum"}).reset_index()
    )

    # Filter states with minimum capacity
    states_to_plot = state_capacity["state"].tolist()

    print(
        f"Plotting {len(states_to_plot)} states with ≥{min_capacity_mw} MW input electricity"
    )

    # Create pie charts for each state
    for state in states_to_plot:
        state_data = capacity_data[capacity_data["state"] == state]

        if len(state_data) == 0:
            continue

        # Get state centroid
        state_centroid = shapes_centroid[shapes_centroid["State"] == state]
        if len(state_centroid) == 0:
            continue

        cent_x = state_centroid["cent_x"].iloc[0]
        cent_y = state_centroid["cent_y"].iloc[0]

        # Prepare pie chart data
        sizes = state_data["p_nom_mw"].values
        labels = state_data["carrier"].values
        colors_list = [carrier_colors[carrier] for carrier in labels]

        # Calculate pie chart radius based on total capacity
        total_capacity = sizes.sum()
        # Scale radius based on capacity (adjusted for MW scale)
        max_capacity = state_capacity["p_nom_mw"].max()
        radius = 0.3 + (total_capacity / max_capacity) * 1.5

        # Create pie chart
        pie_wedges, texts = ax.pie(
            sizes,
            colors=colors_list,
            center=(cent_x, cent_y),
            radius=radius,
            startangle=90,
        )

        # Add capacity label
        ax.annotate(
            f"{total_capacity:.0f} MW",
            xy=(cent_x, cent_y - radius - 0.3),
            ha="center",
            va="top",
            fontsize=12,
        )

    # Create legend
    legend_elements = []
    for carrier, color in carrier_colors.items():
        if carrier in capacity_data["carrier"].values:
            # Clean up carrier names for legend
            display_name = carrier.replace("_", " ").title()
            legend_elements.append(Line2D([0], [0], color="none", label=f"— {group} —"))

    ax.legend(
        handles=legend_elements,
        loc="center left",
        bbox_to_anchor=(1, 0.5),
        fontsize=14,
        title="Electrolyzer Type",
        title_fontsize=16,
    )

    # Step 7: Formatting - Expand map boundaries
    x_buffer = (shapes.total_bounds[2] - shapes.total_bounds[0]) * 0.1
    y_buffer = (shapes.total_bounds[3] - shapes.total_bounds[1]) * 0.1

    ax.set_xlim([-130, -65])
    ax.set_ylim([20, 55])
    ax.set_aspect("equal")
    ax.axis("off")

    ax.set_title(
        "Installed Electrolyzer Capacity by State and Type", fontsize=24, pad=30
    )

    # Add subtitle
    ax.text(
        0.5,
        0.02,
        f"Note: Only states with ≥{min_capacity_mw} MW electrolyzer capacity are shown",
        transform=ax.transAxes,
        ha="center",
        fontsize=14,
        style="italic",
    )

    plt.tight_layout()
    return fig, ax, capacity_data


def print_hydrogen_capacity_summary(capacity_data):
    """Print summary statistics of the hydrogen capacity data"""
    if len(capacity_data) == 0:
        print("No hydrogen capacity data to summarize.")
        return

    print("HYDROGEN ELECTROLYZER CAPACITY SUMMARY")
    print(f"Total installed hydrogen capacity: {capacity_data['p_nom'].sum():.1f} MW")
    print(
        f"Number of states with hydrogen capacity: {capacity_data['state'].nunique()}"
    )
    print(f"Number of electrolyzer types: {capacity_data['carrier'].nunique()}")

    print("\nTOP 10 STATES BY HYDROGEN CAPACITY")
    state_totals = (
        capacity_data.groupby("state")["p_nom"].sum().sort_values(ascending=False)
    )
    for i, (state, capacity) in enumerate(state_totals.head(10).items()):
        print(f"{i + 1:2d}. {state}: {capacity:.1f} MW")

    print("\nELECTROLYZER TYPE MIX (NATIONAL)")
    carrier_totals = (
        capacity_data.groupby("carrier")["p_nom"].sum().sort_values(ascending=False)
    )
    total_national = carrier_totals.sum()
    for carrier, capacity in carrier_totals.items():
        print(
            f"{carrier:25s}: {capacity:8.1f} MW ({capacity / total_national * 100:5.1f}%)"
        )


def create_ft_capacity_by_state_map(
    network,
    path_shapes,
    network_name="Network",
    distance_crs=4326,
    min_capacity_gw=0.1,
    year_title=True,
):
    """
    Create a geographic map showing Fischer–Tropsch (FT) capacity per state in GW input H2.
    The plot shows one bubble per state (aggregated), but the function returns the link-level table.
    CO₂ pipelines are drawn in gray with legend (10–50 MW).
    """

    # Extract year from network name
    year_match = re.search(r"\d{4}", network_name)
    year_str = f"{year_match.group()}" if year_match else ""

    # Filter FT links
    ft_links = network.links[
        network.links["carrier"].str.contains(
            "FT|Fischer|Tropsch", case=False, na=False
        )
        | network.links.index.str.contains("FT|Fischer|Tropsch", case=False, na=False)
    ].copy()
    if ft_links.empty:
        print(f"No FT links found in the network: {network_name}")
        return None, None, None

    # Merge link data with bus coordinates and state info
    links_with_state = ft_links.merge(
        network.buses[["state", "x", "y"]],
        left_on="bus0",
        right_index=True,
        how="left",
    )
    links_with_state["p_nom_gw"] = links_with_state["p_nom_opt"] / 1000

    # Aggregate to state-level capacity for plotting only
    state_capacity = (
        links_with_state.groupby("state", as_index=False)
        .agg({"x": "mean", "y": "mean", "p_nom_gw": "sum"})
        .rename(columns={"p_nom_gw": "total_gw"})
    )
    state_capacity = state_capacity[state_capacity["total_gw"] >= min_capacity_gw]

    # Load shapes and plot base map
    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes["ISO_1"] = shapes["ISO_1"].apply(lambda x: x.split("-")[1])
    shapes.rename(columns={"ISO_1": "State"}, inplace=True)

    fig, ax = plt.subplots(
        figsize=(12, 6), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    bbox = box(-130, 20, -65, 50)
    shapes_clip = shapes.to_crs(epsg=4326).clip(bbox)
    shapes_clip.plot(
        ax=ax, facecolor="whitesmoke", edgecolor="gray", alpha=0.7, linewidth=0.5
    )

    lon_min, lon_max = -130, -65
    lat_min, lat_max = 20, 50

    # --- CO2 PIPELINES -------------------------------------------------------
    co2_links = network.links[
        network.links["carrier"].str.lower() == "co2 pipeline"
    ].copy()
    if not co2_links.empty:
        line_scale = 30
        for _, link in co2_links.iterrows():
            bus0 = network.buses.loc[link.bus0, ["x", "y"]]
            bus1 = network.buses.loc[link.bus1, ["x", "y"]]
            if (
                lon_min < bus0.x < lon_max
                and lat_min < bus0.y < lat_max
                and lon_min < bus1.x < lon_max
                and lat_min < bus1.y < lat_max
            ):
                ax.plot(
                    [bus0.x, bus1.x],
                    [bus0.y, bus1.y],
                    color="dimgray",
                    linewidth=link.p_nom_opt / line_scale,
                    alpha=0.7,
                    transform=ccrs.PlateCarree(),
                    zorder=3,
                )

        legend_caps_MW = [10, 50, 100]
        legend_line_scale_factor = 1
        co2_legend_lines = [
            mlines.Line2D(
                [],
                [],
                color="dimgray",
                linewidth=(cap / line_scale) * legend_line_scale_factor,
                alpha=0.7,
                label=f"{cap} MW",
            )
            for cap in legend_caps_MW
        ]
        co2_legend = ax.legend(
            handles=co2_legend_lines,
            title="CO2 Pipelines",
            title_fontproperties=FontProperties(weight="bold"),
            fontsize=9,
            loc="upper left",
            bbox_to_anchor=(1.05, 1),
            frameon=False,
            labelspacing=1.0,
        )
        ax.add_artist(co2_legend)

    # --- Plot one circle per state ---
    pie_scale = 0.2
    min_radius = 0.1
    max_radius = 3.5

    for _, row in state_capacity.iterrows():
        x, y, total = row["x"], row["y"], row["total_gw"]
        if pd.isna(x) or pd.isna(y):
            continue
        if not (lon_min < x < lon_max and lat_min < y < lat_max):
            continue

        radius = np.clip(total * pie_scale, min_radius, max_radius)
        circle = plt.Circle(
            (x, y),
            radius,
            color="#B22222",
            alpha=0.6,
            transform=ccrs.PlateCarree(),
            zorder=4,
            linewidth=1,
        )
        ax.add_patch(circle)
        ax.text(
            x,
            y - radius - 0.3,
            f"{total:.2f} GW",
            ha="center",
            va="top",
            fontsize=9,
            bbox=dict(facecolor="white", edgecolor="gray", boxstyle="round,pad=0.2"),
            transform=ccrs.PlateCarree(),
        )

    ax.set_extent([lon_min, lon_max, lat_min, lat_max], crs=ccrs.PlateCarree())
    ax.axis("off")
    ax.set_title(
        f"Fischer-Tropsch Capacity by State (GW input H2) - {year_str if year_title else network_name}",
        fontsize=12,
    )
    plt.tight_layout()

    return fig, ax, links_with_state


def create_ft_capacity_by_grid_region_map(
    network,
    path_shapes,
    network_name="Network",
    distance_crs=4326,
    min_capacity_gw=0.1,
    year_title=True,
):
    """
    Create a map showing total FT capacity per grid region in GW input H2,
    aggregating multiple FT links on the same node (bus0) into a single bubble.
    CO2 pipelines are drawn in gray lines with legend (10–50 MW range).
    """

    year_match = re.search(r"\d{4}", network_name)
    year_str = f"{year_match.group()}" if year_match else ""

    ft_links = network.links[
        network.links["carrier"].str.contains(
            "FT|Fischer|Tropsch", case=False, na=False
        )
        | network.links.index.str.contains("FT|Fischer|Tropsch", case=False, na=False)
    ].copy()
    if ft_links.empty:
        print(f"No FT links found in network {network_name}")
        return None, None, None

    links = ft_links.merge(
        network.buses[["grid_region", "x", "y"]],
        left_on="bus0",
        right_index=True,
        how="left",
    )
    links["p_nom_gw"] = links["p_nom_opt"] / 1000

    grid_capacity = (
        links.groupby("bus0", as_index=False)
        .agg({"grid_region": "first", "x": "first", "y": "first", "p_nom_gw": "sum"})
        .rename(columns={"p_nom_gw": "total_gw"})
    )
    grid_capacity = grid_capacity[grid_capacity["total_gw"] >= min_capacity_gw]

    fig, ax = plt.subplots(
        figsize=(12, 6), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    lon_min, lon_max = -130, -65
    lat_min, lat_max = 20, 50
    ax.set_extent([lon_min, lon_max, lat_min, lat_max], crs=ccrs.PlateCarree())

    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes = shapes.to_crs(epsg=4326).clip(box(lon_min, lat_min, lon_max, lat_max))
    shapes.plot(
        ax=ax, facecolor="whitesmoke", edgecolor="gray", alpha=0.7, linewidth=0.5
    )

    # CO2 pipelines
    co2_links = network.links[
        network.links["carrier"].str.lower() == "co2 pipeline"
    ].copy()
    if not co2_links.empty:
        line_scale = 30
        for _, link in co2_links.iterrows():
            bus0 = network.buses.loc[link.bus0, ["x", "y"]]
            bus1 = network.buses.loc[link.bus1, ["x", "y"]]
            if (
                lon_min < bus0.x < lon_max
                and lat_min < bus0.y < lat_max
                and lon_min < bus1.x < lon_max
                and lat_min < bus1.y < lat_max
            ):
                ax.plot(
                    [bus0.x, bus1.x],
                    [bus0.y, bus1.y],
                    color="dimgray",
                    linewidth=link.p_nom_opt / line_scale,
                    alpha=0.7,
                    transform=ccrs.PlateCarree(),
                    zorder=3,
                )

        legend_caps_MW = [10, 50, 100]
        legend_line_scale_factor = 1
        co2_legend_lines = [
            mlines.Line2D(
                [],
                [],
                color="dimgray",
                linewidth=(cap / line_scale) * legend_line_scale_factor,
                alpha=0.7,
                label=f"{cap} MW",
            )
            for cap in legend_caps_MW
        ]
        co2_legend = ax.legend(
            handles=co2_legend_lines,
            title="CO2 Pipelines",
            title_fontproperties=FontProperties(weight="bold"),
            fontsize=9,
            loc="upper left",
            bbox_to_anchor=(1.05, 1),
            frameon=False,
            labelspacing=1.0,
        )
        ax.add_artist(co2_legend)

    pie_scale = 0.2
    min_radius = 0.1
    max_radius = 3.5

    for _, row in grid_capacity.iterrows():
        x, y, total = row["x"], row["y"], row["total_gw"]
        if not (lon_min < x < lon_max and lat_min < y < lat_max):
            continue
        radius = np.clip(total * pie_scale, min_radius, max_radius)  # lineare
        circle = plt.Circle(
            (x, y),
            radius,
            facecolor="#B22222",
            edgecolor="gray",
            alpha=0.6,
            linewidth=1,
            transform=ccrs.PlateCarree(),
            clip_on=True,
            zorder=4,
        )
        ax.add_patch(circle)
        ax.text(
            x,
            y - radius - 0.3,
            f"{total:.2f} GW",
            ha="center",
            va="top",
            fontsize=9,
            bbox=dict(facecolor="white", edgecolor="gray", boxstyle="round,pad=0.2"),
            transform=ccrs.PlateCarree(),
        )

    ax.set_title(
        f"Fischer-Tropsch Capacity by Grid Region (GW input H2) - {year_str if year_title else network_name}",
        fontsize=12,
        pad=20,
    )
    ax.axis("off")
    plt.tight_layout()

    return fig, ax, grid_capacity


def filter_and_group_small_carriers(df, threshold=0.005):
    """
    Filters a DataFrame to group small contributors into an 'other' category.
    This function assumes df contains only non-negative values.
    """
    if df.empty or df.sum().sum() == 0:
        return pd.DataFrame(index=df.index)
    totals = df.sum()
    grand_total = totals.sum()
    significant_carriers = totals[totals / grand_total > threshold].index
    df_filtered = df[significant_carriers].copy()
    other_carriers = totals[~totals.index.isin(significant_carriers)].index
    if not other_carriers.empty:
        df_filtered["other"] = df[other_carriers].sum(axis=1)
    return df_filtered


def calculate_dispatch(n, start_date=None, end_date=None):
    # Select time window
    snapshots_slice = (
        slice(start_date, end_date) if start_date and end_date else slice(None)
    )
    snapshots = n.snapshots[snapshots_slice]

    timestep_hours = (snapshots[1] - snapshots[0]).total_seconds() / 3600

    gen_and_sto_carriers = {
        "csp",
        "solar",
        "onwind",
        "offwind-dc",
        "offwind-ac",
        "nuclear",
        "geothermal",
        "ror",
        "hydro",
        "solar rooftop",
    }
    link_carriers = ["coal", "oil", "OCGT", "CCGT", "biomass", "biomass CHP", "gas CHP"]

    # identify electric buses
    electric_buses = set(
        n.buses.index[
            ~n.buses.carrier.str.contains("heat|gas|H2|oil|coal", case=False, na=False)
        ]
    )

    # Generators
    gen = n.generators[n.generators.carrier.isin(gen_and_sto_carriers)]
    gen_p = n.generators_t.p.loc[snapshots_slice, gen.index].clip(lower=0)
    gen_dispatch = gen_p.groupby(gen["carrier"], axis=1).sum()

    # Storage units
    sto = n.storage_units[n.storage_units.carrier.isin(gen_and_sto_carriers)]
    sto_p = n.storage_units_t.p.loc[snapshots_slice, sto.index].clip(lower=0)
    sto_dispatch = sto_p.groupby(sto["carrier"], axis=1).sum()

    # Links: conventional generation
    link_frames = []
    for carrier in link_carriers:
        links = n.links[
            (n.links.carrier == carrier) & (n.links.bus1.isin(electric_buses))
        ]
        if links.empty:
            continue
        p1 = n.links_t.p1.loc[snapshots_slice, links.index].clip(upper=0)
        p1_positive = -p1
        df = p1_positive.groupby(links["carrier"], axis=1).sum()
        link_frames.append(df)

    # Battery
    battery_links = n.links[n.links.carrier == "battery discharger"]
    if not battery_links.empty:
        p1 = n.links_t.p1.loc[snapshots_slice, battery_links.index].clip(upper=0)
        battery_dispatch = -p1.groupby(battery_links["carrier"], axis=1).sum()
        battery_dispatch.columns = ["battery discharger"]
        link_frames.append(battery_dispatch)

    link_dispatch = (
        pd.concat(link_frames, axis=1) if link_frames else pd.DataFrame(index=snapshots)
    )

    # Combine everything
    supply = pd.concat([gen_dispatch, sto_dispatch, link_dispatch], axis=1)
    supply = supply.groupby(supply.columns, axis=1).sum().clip(lower=0)

    # Convert
    supply_gw = supply / 1e3
    energy_mwh = supply.sum(axis=1) * timestep_hours
    total_gwh = energy_mwh.sum() / 1e3

    return total_gwh, supply_gw


def plot_electricity_dispatch(
    networks, carrier_colors, start_date=None, end_date=None, ymax=None
):
    summary_list = []
    max_y = 0

    # First pass: calculate ymax (for plot scaling) and total GWh
    for key, n in networks.items():
        print(f"Processing network: {key}")
        total_gwh, supply_gw = calculate_dispatch(n, start_date, end_date)
        summary_list.append({"Network": key, "Total Dispatch (GWh)": total_gwh})
        max_y = max(max_y, supply_gw.sum(axis=1).max())

    # Use provided ymax or max across all networks
    y_max_plot = ymax if ymax is not None else max_y

    # Create one subplot per network
    fig, axes = plt.subplots(
        len(networks), 1, figsize=(22, 5 * len(networks)), sharex=True
    )
    if len(networks) == 1:
        axes = [axes]

    # Order of technologies in the stacked plot
    ordered_columns = [
        "nuclear",
        "coal",
        "biomass",
        "biomass CHP",
        "gas CHP",
        "CCGT",
        "OCGT",
        "oil",
        "hydro",
        "ror",
        "geothermal",
        "solar",
        "solar rooftop",
        "csp",
        "onwind",
        "offwind-ac",
        "offwind-dc",
        "battery discharger",
    ]

    # Loop through networks and plot each one
    for ax, (key, n) in zip(axes, networks.items()):
        _, supply_gw = calculate_dispatch(n, start_date, end_date)

        # Convert index to datetime and resample daily averages
        supply_gw.index = pd.to_datetime(supply_gw.index)
        supply_gw = supply_gw.resample("24H").mean()

        # Keep only carriers present in both ordered_columns and supply_gw
        supply_gw = supply_gw[[c for c in ordered_columns if c in supply_gw.columns]]

        # Stacked area plot
        supply_gw.plot.area(
            ax=ax,
            stacked=True,
            linewidth=0,
            color=[carrier_colors.get(c, "gray") for c in supply_gw.columns],
            legend=False,
        )

        # Title and axes formatting
        ax.set_title(f"Electricity dispatch – {key}")
        ax.set_ylabel("Power (GW)")
        ax.set_ylim(0, y_max_plot)
        ax.grid(True)

        # Legend: only keep carriers with nonzero total generation
        handles, labels = ax.get_legend_handles_labels()
        sums = supply_gw.sum()
        filtered = [(h, l) for h, l in zip(handles, labels) if sums.get(l, 0) > 0]

        if filtered:
            handles, labels = zip(*filtered)
            ax.legend(
                handles,
                labels,
                loc="center left",
                bbox_to_anchor=(1.02, 0.5),
                title="Technology",
                fontsize="small",
                title_fontsize="medium",
            )

    # Label x-axis for the bottom plot
    axes[-1].set_xlabel("Time")
    plt.tight_layout(rect=[0, 0, 0.80, 1])
    showfig()

    return summary_list


def compute_and_plot_load(n, key="", ymax=None, start_date=None, end_date=None):
    freq = pd.infer_freq(n.loads_t.p_set.index)
    snapshots = n.snapshots
    snapshot_hours = (snapshots[1] - snapshots[0]).total_seconds() / 3600

    dynamic_load_gw = n.loads_t.p_set.sum(axis=1) / 1e3
    total_dynamic_gwh = (n.loads_t.p_set.sum(axis=1) * snapshot_hours).sum() / 1e3

    static_loads = n.loads[~n.loads.index.isin(n.loads_t.p_set.columns)]
    static_load_gw = static_loads["p_set"].sum() / 1e3
    total_hours = len(n.loads_t.p_set.index) * snapshot_hours
    total_static_gwh = static_load_gw * total_hours

    total_dispatch_gwh, _ = calculate_dispatch(n, start_date, end_date)

    # Plot electric load
    fig, ax = plt.subplots(figsize=(14, 5))
    ax.plot(dynamic_load_gw.index, dynamic_load_gw.values, label="Dynamic Load (GW)")
    ax.hlines(
        static_load_gw,
        dynamic_load_gw.index.min(),
        dynamic_load_gw.index.max(),
        colors="red",
        linestyles="--",
        label="Static Load (GW)",
    )

    start = dynamic_load_gw.index.min().replace(day=1)
    end = dynamic_load_gw.index.max()
    month_starts = pd.date_range(start=start, end=end, freq="MS")

    ax.set_xlim(start, end)
    ax.set_xticks(month_starts)
    ax.set_xticklabels(month_starts.strftime("%b"))
    ax.tick_params(axis="x", rotation=0)

    if ymax:
        ax.set_ylim(0, ymax)
    else:
        ax.set_ylim(bottom=0)

    ax.set_title(f"Electric Load Profile - {key}")
    ax.set_xlabel("Time")
    ax.set_ylabel("Load (GW)")
    ax.legend()
    ax.grid(True)

    plt.tight_layout(rect=[0, 0, 0.88, 1])
    showfig()

    return {
        "key": key,
        "mean_dynamic_load_gw": dynamic_load_gw.mean(),
        "total_dynamic_gwh": total_dynamic_gwh,
        "static_load_gw": static_load_gw,
        "total_static_gwh": total_static_gwh,
        "total_dispatch_gwh": total_dispatch_gwh,
    }


def calculate_lcoe_summary_and_map(n, shapes):
    snapshot_weights = n.snapshot_weightings.generators

    gen_carriers = {
        "csp",
        "solar",
        "onwind",
        "offwind-dc",
        "offwind-ac",
        "nuclear",
        "geothermal",
        "ror",
        "hydro",
        "solar rooftop",
    }

    storage_carriers = {"battery storage", "hydro", "PHS"}

    link_carriers = [
        "coal",
        "oil",
        "OCGT",
        "CCGT",
        "biomass",
        "lignite",
        "urban central solid biomass CHP",
        "urban central gas CHP",
    ]

    electric_buses = set(n.buses[n.buses.carrier == "AC"].index)

    # --- Generators ---
    gen = n.generators[n.generators.carrier.isin(gen_carriers)].copy()
    gen_dispatch = n.generators_t.p[gen.index].multiply(snapshot_weights, axis=0)
    gen["energy"] = gen_dispatch.sum()
    gen = gen[(gen.p_nom_opt > 0) & (gen.energy > 0)]
    gen["lcoe"] = (
        gen.capital_cost * gen.p_nom_opt + gen.marginal_cost * gen.energy
    ) / gen.energy
    gen["type"] = "generator"

    # --- Storage units ---
    sto = n.storage_units[n.storage_units.carrier.isin(storage_carriers)].copy()
    sto_dispatch = (
        n.storage_units_t.p[sto.index].clip(lower=0).multiply(snapshot_weights, axis=0)
    )
    sto["energy"] = sto_dispatch.sum()
    sto = sto[(sto.p_nom_opt > 0) & (sto.energy > 0)]
    sto["lcoe"] = (
        sto.capital_cost * sto.p_nom_opt + sto.marginal_cost * sto.energy
    ) / sto.energy
    sto["type"] = "storage"

    # --- Links ---
    link = n.links[
        (n.links.carrier.isin(link_carriers))
        & (n.links.bus1.isin(electric_buses))
        & (n.links.p_nom_opt > 0)
    ].copy()

    link_dispatch = -n.links_t.p1[link.index].clip(upper=0)
    weighted_link_dispatch = link_dispatch.multiply(snapshot_weights, axis=0)
    link["energy"] = weighted_link_dispatch.sum()

    fuel_usage = n.links_t.p0[link.index].clip(lower=0)
    weighted_fuel_usage = fuel_usage.multiply(snapshot_weights, axis=0)
    link["fuel_usage"] = weighted_fuel_usage.sum()
    link["fuel_cost"] = link.bus0.map(n.generators.marginal_cost)

    # capacity factor
    H = float(snapshot_weights.sum())
    link["CF"] = link["energy"] / (link["p_nom_opt"] * H)

    def lcoe_link(row):
        if row["energy"] <= 0:
            return np.nan
        if row["carrier"] == "oil":
            return np.nan
        # filtro: se CF < 5% → NaN
        if row["CF"] < 0.05:
            return np.nan
        return (
            row["capital_cost"] * row["p_nom_opt"]
            + row["marginal_cost"] * row["fuel_usage"]
            + row["fuel_cost"] * row["fuel_usage"]
        ) / row["energy"]

    link["lcoe"] = link.apply(lcoe_link, axis=1)
    link["type"] = "link"

    # --- Merge data ---
    gen_data = gen[["bus", "carrier", "lcoe", "type", "energy"]]
    sto_data = sto[["bus", "carrier", "lcoe", "type", "energy"]]
    link_data = link[["bus1", "carrier", "lcoe", "type", "energy"]].rename(
        columns={"bus1": "bus"}
    )

    lcoe_data = pd.concat([gen_data, sto_data, link_data], axis=0).dropna()
    lcoe_data = lcoe_data.merge(
        n.buses[["x", "y", "grid_region"]], left_on="bus", right_index=True
    )

    lcoe_by_bus = (
        lcoe_data.groupby("bus")
        .apply(
            lambda df: pd.Series(
                {
                    "weighted_lcoe": (df["lcoe"] * df["energy"]).sum()
                    / df["energy"].sum(),
                    "x": df["x"].iloc[0],
                    "y": df["y"].iloc[0],
                    "grid_region": df["grid_region"].iloc[0],
                }
            )
        )
        .reset_index()
    )

    region_summary = (
        lcoe_data.groupby(["grid_region", "carrier"])
        .agg(
            dispatch_mwh=("energy", "sum"),
            total_cost=("lcoe", lambda x: (x * lcoe_data.loc[x.index, "energy"]).sum()),
        )
        .reset_index()
    )
    region_summary["lcoe"] = (
        region_summary["total_cost"] / region_summary["dispatch_mwh"]
    )
    region_summary["dispatch"] = region_summary["dispatch_mwh"] / 1e6

    table = region_summary.pivot(
        index="grid_region", columns="carrier", values=["lcoe", "dispatch"]
    )
    table.columns = [
        f"{carrier} {metric} ({'USD/MWh' if metric == 'lcoe' else 'TWh'})"
        for metric, carrier in table.columns
    ]
    table = table.reset_index()

    # filtra dispatch e sostituisci lcoe per dispatch bassi
    dispatch_cols = [col for col in table.columns if "dispatch" in col.lower()]
    for col in dispatch_cols:
        table[col] = pd.to_numeric(table[col], errors="coerce").fillna(0.0)

    lcoe_cols = [col for col in table.columns if "lcoe" in col.lower()]

    min_dispatch_threshold = 1  # TWh
    for lcoe_col in lcoe_cols:
        carrier = lcoe_col.split(" ")[0]
        dispatch_col = next(
            (col for col in dispatch_cols if col.startswith(carrier + " ")), None
        )
        if dispatch_col:
            mask = table[dispatch_col] < min_dispatch_threshold
            table.loc[mask, lcoe_col] = np.nan

    table[lcoe_cols] = table[lcoe_cols].applymap(
        lambda x: "-" if pd.isna(x) else round(x, 2)
    )

    grid_region_weighted_lcoe = (
        lcoe_by_bus.merge(lcoe_data[["bus", "energy"]], on="bus", how="left")
        .groupby("grid_region")
        .apply(
            lambda df: (df["weighted_lcoe"] * df["energy"]).sum() / df["energy"].sum()
        )
    )
    table["Weighted Average LCOE (USD/MWh)"] = (
        table["grid_region"].map(grid_region_weighted_lcoe).round(2)
    )

    for col in table.columns:
        if col != "grid_region":
            table[col] = (
                table[col].round(2) if table[col].dtype != object else table[col]
            )

    vmin = lcoe_by_bus["weighted_lcoe"].quantile(0.05)
    vmax = max(
        vmin,
        min(grid_region_weighted_lcoe.max() * 1.1, lcoe_by_bus["weighted_lcoe"].max()),
    )

    geometry = [Point(xy) for xy in zip(lcoe_by_bus["x"], lcoe_by_bus["y"])]
    lcoe_gdf = gpd.GeoDataFrame(lcoe_by_bus, geometry=geometry, crs=shapes.crs).to_crs(
        shapes.crs
    )

    return lcoe_gdf, table, lcoe_by_bus, lcoe_data, vmin, vmax


def plot_lcoe_map_by_grid_region(
    lcoe_by_bus, lcoe_data, shapes, title=None, key=None, ax=None, vmin=None, vmax=None
):
    grid_region_lcoe = (
        lcoe_by_bus.merge(
            lcoe_data[["bus", "energy"]], left_on="bus", right_on="bus", how="left"
        )
        .groupby("grid_region")
        .apply(
            lambda df: (df["weighted_lcoe"] * df["energy"]).sum() / df["energy"].sum()
        )
        .reset_index(name="weighted_lcoe")
    )

    shapes = shapes.rename(columns={"GRID_REGIO": "grid_region"})
    shapes_lcoe = shapes.merge(grid_region_lcoe, on="grid_region", how="left")

    if vmin is None:
        vmin = shapes_lcoe["weighted_lcoe"].quantile(0.05)
    if vmax is None:
        vmax = shapes_lcoe["weighted_lcoe"].quantile(0.95)

    if ax is None:
        fig, ax = plt.subplots(
            figsize=(12, 10), subplot_kw={"projection": ccrs.PlateCarree()}
        )

    shapes_lcoe.plot(
        column="weighted_lcoe",
        cmap=plt.cm.get_cmap("RdYlGn_r"),
        linewidth=0.8,
        edgecolor="0.8",
        legend=True,
        vmin=vmin,
        vmax=vmax,
        ax=ax,
    )

    ax.set_extent([-130, -65, 20, 55], crs=ccrs.PlateCarree())
    ax.axis("off")

    # Title handling
    if title:
        ax.set_title(title)
    else:
        ax.set_title("Weighted average (plant level) LCOE per Grid Region (USD/MWh)")


def plot_h2_capacities_map(network, title, tech_colors, nice_names, regions_onshore):
    h2_carriers_buses = ["Alkaline electrolyzer large", "PEM electrolyzer", "SOEC"]

    net = network.copy()
    h2_capacity_data = compute_h2_capacities(net)[h2_carriers_buses]

    valid_buses = net.buses.dropna(subset=["x", "y"])
    valid_buses = valid_buses[
        (valid_buses["x"] > -200)
        & (valid_buses["x"] < 200)
        & (valid_buses["y"] > -90)
        & (valid_buses["y"] < 90)
    ]

    fig, ax = plt.subplots(
        figsize=(14, 10), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    bbox = box(-130, 20, -60, 50)
    regions_onshore_clipped = regions_onshore.to_crs(epsg=4326).clip(bbox)
    regions_onshore_clipped.plot(
        ax=ax,
        facecolor="whitesmoke",
        edgecolor="gray",
        alpha=0.7,
        linewidth=0.5,
        zorder=0,
    )

    max_cap = h2_capacity_data.sum(axis=1).max()

    for bus_id, capacities in h2_capacity_data.iterrows():
        if bus_id not in valid_buses.index:
            continue
        x, y = valid_buses.loc[bus_id, ["x", "y"]]
        if not bbox.contains(gpd.points_from_xy([x], [y])[0]):
            continue

        total = capacities.sum()
        if total < 10:
            continue

        radius = np.clip(np.sqrt(total) * 0.02, 0.3, 2.0)
        colors = [tech_colors.get(c, "gray") for c in capacities.index]

        start_angle = 0
        for val, color in zip(capacities.values, colors):
            if val == 0:
                continue
            angle = 360 * val / total
            wedge = Wedge(
                center=(x, y),
                r=radius,
                theta1=start_angle,
                theta2=start_angle + angle,
                facecolor=color,
                transform=ccrs.PlateCarree()._as_mpl_transform(ax),
                zorder=5,
            )
            ax.add_patch(wedge)
            start_angle += angle

        ax.text(
            x,
            y + radius + 0.3,
            f"{total:.1e} MW",
            transform=ccrs.PlateCarree(),
            fontsize=8,
            ha="center",
            va="bottom",
            zorder=6,
            bbox=dict(
                facecolor="white", edgecolor="gray", boxstyle="round,pad=0.2", alpha=0.7
            ),
        )

    # Legends
    legend_anchor_x = 1.05
    bold_fp = FontProperties(weight="bold", size=10)

    # Electrolyzer Capacity Legend
    legend_caps = [1e1, 1e2, 1e3]
    legend_elements = [
        plt.Line2D(
            [0],
            [0],
            marker="o",
            color="gray",
            markersize=np.clip(np.sqrt(cap) * 0.02, 0.3, 2.0) * 20,
            alpha=0.4,
            linestyle="None",
            label=f"{cap:.0e} MW",
        )
        for cap in legend_caps
    ]
    cap_legend = ax.legend(
        handles=legend_elements,
        title="Electrolyzer Capacity",
        title_fontproperties=bold_fp,
        fontsize=9,
        loc="upper left",
        bbox_to_anchor=(legend_anchor_x, 1),
        frameon=False,
        labelspacing=2.2,
        handletextpad=1.0,
    )

    # Technology legend
    carrier_handles = [
        mpatches.Patch(color=tech_colors.get(c, "gray"), label=nice_names.get(c, c))
        for c in sorted(h2_capacity_data.columns)
        if h2_capacity_data[c].sum() > 0
    ]
    tech_legend = ax.legend(
        handles=carrier_handles,
        title="Electrolyzer technologies",
        title_fontproperties=FontProperties(weight="bold"),
        fontsize=9,
        loc="upper left",
        bbox_to_anchor=(legend_anchor_x, 0.60),
        frameon=False,
        labelspacing=1.0,
    )

    tech_legend._legend_title_box._text.set_ha("left")

    ax.add_artist(cap_legend)
    ax.add_artist(tech_legend)

    ax.set_extent([-130, -65, 20, 55], crs=ccrs.PlateCarree())

    ax.set_title(
        f"Installed electrolyzer capacity (MW input electricity) - {title} (only nodes ≥ 10 MW)\n"
    )
    plt.tight_layout()
    showfig()


def plot_lcoh_maps_by_grid_region_marginal(
    networks,
    shapes,
    h2_carriers,
    regional_fees,
    emm_mapping,
    output_threshold=1.0,
    include_baseload=True,
    baseload_charge_path="./data/energy_charge_rate.csv",
    customer_charge_mw=400.0,
    demand_charge_rate=9.0,
    baseload_percentages=None,
    year_title=True,
):
    """
    Plot weighted average LCOH by grid region (USD/kg H2),
    using marginal electricity prices and fully coherent
    with calculate_lcoh_by_region.

    LCOH =
        Electrolysis CAPEX
      + Electrolysis OPEX
      + Electricity (net of baseload)
      + Baseload charges
      + Transmission fees
    """

    # -----------------------------
    # Conversion (LHV)
    # -----------------------------
    conv = 1000.0 / 33.0  # kg H2 per MWh H2

    # -----------------------------
    # Normalize grid_region column
    # -----------------------------
    for col in ["grid_region", "Grid Region", "GRID_REGIO"]:
        if col in shapes.columns:
            shapes = shapes.rename(columns={col: "grid_region"})
            break
    else:
        raise KeyError("No grid_region column found in shapes")

    # -----------------------------
    # Baseload charges (optional)
    # -----------------------------
    baseload_charges = {}
    if include_baseload:
        baseload_charges = calculate_baseload_charge(
            networks=networks,
            h2_carriers=h2_carriers,
            emm_mapping=emm_mapping,
            energy_charge_path=baseload_charge_path,
            customer_charge_mw=customer_charge_mw,
            demand_charge_rate=demand_charge_rate,
            baseload_percentages=baseload_percentages,
            output_threshold=output_threshold,
            verbose=False,
            year_title=year_title,
        )

    all_results = []

    # -----------------------------
    # Loop over networks
    # -----------------------------
    for year_key, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(year_key)).group())

        # Exclude Base_2023
        if scen_year == 2023:
            continue

        key = scen_year if year_title else year_key

        links = net.links[net.links.carrier.isin(h2_carriers)]
        if links.empty:
            continue

        # Flows
        p0 = net.links_t.p0[links.index]  # electricity input (MW)
        p1 = net.links_t.p1[links.index]  # H2 output (negative)
        w = net.snapshot_weightings.generators

        cons = p0.clip(lower=0).multiply(w, axis=0)  # MWh_el
        h2 = (-p1).clip(lower=0).multiply(w, axis=0)  # MWh_H2
        h2_out = h2.sum()

        valid = h2_out > output_threshold
        if valid.sum() == 0:
            continue

        out_valid = h2_out[valid]

        # Skip entire year if total H2 production is below threshold (align with table)
        if h2_out[valid].sum() <= output_threshold:
            continue

        # -----------------------------
        # CAPEX / OPEX (electrolyzers)
        # -----------------------------
        capex = links.loc[valid, "capital_cost"] * links.loc[valid, "p_nom_opt"]
        opex = links.loc[valid, "marginal_cost"] * cons.loc[:, valid].sum(axis=0)

        capex_val = capex / out_valid / conv
        opex_val = opex / out_valid / conv

        # -----------------------------
        # Electricity cost (marginal)
        # -----------------------------
        elec_cost = {}
        for l in valid.index[valid]:
            bus = links.at[l, "bus0"]
            elec_cost[l] = (cons[l] * net.buses_t.marginal_price[bus]).sum()

        elec_val = pd.Series(elec_cost) / out_valid / conv  # USD/kg H2

        # -----------------------------
        # Base dataframe
        # -----------------------------
        df = pd.DataFrame(
            {
                "Electrolysis CAPEX (USD/kg H2)": capex_val,
                "Electrolysis OPEX (USD/kg H2)": opex_val,
                "Electricity (USD/kg H2)": elec_val,
                "h2_out": out_valid,
                "bus": links.loc[valid, "bus0"],
            }
        )

        df["grid_region"] = df["bus"].map(net.buses["grid_region"])

        # -----------------------------
        # Transmission fees
        # -----------------------------
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year, ["region", "Transmission nom USD/MWh"]
        ].set_index("region")

        df["EMM"] = df["grid_region"].map(emm_mapping)
        fee_trans = df["EMM"].map(fee_map["Transmission nom USD/MWh"]).fillna(0.0)

        elec_rate = cons.loc[:, valid].sum(axis=0) / out_valid  # MWh_el / MWh_H2

        df["Transmission (USD/kg H2)"] = (
            (fee_trans * elec_rate / conv).reindex(df.index).fillna(0.0)
        )

        # -----------------------------
        # Baseload accounting
        # -----------------------------
        if include_baseload and key in baseload_charges:
            baseload_df = baseload_charges[key]

            baseload_frac = baseload_df.set_index("grid_region")["baseload_pct"].div(
                100.0
            )

            baseload_cost = baseload_df.set_index("grid_region")[
                "baseload_cost_per_mwh_h2"
            ]

            # Activity factor (same as calculate_lcoh_by_region)
            activity_factor = {}
            for l in valid.index[valid]:
                p_prod = (-net.links_t.p1[l]).clip(lower=0)
                activity_factor[l] = (p_prod > 0).sum() / len(p_prod)

            df["activity_factor"] = df.index.map(activity_factor)
            df["baseload_frac"] = df["grid_region"].map(baseload_frac).fillna(0.0)

            df["Electricity net of baseload (USD/kg H2)"] = df[
                "Electricity (USD/kg H2)"
            ] * (1 - df["baseload_frac"] * df["activity_factor"])

            df["Baseload charges (USD/kg H2)"] = (
                df["grid_region"].map(baseload_cost).fillna(0.0) / conv
            )

        else:
            df["Electricity net of baseload (USD/kg H2)"] = df[
                "Electricity (USD/kg H2)"
            ]
            df["Baseload charges (USD/kg H2)"] = 0.0

        # -----------------------------
        # Final LCOH
        # -----------------------------
        df["LCOH incl. Transmission + Baseload"] = (
            df["Electrolysis CAPEX (USD/kg H2)"]
            + df["Electrolysis OPEX (USD/kg H2)"]
            + df["Electricity net of baseload (USD/kg H2)"]
            + df["Baseload charges (USD/kg H2)"]
            + df["Transmission (USD/kg H2)"]
        )

        df["year"] = key

        all_results.append(
            df[
                [
                    "grid_region",
                    "year",
                    "h2_out",
                    "LCOH incl. Transmission + Baseload",
                ]
            ]
        )

    if not all_results:
        print("No valid data.")
        return

    # -----------------------------
    # Aggregate by region
    # -----------------------------
    all_df = pd.concat(all_results, ignore_index=True)

    region_lcoh = (
        all_df.groupby(["grid_region", "year"])
        .apply(
            lambda g: pd.Series(
                {
                    "weighted_lcoh": (
                        g["LCOH incl. Transmission + Baseload"] * g["h2_out"]
                    ).sum()
                    / g["h2_out"].sum()
                }
            )
        )
        .reset_index()
    )

    plot_df = shapes.merge(region_lcoh, on="grid_region", how="left")

    vmin = plot_df["weighted_lcoh"].quantile(0.05)
    vmax = plot_df["weighted_lcoh"].quantile(0.95)

    # -----------------------------
    # Plot
    # -----------------------------
    for y in sorted(region_lcoh.year.unique()):
        fig, ax = plt.subplots(
            figsize=(12, 10),
            subplot_kw={"projection": ccrs.PlateCarree()},
        )

        year_df = plot_df[plot_df.year == y]

        year_df.plot(
            column="weighted_lcoh",
            cmap="RdYlGn_r",
            linewidth=0.8,
            edgecolor="0.8",
            legend=True,
            legend_kwds={"label": "LCOH incl. Transmission + Baseload (USD/kg H2)"},
            vmin=vmin,
            vmax=vmax,
            ax=ax,
        )

        ax.set_extent([-130, -65, 20, 55])
        ax.axis("off")
        ax.set_title(
            f"LCOH (incl. Transmission fees & Baseload charges, elec.cost = marginal) – {y}"
        )

        showfig()


def plot_lcoh_maps_by_grid_region_lcoe(
    networks,
    shapes,
    h2_carriers,
    regional_fees,
    emm_mapping,
    grid_region_lcoe,
    output_threshold=1.0,
    include_baseload=True,
    baseload_charge_path="./data/energy_charge_rate.csv",
    customer_charge_mw=400.0,
    demand_charge_rate=9.0,
    baseload_percentages=None,
    year_title=True,
    vmin=None,
    vmax=None,
):
    """
    Plot production-weighted regional LCOH (USD/kg H2), LCOE-based electricity cost.

    EXACT definition:
    LCOH = CAPEX + OPEX + Electricity (regional LCOE)
           + Transmission fees + Baseload charges

    NO Distribution costs included.
    Fully aligned with:
    'LCOH + Transmission fees + Baseload charges (USD/kg H2)' in the table.
    """

    # -----------------------------
    # Constants
    # -----------------------------
    conv = 1000.0 / 33.0  # kg H2 per MWh H2

    # -----------------------------
    # Normalize grid_region column
    # -----------------------------
    for col in ["grid_region", "Grid Region", "GRID_REGIO"]:
        if col in shapes.columns:
            shapes = shapes.rename(columns={col: "grid_region"})
            break
    else:
        raise KeyError("No grid_region column found in shapes")

    # -----------------------------
    # Baseload charges (optional)
    # -----------------------------
    baseload_charges = {}
    if include_baseload:
        baseload_charges = calculate_baseload_charge(
            networks=networks,
            h2_carriers=h2_carriers,
            emm_mapping=emm_mapping,
            energy_charge_path=baseload_charge_path,
            customer_charge_mw=customer_charge_mw,
            demand_charge_rate=demand_charge_rate,
            baseload_percentages=baseload_percentages,
            output_threshold=output_threshold,
            verbose=False,
            year_title=year_title,
        )

    rows = []

    # -----------------------------
    # Loop over networks
    # -----------------------------
    for year_key, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(year_key)).group())

        # Exclude Base_2023
        if scen_year == 2023:
            continue

        year_lbl = scen_year if year_title else year_key

        links = net.links[net.links.carrier.isin(h2_carriers)]
        if links.empty:
            continue

        # Flows
        p0 = net.links_t.p0[links.index]  # MW electricity
        p1 = net.links_t.p1[links.index]  # MW H2 (negative)
        w = net.snapshot_weightings.generators

        cons = p0.clip(lower=0).multiply(w, axis=0)  # MWh_el
        h2 = (-p1).clip(lower=0).multiply(w, axis=0)  # MWh_H2
        h2_out = h2.sum()

        valid = h2_out > output_threshold
        if valid.sum() == 0:
            continue

        out_valid = h2_out[valid]

        # Skip entire year if total H2 production is below threshold
        if out_valid.sum() <= output_threshold:
            continue

        # -----------------------------
        # CAPEX / OPEX
        # -----------------------------
        capex = links.loc[valid, "capital_cost"] * links.loc[valid, "p_nom_opt"]
        opex = links.loc[valid, "marginal_cost"] * cons.loc[:, valid].sum(axis=0)

        capex_val = capex / out_valid / conv
        opex_val = opex / out_valid / conv

        # -----------------------------
        # Electricity cost (regional LCOE)
        # -----------------------------
        elec_val = {}
        for l in out_valid.index:
            bus = links.at[l, "bus0"]
            region = net.buses.at[bus, "grid_region"]
            lcoe = grid_region_lcoe.get(region, None)
            if lcoe is None:
                elec_val[l] = float("nan")
            else:
                elec_val[l] = (cons[l].sum() * lcoe) / out_valid[l] / conv

        elec_val = pd.Series(elec_val)

        # -----------------------------
        # Base dataframe (link level)
        # -----------------------------
        df = pd.DataFrame(
            {
                "grid_region": links.loc[valid, "bus0"].map(net.buses["grid_region"]),
                "h2_out": out_valid,
                "CAPEX": capex_val,
                "OPEX": opex_val,
                "Electricity": elec_val,
            }
        )

        # -----------------------------
        # Transmission fees ONLY
        # -----------------------------
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year, ["region", "Transmission nom USD/MWh"]
        ].set_index("region")

        df["EMM"] = df["grid_region"].map(emm_mapping)
        fee_trans = df["EMM"].map(fee_map["Transmission nom USD/MWh"]).fillna(0.0)

        elec_rate = cons.loc[:, valid].sum(axis=0) / out_valid
        df["Transmission"] = fee_trans * elec_rate / conv

        # -----------------------------
        # Baseload charges
        # -----------------------------
        if include_baseload and year_lbl in baseload_charges:
            bl = baseload_charges[year_lbl]
            bl_cost = bl.set_index("grid_region")["baseload_cost_per_mwh_h2"]
            df["Baseload"] = df["grid_region"].map(bl_cost).fillna(0.0) / conv
        else:
            df["Baseload"] = 0.0

        # -----------------------------
        # FULL LCOH (Transmission-only)
        # -----------------------------
        df["LCOH"] = (
            df["CAPEX"]
            + df["OPEX"]
            + df["Electricity"]
            + df["Transmission"]
            + df["Baseload"]
        )

        df["year"] = year_lbl
        rows.append(df[["grid_region", "year", "h2_out", "LCOH"]])

    if not rows:
        print("No valid data.")
        return

    all_df = pd.concat(rows, ignore_index=True)

    # -----------------------------
    # Production-weighted aggregation
    # -----------------------------
    region_lcoh = (
        all_df.groupby(["grid_region", "year"])
        .apply(
            lambda g: pd.Series(
                {"weighted_lcoh": (g["LCOH"] * g["h2_out"]).sum() / g["h2_out"].sum()}
            )
        )
        .reset_index()
    )

    plot_df = shapes.merge(region_lcoh, on="grid_region", how="left")

    if vmin is None:
        vmin = plot_df["weighted_lcoh"].min()
    if vmax is None:
        vmax = plot_df["weighted_lcoh"].max()

    # -----------------------------
    # Plot
    # -----------------------------
    for y in sorted(region_lcoh.year.unique()):
        fig, ax = plt.subplots(
            figsize=(12, 10), subplot_kw={"projection": ccrs.PlateCarree()}
        )

        plot_df[plot_df.year == y].plot(
            column="weighted_lcoh",
            cmap="RdYlGn_r",
            linewidth=0.8,
            edgecolor="0.8",
            legend=True,
            legend_kwds={"label": "LCOH incl. Transmission + Baseload (USD/kg H2)"},
            vmin=vmin,
            vmax=vmax,
            ax=ax,
        )

        ax.set_extent([-130, -65, 20, 55])
        ax.axis("off")
        ax.set_title(f"LCOH (incl. Transmission & Baseload, elec.cost = LCOE) – {y}")

        showfig()


def calculate_weighted_lcoh_table_by_year(
    networks,
    h2_carriers,
    regional_fees,
    emm_mapping,
    output_threshold=1.0,
    year_title=True,
    include_baseload=True,
    baseload_charge_path="./data/energy_charge_rate.csv",
    customer_charge_mw=400.0,
    demand_charge_rate=9.0,
    baseload_percentages=None,
):
    """
    Calculate weighted average LCOH by grid region and year (USD/kg H2),
    including Transmission fees and Baseload charges.

    Notes
    -----
    - LCOH is computed from CAPEX, OPEX, and electricity cost (marginal),
      including transmission fees and optional baseload charges.
    - Results are weighted by hydrogen output for each grid region.
    """

    results = {}
    for year_key, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(year_key)).group())

        df = calculate_lcoh_by_region(
            {year_key: net},
            h2_carriers=h2_carriers,
            regional_fees=regional_fees,
            emm_mapping=emm_mapping,
            output_threshold=output_threshold,
            year_title=year_title,
            include_baseload=include_baseload,
            baseload_charge_path=baseload_charge_path,
            customer_charge_mw=customer_charge_mw,
            demand_charge_rate=demand_charge_rate,
            baseload_percentages=baseload_percentages,
        )

        results.update(df)

    return results


def calculate_total_generation_by_carrier(network, start_date=None, end_date=None):
    # Time setup
    snapshots_slice = (
        slice(start_date, end_date) if start_date and end_date else slice(None)
    )
    snapshots = network.snapshots[snapshots_slice]
    timestep_h = (snapshots[1] - snapshots[0]).total_seconds() / 3600

    # Define relevant carriers ===
    gen_and_sto_carriers = {
        "csp",
        "solar",
        "onwind",
        "offwind-dc",
        "offwind-ac",
        "nuclear",
        "geothermal",
        "ror",
        "hydro",
        "solar rooftop",
    }
    link_carriers = [
        "coal",
        "oil",
        "OCGT",
        "CCGT",
        "biomass",
        "lignite",
        "urban central solid biomass CHP",
        "urban central gas CHP",
        "battery discharger",
    ]

    # Identify electric buses
    electric_buses = set(
        network.buses.index[
            ~network.buses.carrier.str.contains(
                "heat|gas|H2|oil|coal", case=False, na=False
            )
        ]
    )

    # Generators
    gen = network.generators[network.generators.carrier.isin(gen_and_sto_carriers)]
    gen_p = network.generators_t.p.loc[snapshots_slice, gen.index].clip(lower=0)
    gen_dispatch = gen_p.groupby(gen["carrier"], axis=1).sum()
    gen_energy_mwh = gen_dispatch.sum() * timestep_h

    # Storage units
    sto = network.storage_units[
        network.storage_units.carrier.isin(gen_and_sto_carriers)
    ]
    sto_p = network.storage_units_t.p.loc[snapshots_slice, sto.index].clip(lower=0)
    sto_dispatch = sto_p.groupby(sto["carrier"], axis=1).sum()
    sto_energy_mwh = sto_dispatch.sum() * timestep_h

    # Link-based generation
    link_energy_twh = {}

    for carrier in link_carriers:
        links = network.links[
            (network.links.carrier == carrier)
            & (network.links.bus1.isin(electric_buses))
        ]

        if links.empty:
            link_energy_twh[carrier] = 0.0
            continue

        p1 = network.links_t.p1.loc[snapshots_slice, links.index]
        p1_positive = -p1.clip(upper=0)
        energy_mwh = p1_positive.sum().sum() * timestep_h
        link_energy_twh[carrier] = energy_mwh / 1e6  # MWh → TWh

    link_dispatch = pd.Series(link_energy_twh)

    # Combine all sources
    total_energy_twh = pd.concat(
        [
            gen_energy_mwh / 1e6,  # MW → TWh
            sto_energy_mwh / 1e6,
            link_dispatch,
        ]
    )

    total_energy_twh = total_energy_twh.groupby(total_energy_twh.index).sum()
    total_energy_twh = total_energy_twh[total_energy_twh > 0].round(2)
    total_energy_twh = total_energy_twh.sort_values(ascending=False)

    return total_energy_twh


def plot_hydrogen_dispatch(
    networks, h2_carriers, output_threshold=1.0, year_title=True
):
    """
    Plot hourly hydrogen dispatch per carrier (stacked area plot) for each network in the input dictionary.
    All plots share the same y-axis scale.

    Parameters:
    - networks: dict of PyPSA networks (e.g. {'scenario_2025': network, ...})
    - h2_carriers: list of hydrogen-producing carrier names (e.g. ['PEM', 'SOEC'])
    - output_threshold: minimum total energy (MWh) to include a link in the analysis
    """
    # First pass: find global max dispatch to fix y-axis scale
    global_max = 0
    dispatch_series_by_network = {}

    for key, network in networks.items():
        h2_links = network.links[network.links.carrier.isin(h2_carriers)]
        if h2_links.empty:
            continue

        link_ids = h2_links.index
        p1 = network.links_t.p1[link_ids]
        h2_output = -p1  # Flip sign: PyPSA convention

        data = {}
        for carrier in h2_carriers:
            carrier_links = h2_links[h2_links.carrier == carrier].index
            if carrier_links.empty:
                continue

            output = h2_output[carrier_links]
            output = output.loc[:, output.sum() > output_threshold]
            if output.empty:
                continue

            data[carrier] = output.sum(axis=1)

        if not data:
            continue

        df = pd.DataFrame(data)
        df.index = pd.to_datetime(df.index)
        df = df.resample("24H").mean()
        df = df * 0.03  # Convert to tons
        dispatch_series_by_network[key] = df

        max_dispatch = df.sum(axis=1).max()
        if max_dispatch > global_max:
            global_max = max_dispatch

    if not dispatch_series_by_network:
        print("No valid hydrogen dispatch data found.")
        return pd.DataFrame()

    merged_dispatch_df = pd.concat(dispatch_series_by_network, axis=1)

    # Second pass: generate plots with fixed y-axis
    for key, df in dispatch_series_by_network.items():
        fig, ax = plt.subplots(figsize=(15, 5))
        df.plot.area(ax=ax, linewidth=0)
        year = key[-4:]  # Extract the year
        ax.set_title(f"Electricity Dispatch – {year if year_title else key}")
        ax.set_title(
            f"Hydrogen Dispatch by technology – {year if year_title else key}",
            fontsize=14,
        )
        ax.set_ylabel("Hydrogen Dispatch (tons/hour)")
        ax.set_xlabel("Time")
        ax.set_ylim(0, global_max * 1.05)  # add 5% headroom
        ax.grid(axis="y", linestyle="--", alpha=0.5)

        start = df.index.min().replace(day=1)
        end = df.index.max()
        month_starts = pd.date_range(start=start, end=end, freq="MS")

        ax.set_xlim(start, end)
        ax.set_xticks(month_starts)
        ax.set_xticklabels(month_starts.strftime("%b"))
        ax.tick_params(axis="x", rotation=0)

        ax.legend(
            title="Technology",
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            frameon=False,
        )

        plt.tight_layout(rect=[0, 0, 0.85, 1])
        showfig()

    return merged_dispatch_df


def analyze_ft_costs_by_region(networks: dict, year_title=True):
    """
    Compute and display total Fischer-Tropsch fuel production and
    total marginal cost (USD/MWh) by grid region for each network.
    """
    for name, n in networks.items():
        # Identify Fischer-Tropsch links that are built or extendable with capacity
        ft_links = n.links[
            (n.links.carrier.str.contains("Fischer", case=False, na=False))
            & (
                (n.links.get("p_nom_opt", 0) > 0)
                | (
                    (n.links.get("p_nom", 0) > 0)
                    & (n.links.get("p_nom_extendable", False) == False)
                )
            )
        ].copy()

        if ft_links.empty:
            print(f"\n{name}: No active Fischer-Tropsch links found.")
            continue

        # Filter out links that don't appear in all links_t.p* time series
        ft_link_ids = [
            link
            for link in ft_links.index
            if all(
                link in getattr(n.links_t, attr).columns
                for attr in ["p0", "p1", "p2", "p3"]
            )
        ]

        if not ft_link_ids:
            print(f"\n{name}: No Fischer-Tropsch links with time series data.")
            continue

        # Extract hourly marginal prices for input buses (H2, CO2, electricity)
        price_dict = {}
        for link in ft_link_ids:
            price_dict[link] = {
                "h2_price": n.buses_t.marginal_price[ft_links.at[link, "bus0"]],
                "co2_price": n.buses_t.marginal_price[ft_links.at[link, "bus2"]],
                "elec_price": n.buses_t.marginal_price[ft_links.at[link, "bus3"]],
            }

        # Multiply prices × flows to compute total input cost per link
        p0 = n.links_t.p0[ft_link_ids]
        p2 = n.links_t.p2[ft_link_ids]
        p3 = n.links_t.p3[ft_link_ids]
        p1 = n.links_t.p1[ft_link_ids]

        marginal_cost_inputs = {}
        for link in ft_link_ids:
            cost_h2 = (p0[link] * price_dict[link]["h2_price"]).sum()
            cost_co2 = (p2[link] * price_dict[link]["co2_price"]).sum()
            cost_elec = (p3[link] * price_dict[link]["elec_price"]).sum()
            total_input_cost = cost_h2 + cost_co2 + cost_elec
            marginal_cost_inputs[link] = total_input_cost

        # Compute total marginal cost per MWh of fuel output (input + technical cost)
        marginal_cost_total = {}
        for link in ft_link_ids:
            output_mwh = -p1[link].sum()
            if output_mwh <= 0:
                continue
            tech_cost = ft_links.at[link, "marginal_cost"] * output_mwh
            total_cost_per_mwh = (marginal_cost_inputs[link] + tech_cost) / output_mwh
            marginal_cost_total[link] = {
                "bus": ft_links.at[link, "bus1"],
                "production (MWh)": output_mwh,
                "marginal_cost_total (USD/MWh)": total_cost_per_mwh,
            }

        # Map each link's output bus to its grid_region
        df_links = pd.DataFrame.from_dict(marginal_cost_total, orient="index")
        bus_to_region = n.buses["grid_region"].to_dict()
        df_links["grid_region"] = df_links["bus"].map(bus_to_region)
        df_links = df_links.dropna(subset=["grid_region"])

        # Aggregate by grid_region (sum production, weighted average of cost)
        grouped = df_links.groupby("grid_region")
        sum_prod = grouped["production (MWh)"].sum()
        weighted_cost = grouped.apply(
            lambda g: (g["marginal_cost_total (USD/MWh)"] * g["production (MWh)"]).sum()
            / g["production (MWh)"].sum()
        )

        df_region_result = pd.DataFrame(
            {
                "production (MWh)": sum_prod,
                "marginal_cost_total (USD/MWh)": weighted_cost,
            }
        )

        # Round to 2 decimals
        df_region_result = df_region_result.round(2)

        # Reset index to make 'grid_region' a visible column
        df_region_result = df_region_result.reset_index()

        # Rename columns
        df_region_result = df_region_result.rename(
            columns={
                "grid_region": "Grid region",
                "production (MWh)": "Production (MWh)",
                "marginal_cost_total (USD/MWh)": "e-kerosene marginal cost (USD/MWh)",
            }
        )

        # Format numbers and hide index
        styled = df_region_result.style.format(
            {
                "Production (MWh)": "{:,.2f}",
                "e-kerosene marginal cost (USD/MWh)": "{:,.2f}",
            }
        ).hide(axis="index")

        # Extract year from network name
        match = re.search(r"\d{4}", name)
        year = match.group() if match else "unknown"

        print(f"\nYear: {year if year_title else name}\n")
        display(styled)


def compute_aviation_fuel_demand(
    networks,
    include_scenario: bool = False,
    scenario_as_index: bool = False,
    wide: bool = False,
):
    """
    Compute kerosene / e-kerosene demand per scenario.

    Parameters
    ----------
    networks : dict[str, pypsa.Network]
    include_scenario : bool
        (Ignored when wide=True; kept for backward compatibility in long mode)
    scenario_as_index : bool
        (Ignored when wide=True)
    wide : bool
        If True returns a wide table with MultiIndex columns (Scenario, Metric) and a single Year column.

    Returns
    -------
    pandas.DataFrame
    """

    results = {}

    for name, n in networks.items():
        m = re.search(r"(?:scenario_(\d{2})|Base)_(\d{4})", name)
        if m:
            scenario = f"scenario_{m.group(1)}" if m.group(1) else "Base"
            year = int(m.group(2))
        else:
            scenario = name
            digits = "".join(filter(str.isdigit, name[-4:]))
            year = int(digits) if digits.isdigit() else None

        kerosene_idx = n.loads.index[n.loads.carrier == "kerosene for aviation"]
        ekerosene_idx = n.loads.index[n.loads.carrier == "e-kerosene for aviation"]

        if kerosene_idx.empty and ekerosene_idx.empty:
            kerosene_twh = 0.0
            ekerosene_twh = 0.0
        else:
            w = n.snapshot_weightings.generators
            kerosene_mwh = (
                n.loads_t.p[kerosene_idx].multiply(w, axis=0).sum().sum()
                if len(kerosene_idx)
                else 0.0
            )
            ekerosene_mwh = (
                n.loads_t.p[ekerosene_idx].multiply(w, axis=0).sum().sum()
                if len(ekerosene_idx)
                else 0.0
            )
            kerosene_twh = kerosene_mwh / 1e6
            ekerosene_twh = ekerosene_mwh / 1e6

        results[name] = {
            "Scenario": scenario,
            "Year": year,
            "Kerosene (TWh)": kerosene_twh,
            "e-Kerosene (TWh)": ekerosene_twh,
        }

    df = pd.DataFrame.from_dict(results, orient="index").reset_index(drop=True)

    df["Total (TWh)"] = df["Kerosene (TWh)"] + df["e-Kerosene (TWh)"]
    df["e-Kerosene Share (%)"] = df.apply(
        lambda r: 0.0
        if r["Total (TWh)"] == 0
        else 100 * r["e-Kerosene (TWh)"] / r["Total (TWh)"],
        axis=1,
    )

    num_cols = df.select_dtypes(include="number").columns
    df[num_cols] = df[num_cols].applymap(lambda x: 0.0 if abs(x) < 1e-9 else x)

    if not wide:
        if not include_scenario:
            df = df.drop(columns=["Scenario"])
        else:
            if scenario_as_index:
                df = df.set_index("Scenario")
        df = df.sort_values(
            ["Year"]
            + (["Scenario"] if include_scenario and not scenario_as_index else [])
        )
        return df.reset_index(drop=not scenario_as_index)

    # Wide: columns (Scenario, Metric), single Year column
    metrics = [
        "Kerosene (TWh)",
        "e-Kerosene (TWh)",
        "Total (TWh)",
        "e-Kerosene Share (%)",
    ]
    wide_df = df.pivot_table(
        index="Year", columns="Scenario", values=metrics, aggfunc="first"
    )

    # Current pivot gives (Metric, Scenario); swap to (Scenario, Metric)
    wide_df.columns = wide_df.columns.swaplevel(0, 1)
    wide_df = wide_df.sort_index(axis=1, level=[0, 1])

    wide_df = wide_df.reset_index()  # keep single Year column
    return wide_df


def compute_emissions_from_links(net, net_definition="atmosphere"):
    """
    Compute CO2 flows (atmosphere, captured, sequestered) by process.

    Columns:
      - CO2 to Atmosphere (Mt CO2/year): flows to 'co2 atmosphere' (typo-safe)
      - CO2 Captured (Mt CO2/year): flows to generic CO2 buses incl. 'co2 stored'/'co2 storage'
      - CO2 Sequestered (Mt CO2/year): flows to 'co2 geological sequestration'
      - Net CO2 Emissions (Mt CO2/year):
          * 'atmosphere'  -> Atmosphere - Sequestered  (recommended; DAC negative)
          * 'neutral'     -> Atmosphere + Captured - Sequestered (DAC = 0)

    Units: Mt CO2/year
    """

    results = []
    bus_cols = [c for c in net.links.columns if re.fullmatch(r"bus\d+", c)]

    for link_name, row in net.links.iterrows():
        carrier = str(row["carrier"]).strip()

        co2_atm = 0.0
        co2_cap = 0.0
        co2_seq = 0.0

        for j, bus_col in enumerate(bus_cols):
            bus_val = str(row[bus_col]).lower().strip()
            p_col = f"p{j}"
            if p_col not in net.links_t or link_name not in net.links_t[p_col]:
                continue

            # tCO2/year (assuming 1 MWh on CO2 buses ~ 1 tCO2)
            flow = net.links_t[p_col][link_name].mean() * 8760.0

            # Order matters: atmosphere -> sequestration -> stored/storage -> generic CO2
            if ("co2 atmosphere" in bus_val) or ("co2 atmoshpere" in bus_val):
                co2_atm -= flow
            elif "co2 geological sequestration" in bus_val:
                co2_seq -= flow
            elif ("co2 stored" in bus_val) or ("co2 storage" in bus_val):
                co2_cap -= flow
            elif "co2" in bus_val:
                co2_cap -= flow

        results.append(
            {
                "Process": carrier,
                "CO2 to Atmosphere (Mt CO2/year)": co2_atm / 1e6,
                "CO2 Captured (Mt CO2/year)": co2_cap / 1e6,
                "CO2 Sequestered (Mt CO2/year)": co2_seq / 1e6,
            }
        )

    df = pd.DataFrame(results)
    summary = df.groupby("Process", as_index=False).sum()

    if net_definition == "neutral":
        net = (
            summary["CO2 to Atmosphere (Mt CO2/year)"]
            + summary["CO2 Captured (Mt CO2/year)"]
            - summary["CO2 Sequestered (Mt CO2/year)"]
        )
    else:  # 'atmosphere' (default)
        net = (
            summary["CO2 to Atmosphere (Mt CO2/year)"]
            - summary["CO2 Sequestered (Mt CO2/year)"]
        )

    summary["Net CO2 Emissions (Mt CO2/year)"] = net

    # Round numeric columns
    for c in summary.columns:
        if c != "Process":
            summary[c] = summary[c].round(2)

    return summary


def compute_emissions_grouped(net, carrier_groups):
    """
    Aggregate CO2 flows (captured, to atmosphere, sequestered) by carrier groups.
    Net emissions = Atmosphere - Sequestered.
    Units: Mt CO2/year
    """

    results = []
    bus_cols = [col for col in net.links.columns if re.fullmatch(r"bus\d+", col)]

    for link_name, row in net.links.iterrows():
        carrier = row["carrier"]

        # skip if not in any group
        if not any(carrier in carriers for carriers in carrier_groups.values()):
            continue

        # skip technical storage carriers
        if "storage" in carrier.lower():
            continue

        co2_atmos, co2_captured, co2_sequestered = 0.0, 0.0, 0.0

        for j, bus_col in enumerate(bus_cols):
            bus_val = str(row[bus_col]).lower().strip()
            p_col = f"p{j}"
            if p_col not in net.links_t or link_name not in net.links_t[p_col]:
                continue

            flow = net.links_t[p_col][link_name].mean() * 8760

            if "co2 atmosphere" in bus_val or "co2 atmoshpere" in bus_val:
                co2_atmos -= flow
            elif "geological sequestration" in bus_val:
                co2_sequestered -= flow
            elif "co2" in bus_val:
                co2_captured -= flow

        results.append(
            {
                "carrier": carrier,
                "co2_atmosphere": co2_atmos,
                "co2_captured": co2_captured,
                "co2_sequestered": co2_sequestered,
            }
        )

    df = pd.DataFrame(results)

    group_results = []
    for group_name, group_carriers in carrier_groups.items():
        group_df = df[df["carrier"].isin(group_carriers)]
        if group_df.empty:
            continue

        atm = group_df["co2_atmosphere"].sum()
        captured = group_df["co2_captured"].sum()
        sequestered = group_df["co2_sequestered"].sum()

        group_results.append(
            {
                "carrier group": group_name,
                "CO2 to Atmosphere (Mt CO2/year)": atm / 1e6,
                "CO2 Captured (Mt CO2/year)": captured / 1e6,
                "CO2 Sequestered (Mt CO2/year)": sequestered / 1e6,
                "Net CO2 Emissions (Mt CO2/year)": (atm - sequestered) / 1e6,
            }
        )

    return pd.DataFrame(group_results).round(2)


def compute_emissions_by_state(net, carrier_groups):
    """
    Compute CO2 flows (to atmosphere, stored/sequestered) by State and process group.
    Net emissions = Atmosphere - Sequestered.
    Units: Mt CO2/year
    """

    results = []
    bus_cols = [col for col in net.links.columns if re.fullmatch(r"bus\d+", col)]

    for link_name, row in net.links.iterrows():
        carrier = row["carrier"]

        # assign process group
        group = next(
            (g for g, carriers in carrier_groups.items() if carrier in carriers), None
        )
        if group is None:
            continue

        co2_atmos = 0.0
        co2_sequestered = 0.0

        for j, bus_col in enumerate(bus_cols):
            bus_val = str(row[bus_col]).lower().strip()
            p_col = f"p{j}"
            if p_col not in net.links_t or link_name not in net.links_t[p_col]:
                continue

            flow = net.links_t[p_col][link_name].mean() * 8760

            if "co2 atmosphere" in bus_val or "co2 atmoshpere" in bus_val:
                co2_atmos -= flow
            elif "geological sequestration" in bus_val or "co2 stored" in bus_val:
                co2_sequestered -= flow

        # infer State from connected buses
        state = "Unknown"
        for bus_col in bus_cols:
            bus = row[bus_col]
            if bus in net.buses.index and "state" in net.buses.columns:
                s = net.buses.loc[bus, "state"]
                if pd.notna(s) and s != "Unknown":
                    state = s
                    break

        results.append(
            {
                "State": state,
                "group": group,
                "CO2 to Atmosphere (Mt CO2/year)": co2_atmos,
                "CO2 Sequestered (Mt CO2/year)": co2_sequestered,
            }
        )

    df = pd.DataFrame(results)

    if df.empty:
        return pd.DataFrame(
            columns=[
                "State",
                "group",
                "CO2 to Atmosphere (Mt CO2/year)",
                "CO2 Sequestered (Mt CO2/year)",
                "Net CO2 Emissions (Mt CO2/year)",
            ]
        )

    summary = (
        df.groupby(["State", "group"])[
            ["CO2 to Atmosphere (Mt CO2/year)", "CO2 Sequestered (Mt CO2/year)"]
        ]
        .sum()
        .reset_index()
    )

    summary["Net CO2 Emissions (Mt CO2/year)"] = (
        summary["CO2 to Atmosphere (Mt CO2/year)"]
        - summary["CO2 Sequestered (Mt CO2/year)"]
    )

    for c in [
        "CO2 to Atmosphere (Mt CO2/year)",
        "CO2 Sequestered (Mt CO2/year)",
        "Net CO2 Emissions (Mt CO2/year)",
    ]:
        summary[c] = (summary[c] / 1e6).round(2)

    return summary


def plot_emissions_maps_by_group(
    all_state_emissions,
    path_shapes,
    title,
    column: str = "Net CO2 Emissions (Mt CO2/year)",
    vmin=None,
    vmax=None,
):
    """
    Plot CO2 emissions maps by process group and State.

    Parameters
    ----------
    all_state_emissions : pd.DataFrame
        Output from compute_emissions_by_state (must contain 'State' and 'group').
    path_shapes : str
        Path to state shapefile/geojson.
    title : str
        Title of the figure (e.g. scenario or year).
    column : str
        Column to plot (default: "Net CO2 Emissions (Mt CO2/year)").
    vmin, vmax : float
        Color scale limits.
    """

    if column not in all_state_emissions.columns:
        raise KeyError(
            f"Column '{column}' not found in DataFrame. "
            f"Available columns: {list(all_state_emissions.columns)}"
        )

    # Load shapefile and enforce CRS
    gdf_states = gpd.read_file(path_shapes).to_crs("EPSG:4326")
    gdf_states["State"] = gdf_states["ISO_1"].str[-2:]

    groups = all_state_emissions["group"].unique()
    n = len(groups)
    ncols = 2
    nrows = int(np.ceil(n / ncols))

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(8 * ncols, 6 * nrows))
    axes = axes.flat if n > 1 else [axes]

    for i, group in enumerate(groups):
        ax = axes[i]
        df_group = all_state_emissions[all_state_emissions["group"] == group].copy()

        merged = gdf_states.merge(df_group, on="State", how="left")

        # Replace exact zeros with NaN so they are plotted as white
        plot_col = f"{column}_plot"
        merged[plot_col] = merged[column].replace(0, np.nan)

        merged.plot(
            column=plot_col,
            # red = positive (emissions), green = negative (removals)
            cmap="RdYlGn_r",
            legend=True,
            ax=ax,
            vmin=vmin,
            vmax=vmax,
            # white = 0 or missing
            missing_kwds={"color": "white", "label": "0 or no data"},
            edgecolor="black",
        )

        ax.set_title(f"{group}", fontsize=12)
        ax.set_xlim([-130, -65])
        ax.set_ylim([20, 55])
        ax.axis("off")

        # format legend
        leg = ax.get_legend()
        if leg:
            leg.set_bbox_to_anchor((1.05, 0.5))
            leg.set_title("Mt CO2/year", prop={"size": 8})
            for t in leg.get_texts():
                t.set_fontsize(8)

    # hide unused subplots
    for j in range(i + 1, len(axes)):
        axes[j].axis("off")

    fig.suptitle(f"{column} by process group and State – {title}", fontsize=14)
    plt.tight_layout()
    plt.subplots_adjust(top=0.9, right=0.85)
    showfig()


def evaluate_res_ces_by_state(
    networks, ces, res, ces_carriers, res_carriers, multiple_scenarios=False
):
    results = {}

    for name, network in networks.items():
        year = int(name[-4:])
        year_str = str(year)

        snapshots = network.snapshots
        timestep_h = (snapshots[1] - snapshots[0]).total_seconds() / 3600
        snapshots_slice = slice(None)

        gen_and_sto_carriers = {
            "csp",
            "solar",
            "onwind",
            "offwind-dc",
            "offwind-ac",
            "nuclear",
            "geothermal",
            "ror",
            "hydro",
            "solar rooftop",
        }
        link_carriers = ["coal", "oil", "OCGT", "CCGT", "biomass", "lignite"]

        electric_buses = set(
            network.buses.index[
                ~network.buses.carrier.str.contains(
                    "heat|gas|H2|oil|coal", case=False, na=False
                )
            ]
        )

        # Generators
        gen = network.generators[
            network.generators.carrier.isin(gen_and_sto_carriers)
        ].copy()
        gen["state"] = gen["bus"].map(network.buses["state"])
        gen = gen[gen["state"].notna()]

        gen_p = network.generators_t.p.loc[snapshots_slice, gen.index].clip(lower=0)
        gen_energy = gen_p.multiply(timestep_h).sum()  # MWh per generator
        gen_energy = gen_energy.to_frame(name="energy_mwh")
        gen_energy["carrier"] = gen.loc[gen_energy.index, "carrier"]
        gen_energy["state"] = gen.loc[gen_energy.index, "state"]

        # Storage
        sto = network.storage_units[
            network.storage_units.carrier.isin(gen_and_sto_carriers)
        ].copy()
        sto["state"] = sto["bus"].map(network.buses["state"])
        sto = sto[sto["state"].notna()]

        sto_p = network.storage_units_t.p.loc[snapshots_slice, sto.index].clip(lower=0)
        sto_energy = sto_p.multiply(timestep_h).sum()
        sto_energy = sto_energy.to_frame(name="energy_mwh")
        sto_energy["carrier"] = sto.loc[sto_energy.index, "carrier"]
        sto_energy["state"] = sto.loc[sto_energy.index, "state"]

        # Links
        link_data = []
        for i, link in network.links.iterrows():
            if (
                link["carrier"] in link_carriers
                and link["bus1"] in electric_buses
                and pd.notna(network.buses.loc[link["bus1"], "state"])
            ):
                p1 = -network.links_t.p1.loc[snapshots_slice, i].clip(upper=0)
                energy_mwh = p1.sum() * timestep_h
                link_data.append(
                    {
                        "carrier": link["carrier"],
                        "state": network.buses.loc[link["bus1"], "state"],
                        "energy_mwh": energy_mwh,
                    }
                )

        link_energy = pd.DataFrame(link_data)

        # Combine all generations
        all_energy = pd.concat(
            [
                gen_energy[["carrier", "state", "energy_mwh"]],
                sto_energy[["carrier", "state", "energy_mwh"]],
                link_energy[["carrier", "state", "energy_mwh"]],
            ]
        )

        # Aggregate by State
        state_totals = all_energy.groupby("state")["energy_mwh"].sum()
        state_ces = (
            all_energy[all_energy["carrier"].isin(ces_carriers)]
            .groupby("state")["energy_mwh"]
            .sum()
        )
        state_res = (
            all_energy[all_energy["carrier"].isin(res_carriers)]
            .groupby("state")["energy_mwh"]
            .sum()
        )

        df = pd.DataFrame(
            {
                "Total (MWh)": state_totals,
                "CES_energy": state_ces,
                "RES_energy": state_res,
            }
        ).fillna(0)

        df["% CES"] = 100 * df["CES_energy"] / df["Total (MWh)"]
        df["% RES"] = 100 * df["RES_energy"] / df["Total (MWh)"]

        # Targets
        if year_str in ces.columns:
            df["% CES target"] = df.index.map(
                lambda state: ces[year_str].get(state, float("nan"))
            )
        else:
            df["% CES target"] = float("nan")

        if year_str in res.columns:
            df["% RES target"] = df.index.map(
                lambda state: res[year_str].get(state, float("nan"))
            )
        else:
            df["% RES target"] = float("nan")

        df["% RES target"] = df["% RES target"].apply(
            lambda x: "N/A" if pd.isna(x) else round(x * 100, 2)
        )
        df["% CES target"] = df["% CES target"].apply(
            lambda x: "N/A" if pd.isna(x) else round(x * 100, 2)
        )

        df = df[["% RES", "% RES target", "% CES", "% CES target"]].round(2)
        if multiple_scenarios:
            results[name] = df.sort_index()
        else:
            results[year] = df.sort_index()
    return results


def plot_network_generation_and_transmission(
    n, key, tech_colors, nice_names, regions_onshore, title_year=True
):
    # Define generation and link carriers
    gen_carriers = {
        "onwind",
        "offwind-ac",
        "offwind-dc",
        "solar",
        "solar rooftop",
        "csp",
        "nuclear",
        "geothermal",
        "ror",
        "PHS",
        "battery discharger",
    }
    link_carriers = {
        "OCGT",
        "CCGT",
        "coal",
        "oil",
        "biomass",
        "urban central solid biomass CHP",
        "urban central gas CHP",
        "battery discharger",
    }

    # Generator and storage capacity
    gen_p_nom_opt = n.generators[n.generators.carrier.isin(gen_carriers)]
    gen_p_nom_opt = gen_p_nom_opt.groupby(["bus", "carrier"]).p_nom_opt.sum()

    sto_p_nom_opt = n.storage_units[n.storage_units.carrier.isin(gen_carriers)]
    sto_p_nom_opt = sto_p_nom_opt.groupby(["bus", "carrier"]).p_nom_opt.sum()

    # Link capacity (scaled by efficiency)
    link_mask = (
        n.links.efficiency.notnull()
        & (n.links.p_nom_opt > 0)
        & n.links.carrier.isin(link_carriers)
    )
    electricity_links = n.links[link_mask].copy()
    electricity_links["electric_output"] = (
        electricity_links.p_nom_opt * electricity_links.efficiency
    )
    link_p_nom_opt = electricity_links.groupby(
        ["bus1", "carrier"]
    ).electric_output.sum()
    link_p_nom_opt.index = link_p_nom_opt.index.set_names(["bus", "carrier"])

    # Combine all sources
    bus_carrier_capacity = pd.concat([gen_p_nom_opt, sto_p_nom_opt, link_p_nom_opt])
    bus_carrier_capacity = bus_carrier_capacity.groupby(level=[0, 1]).sum()
    bus_carrier_capacity = bus_carrier_capacity[bus_carrier_capacity > 0]

    # Keep only buses with valid coordinates
    valid_buses = n.buses.dropna(subset=["x", "y"])
    valid_buses = valid_buses[
        (valid_buses["x"] > -200)
        & (valid_buses["x"] < 200)
        & (valid_buses["y"] > -90)
        & (valid_buses["y"] < 90)
    ]

    # Normalize bus names (remove " low voltage")
    def normalize_bus_name(bus_name):
        return bus_name.replace(" low voltage", "")

    bus_carrier_capacity = bus_carrier_capacity.reset_index()
    bus_carrier_capacity["bus"] = bus_carrier_capacity["bus"].apply(normalize_bus_name)
    bus_carrier_capacity["carrier"] = bus_carrier_capacity["carrier"].replace(
        {"offwind-ac": "offwind", "offwind-dc": "offwind"}
    )
    bus_carrier_capacity = bus_carrier_capacity.groupby(
        ["bus", "carrier"], as_index=False
    ).sum()
    bus_carrier_capacity = bus_carrier_capacity.set_index(["bus", "carrier"]).squeeze()
    capacity_df = bus_carrier_capacity.unstack(fill_value=0)
    capacity_df = capacity_df.loc[capacity_df.index.intersection(valid_buses.index)]

    # Setup map background
    fig, ax = plt.subplots(
        figsize=(28, 10), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    bbox = box(-130, 20, -60, 50)
    regions_onshore_clipped = regions_onshore.to_crs(epsg=4326).clip(bbox)

    regions_onshore_clipped.plot(
        ax=ax,
        facecolor="whitesmoke",
        edgecolor="gray",
        alpha=0.7,
        linewidth=0.5,
        zorder=0,
    )

    # Plot only DC links
    original_links = n.links.copy()
    n.links = n.links[n.links.carrier == "DC"]

    line_scale = 5e3
    n.plot(
        ax=ax,
        bus_sizes=0,
        bus_alpha=0,
        line_widths=n.lines.s_nom_opt / line_scale,
        link_widths=n.links.p_nom_opt / line_scale,
        line_colors="teal",
        link_colors="turquoise",
        color_geomap=False,
        flow=None,
    )
    n.links = original_links

    # Draw pie charts at bus locations
    pie_scale = 0.003
    for bus_id, capacities in capacity_df.iterrows():
        x, y = valid_buses.loc[bus_id, ["x", "y"]]
        if not bbox.contains(gpd.points_from_xy([x], [y])[0]):
            continue
        values = capacities.values
        total = values.sum()
        if total == 0:
            continue
        size = np.clip(np.sqrt(total) * pie_scale, 0.1, 1.5)
        colors = [tech_colors.get(c, "gray") for c in capacities.index]
        start_angle = 0
        for val, color in zip(values, colors):
            if val == 0:
                continue
            angle = 360 * val / total
            wedge = Wedge(
                center=(x, y),
                r=size,
                theta1=start_angle,
                theta2=start_angle + angle,
                facecolor=color,
                edgecolor="k",
                linewidth=0.3,
                transform=ccrs.PlateCarree()._as_mpl_transform(ax),
                zorder=5,
            )
            ax.add_patch(wedge)
            start_angle += angle

    # Legends

    # Bus capacity (marker size legend)
    bus_caps = [5, 10, 50]
    bus_marker_sizes = [np.sqrt(cap) * pie_scale * 1000 for cap in bus_caps]
    bus_patches = [
        mlines.Line2D(
            [],
            [],
            linestyle="None",
            marker="o",
            color="gray",
            markersize=size,
            label=f"{cap} GW",
            markerfacecolor="gray",
            alpha=0.5,
        )
        for cap, size in zip(bus_caps, bus_marker_sizes)
    ]
    bus_legend = ax.legend(
        handles=bus_patches,
        title="Bus Capacity",
        title_fontsize=12,
        fontsize=10,
        frameon=False,
        loc="upper right",
        bbox_to_anchor=(1.085, 1.0),
        labelspacing=1.4,
    )

    # AC line capacity
    ac_caps = [5e3, 20e3, 50e3]
    ac_patches = [
        mlines.Line2D(
            [],
            [],
            color="teal",
            linewidth=cap / line_scale,
            label=f"{int(cap / 1e3)} GW",
        )
        for cap in ac_caps
    ]
    ac_legend = ax.legend(
        handles=ac_patches,
        title="AC Line Capacity",
        title_fontsize=12,
        fontsize=10,
        frameon=False,
        loc="upper right",
        bbox_to_anchor=(1.1, 0.83),
        labelspacing=1.1,
    )

    # DC link capacity
    dc_caps = [2e3, 5e3, 10e3]
    dc_patches = [
        mlines.Line2D(
            [],
            [],
            color="turquoise",
            linewidth=cap / line_scale,
            label=f"{int(cap / 1e3)} GW",
        )
        for cap in dc_caps
    ]
    dc_legend = ax.legend(
        handles=dc_patches,
        title="DC Link Capacity",
        title_fontsize=12,
        fontsize=10,
        frameon=False,
        loc="upper right",
        bbox_to_anchor=(1.1, 0.68),
        labelspacing=1.1,
    )

    # Carrier legend (force preferred order with nice_names)
    preferred_order = [
        "Coal",
        "Gas CCGT",
        "Gas OCGT",
        "Gas CHP",
        "Oil",
        "Nuclear",
        "Biomass",
        "Biomass CHP",
        "Conventional hydro",
        "Run-of-River hydro",
        "Pumped hydro storage",
        "Utility-scale solar",
        "Rooftop solar",
        "CSP",
        "Onshore wind",
        "Offshore wind",
        "Battery",
    ]

    # Map raw carriers to pretty names
    carriers_present = {
        nice_names.get(c, c): c for c in capacity_df.columns if capacity_df[c].sum() > 0
    }

    # Keep only the carriers that are in the preferred order and present in data
    ordered_carriers = [c for c in preferred_order if c in carriers_present.keys()]

    # Build handles with the correct color from the raw key
    carrier_handles = [
        mpatches.Patch(color=tech_colors.get(carriers_present[c], "gray"), label=c)
        for c in ordered_carriers
    ]

    carrier_legend = ax.legend(
        handles=carrier_handles,
        title="Technology",
        title_fontsize=13,
        fontsize=11,
        frameon=False,
        loc="upper right",
        bbox_to_anchor=(1.125, 0.54),
        ncol=2,
        labelspacing=1.05,
    )

    ax.add_artist(bus_legend)
    ax.add_artist(ac_legend)
    ax.add_artist(dc_legend)
    ax.add_artist(carrier_legend)

    # Set map extent
    ax.set_extent([-125, -65, 20, 55], crs=ccrs.PlateCarree())
    ax.autoscale(False)

    # Title
    year = key[-4:]
    ax.set_title(
        f"Installed electricity generation and transmission capacity – {year if title_year else key}",
        fontsize=14,
    )

    plt.tight_layout()
    showfig()


def compute_installed_capacity_by_carrier(
    networks, nice_names=None, display_result=True, column_year=True
):
    totals_by_carrier = {}

    for name, net in networks.items():
        # Conventional generator and storage carriers
        gen_carriers = {
            "onwind",
            "offwind",
            "solar",
            "solar rooftop",
            "csp",
            "nuclear",
            "geothermal",
            "ror",
            "PHS",
            "hydro",
        }
        link_carriers = {
            "OCGT",
            "CCGT",
            "coal",
            "oil",
            "biomass",
            "urban central solid biomass CHP",
            "urban central gas CHP",
            "battery discharger",
        }

        # Generators
        gen = net.generators.copy()
        gen = gen[gen.carrier.isin(gen_carriers)]
        gen_totals = gen.groupby("carrier")["p_nom_opt"].sum()

        # Storage units
        sto = net.storage_units.copy()
        sto = sto[sto.carrier.isin(gen_carriers)]
        sto_totals = sto.groupby("carrier")["p_nom_opt"].sum()

        # Links (efficiency-scaled output)
        links = net.links.copy()
        mask = (
            links.efficiency.notnull()
            & (links.p_nom_opt > 0)
            & links.carrier.isin(link_carriers)
        )
        links = links[mask]

        links_totals = links.groupby("carrier").apply(
            lambda df: (df["p_nom_opt"] * df["efficiency"]).sum()
        )

        # Combine all contributions
        all_totals = pd.concat([gen_totals, sto_totals, links_totals])
        all_totals = all_totals.groupby(all_totals.index).sum()
        all_totals = all_totals[all_totals > 0]
        totals_by_carrier[name] = all_totals

    # Assemble final dataframe
    carrier_capacity_df = pd.DataFrame(totals_by_carrier).fillna(0)

    # Extract years and sort
    if column_year:
        carrier_capacity_df.columns = [
            int(name[-4:]) for name in carrier_capacity_df.columns
        ]
    carrier_capacity_df = carrier_capacity_df[sorted(carrier_capacity_df.columns)]

    # Convert to GW
    carrier_capacity_df = (carrier_capacity_df / 1000).round(2)

    # Rename index if nice_names is provided
    if nice_names:
        carrier_capacity_df = carrier_capacity_df.rename(index=nice_names)

    # Apply preferred order
    preferred_order = [
        "Coal",
        "Gas CCGT",
        "Gas OCGT",
        "Gas CHP",
        "Oil",
        "Nuclear",
        "Biomass",
        "Biomass CHP",
        "Geothermal",
        "Conventional hydro",
        "Run-of-River hydro",
        "Pumped hydro storage",
        "Onshore wind",
        "Offshore wind",
        "Utility-scale solar",
        "Rooftop solar",
        "CSP",
        "Battery",
    ]
    available = carrier_capacity_df.index.tolist()
    ordered_index = [c for c in preferred_order if c in available] + [
        c for c in available if c not in preferred_order
    ]
    carrier_capacity_df = carrier_capacity_df.loc[ordered_index]

    if display_result:
        print("\nInstalled capacity by technology (GW)\n")
        display(carrier_capacity_df)

    return carrier_capacity_df


def compute_system_costs(network, rename_capex, rename_opex, name_tag):
    """
    Compute CAPEX and OPEX (including input-cost OPEX for industrial and link components).
    Aggregates all values to billion EUR.

    Parameters
    ----------
    network : pypsa.Network
        Solved PyPSA network.
    rename_capex : dict
        Mapping of carrier names for CAPEX aggregation.
    rename_opex : dict
        Mapping of carrier names for OPEX aggregation.
    name_tag : str
        Scenario name containing the year (e.g. 'Base_2030').

    Returns
    -------
    pandas.DataFrame
        With columns: tech_label, main_category, cost_type,
        cost_billion, year, scenario.
    """

    # --- PyPSA statistics ---
    costs_raw = network.statistics()[["Capital Expenditure", "Operational Expenditure"]]
    year_str = name_tag[-4:]

    # CAPEX
    capex_raw = costs_raw[["Capital Expenditure"]].reset_index()
    capex_raw["tech_label"] = (
        capex_raw["carrier"].map(rename_capex).fillna(capex_raw["carrier"])
    )
    capex_raw["main_category"] = capex_raw["tech_label"]

    capex_grouped = (
        capex_raw.groupby("tech_label", as_index=False)
        .agg({"Capital Expenditure": "sum", "main_category": "first"})
        .rename(columns={"Capital Expenditure": "cost_billion"})
    )
    capex_grouped["cost_billion"] /= 1e9
    capex_grouped["cost_type"] = "Capital expenditure"
    capex_grouped["year"] = year_str
    capex_grouped["scenario"] = name_tag

    # OPEX (base)
    opex_raw = costs_raw[["Operational Expenditure"]].reset_index()
    opex_raw["tech_label"] = (
        opex_raw["carrier"].map(rename_opex).fillna(opex_raw["carrier"])
    )
    opex_raw["main_category"] = opex_raw["tech_label"]

    opex_grouped = (
        opex_raw.groupby("tech_label", as_index=False)
        .agg({"Operational Expenditure": "sum", "main_category": "first"})
        .rename(columns={"Operational Expenditure": "cost_billion"})
    )
    opex_grouped["cost_billion"] /= 1e9
    opex_grouped["cost_type"] = "Operational expenditure"
    opex_grouped["year"] = year_str
    opex_grouped["scenario"] = name_tag

    # EXTRA OPEX FROM INPUT FLOWS (GENERIC)
    w = network.snapshot_weightings["objective"]
    bus_cols = [c for c in network.links.columns if c.startswith("bus")]
    results_extra = []

    for link_id, row in network.links.iterrows():
        tech = row["carrier"]

        for bcol in bus_cols:
            bus = row[bcol]
            if pd.isna(bus):
                continue

            # Only count buses with marginal price
            if bus not in network.buses_t.marginal_price.columns:
                continue

            # dispatch column p0,p1,...
            idx = bcol[3:]  # bus2 : "2"
            pcol = f"p{idx}"  # : p2

            if pcol not in network.links_t or link_id not in network.links_t[pcol]:
                continue

            # annual flow
            flow = (network.links_t[pcol][link_id] * w).sum()

            # INPUT = flow > 0   (100% consistent with CC calculations)
            inflow = max(flow, 0.0)
            if inflow <= 0:
                continue

            price = float(network.buses_t.marginal_price[bus].mean())
            fuel_cost = inflow * price  # EUR

            results_extra.append(
                {
                    "tech_label": tech,
                    "main_category": tech,
                    "cost_type": "Operational expenditure",
                    "cost_billion": fuel_cost / 1e9,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

    link_opex_df = pd.DataFrame(results_extra)

    # Apply renaming rules to extra OPEX
    if not link_opex_df.empty:
        link_opex_df["tech_label"] = link_opex_df["tech_label"].replace(rename_opex)
        link_opex_df["main_category"] = link_opex_df["tech_label"]

    # MERGE ALL COSTS
    df_all = pd.concat([capex_grouped, opex_grouped, link_opex_df], ignore_index=True)
    return df_all


def assign_macro_category(row, categories_capex, categories_opex):
    if row["cost_type"] == "Capital expenditure":
        return categories_capex.get(row["tech_label"], "Other")
    elif row["cost_type"] == "Operational expenditure":
        return categories_opex.get(row["tech_label"], "Other")
    else:
        return "Other"


def calculate_total_inputs_outputs_ft(
    networks: dict,
    ft_carrier: str = "Fischer-Tropsch",
    year_index: bool = True,
    include_scenario: bool = False,
    scenario_regex: str = r"(scenario_\d{2})",
    keep_empty: bool = False,
    wide: bool = False,
    scenario_first_level: bool = True,
):
    """
    Calculate Fischer-Tropsch (FT) process inputs/outputs per network.

    For each network (scenario-year):
      - Used electricity (TWh)    : links_t.p3
      - Used hydrogen (TWh / t)   : links_t.p0
      - Used CO2 (Mt)             : links_t.p2  (sum of flow; negative convention => absolute used)
      - Produced e-kerosene (TWh) : links_t.p1  (output taken as negative of p1 aggregated)

    Parameters
    ----------
    networks : dict
        {name: pypsa.Network}
    ft_carrier : str
        Link carrier name to filter Fischer-Tropsch units.
    year_index : bool
        If True (default) the output keeps a numeric Year column (and sorts by it).
        If False, original network name retained in 'Year' column (for backward compatibility).
    include_scenario : bool
        If True, adds a 'Scenario' column (parsed with scenario_regex; 'Base' if not found).
    scenario_regex : str
        Regex to extract scenario label from the network name.
    keep_empty : bool
        If True, include rows for scenarios without FT links (filled with zeros).
    wide : bool
        If True, returns a pivoted (wide) table:
            - Index: Year (or Scenario) depending on flags
            - Columns: metrics (or MultiIndex with scenario + metric)
    scenario_first_level : bool
        If wide=True and include_scenario=True: choose MultiIndex order:
            True  => (Scenario, Metric)
            False => (Metric, Scenario)

    Returns
    -------
    pd.DataFrame
        Long or wide formatted table with FT energy/material balances.

    Notes
    -----
    - Previous behavior preserved when include_scenario=False, wide=False.
    - Hydrogen tons conversion: TWh -> kWh -> kg -> t  (1 TWh = 1e9 kWh; LHV ≈ 33 kWh/kg).
    """

    rows = []

    for name, net in networks.items():
        # Extract scenario
        scenario_match = re.search(scenario_regex, name) if include_scenario else None
        scenario = (
            scenario_match.group(1)
            if (scenario_match and include_scenario)
            else ("Base" if include_scenario else None)
        )

        ft_links = net.links[net.links.carrier == ft_carrier]
        if ft_links.empty:
            if keep_empty:
                # Extract year anyway
                year_match = re.search(r"\d{4}", name)
                yr_val = int(year_match.group()) if year_match else None
                rows.append(
                    {
                        "Year": (yr_val if year_index else name),
                        "Scenario": scenario,
                        "Used electricity (TWh)": 0.0,
                        "Used hydrogen (TWh)": 0.0,
                        "Used hydrogen (t)": 0.0,
                        "Used CO2 (Mt)": 0.0,
                        "Produced e-kerosene (TWh)": 0.0,
                    }
                )
            continue

        ft_ids = ft_links.index

        # Timestep hours
        timestep_h = (
            (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600
            if len(net.snapshots) > 1
            else 1.0
        )

        def _energy(df, ids):
            sel = df.reindex(columns=ids, fill_value=0.0)
            return (sel * timestep_h).sum().sum() / 1e6  # MWh -> TWh

        elec_twh = _energy(net.links_t.p3, ft_ids)
        h2_twh = _energy(net.links_t.p0, ft_ids)
        kerosene_twh_raw = _energy(net.links_t.p1, ft_ids)
        # p1 likely negative for output (depending on sign convention). Keep positive production:
        kerosene_twh = -kerosene_twh_raw

        def _co2(df, ids):
            sel = df.reindex(columns=ids, fill_value=0.0)
            # Sum raw (can be negative). For clarity assume negative means consumption:
            total = (sel * timestep_h).sum().sum() / 1e6  # t -> Mt
            # Use absolute if negative is consumption; keep sign if desired. Here make it positive usage.
            return -total if total < 0 else total

        co2_mt = _co2(net.links_t.p2, ft_ids)

        # Hydrogen tons
        h2_tons = h2_twh * 1e9 / 33 / 1000  # TWh -> kWh -> kg -> t

        year_match = re.search(r"\d{4}", name)
        if not year_match:
            continue
        year_val = int(year_match.group())

        row = {
            "Year": (year_val if year_index else name),
            "Used electricity (TWh)": elec_twh,
            "Used hydrogen (TWh)": h2_twh,
            "Used hydrogen (t)": h2_tons,
            "Used CO2 (Mt)": co2_mt,
            "Produced e-kerosene (TWh)": kerosene_twh,
        }
        if include_scenario:
            row["Scenario"] = scenario
        rows.append(row)

    if not rows:
        return pd.DataFrame(
            columns=[
                "Year",
                *(["Scenario"] if include_scenario else []),
                "Used electricity (TWh)",
                "Used hydrogen (TWh)",
                "Used hydrogen (t)",
                "Used CO2 (Mt)",
                "Produced e-kerosene (TWh)",
            ]
        )

    df = pd.DataFrame(rows)

    # Ensure proper dtypes
    if year_index and "Year" in df.columns:
        df["Year"] = pd.to_numeric(df["Year"], errors="coerce")

    # Sort
    sort_cols = []
    if include_scenario:
        sort_cols.append("Scenario")
    if "Year" in df.columns:
        sort_cols.append("Year")
    if sort_cols:
        df = df.sort_values(sort_cols)

    if not wide:
        return df.reset_index(drop=True)

    # Wide formatting
    value_cols = [
        "Used electricity (TWh)",
        "Used hydrogen (TWh)",
        "Used hydrogen (t)",
        "Used CO2 (Mt)",
        "Produced e-kerosene (TWh)",
    ]

    if include_scenario:
        if scenario_first_level:
            # Columns: Scenario -> Metric
            pivot_index = "Year" if year_index else None
            if pivot_index:
                wide_df = df.pivot(
                    index=pivot_index, columns="Scenario", values=value_cols
                )
                # Reorder to Scenario first => (Scenario, Metric)
                wide_df = wide_df.swaplevel(0, 1, axis=1).sort_index(axis=1)
            else:
                # No year index: just aggregate by scenario (single row per scenario)
                wide_df = df.groupby("Scenario")[value_cols].sum()
        else:
            # Columns: Metric -> Scenario
            pivot_index = "Year" if year_index else None
            if pivot_index:
                wide_df = df.pivot(
                    index=pivot_index, columns="Scenario", values=value_cols
                )
            else:
                wide_df = df.groupby("Scenario")[value_cols].sum().T
        return wide_df

    # include_scenario False in wide mode
    if year_index:
        wide_df = df.set_index("Year")[value_cols]
    else:
        # 'Year' holds original name when year_index=False
        wide_df = df.set_index("Year")[value_cols]
    return wide_df


def compute_ekerosene_production_cost_by_region(
    networks: dict,
    year_title: bool = True,
    aggregate: bool = False,
    wide: bool = False,
    scenario_regex: str = r"(?:scenario_(\d{2})|Base)",
    scenario_first_level: bool = True,
    min_production_twh: float = 1e-3,
    return_weighted_average: bool = False,
    expected_scenarios: list | None = None,
    expected_years: list | None = None,
    fill_cost_with: float | None = None,  # None → NaN
):
    """
    Extended: ensure ALL expected scenarios and years (even with no production)
    appear in the output (long and wide). Missing combinations get:
        Production (TWh) = 0
        Costs = NaN (or fill_cost_with if provided)

    Parameters added
    ----------------
    expected_scenarios : list | None
        Full list of scenarios to include (e.g. ['Base','scenario_01',...]).
        If None -> inferred from networks.
    expected_years : list | None
        Full list of model years to include (e.g. [2023,2030,2035,2040]).
        If None -> inferred from networks.
    fill_cost_with : float | None
        Value to fill missing cost metrics. Default None keeps NaN.
    """

    all_rows = []

    # Infer scenarios / years if not provided
    inferred_scenarios = set()
    inferred_years = set()

    for name in networks.keys():
        m = re.search(r"(?:scenario_(\d{2})|Base)_(\d{4})", name)
        if not m:
            continue
        scen = f"scenario_{m.group(1)}" if m.group(1) else "Base"
        yr = int(m.group(2))
        inferred_scenarios.add(scen)
        inferred_years.add(yr)

    if expected_scenarios is None:
        expected_scenarios = sorted(inferred_scenarios, key=lambda x: (x != "Base", x))
    if expected_years is None:
        expected_years = sorted(inferred_years)

    # --- Energy content conversion for e-kerosene ---
    # 1 MWh = 3600 MJ
    # Energy density of kerosene ≈ 34 MJ/L
    # 1 US gallon = 3.78541 L
    # Energy per liter = 34 / 3600 = 0.009444... MWh/L
    # Energy per gallon = 0.009444... × 3.78541 ≈ 0.0357 MWh/gallon
    MWH_PER_GALLON = 34.0 / 3600.0 * 3.78541

    cost_cols = [
        "Electricity cost (USD/MWh e-kerosene)",
        "Hydrogen cost (USD/MWh e-kerosene)",
        "CO2 cost (USD/MWh e-kerosene)",
        "Total production cost (USD/MWh e-kerosene)",
        "Total cost (USD/gallon e-kerosene)",
    ]
    metrics_all = ["Production (TWh)"] + cost_cols

    for name, net in networks.items():
        year_match = re.search(r"\d{4}", name)
        if not year_match:
            continue
        year = int(year_match.group())

        scen_match = re.search(scenario_regex, name)
        if scen_match:
            scen_raw = scen_match.group(0)
            scenario = "Base" if "Base" in scen_raw else scen_raw
        else:
            scenario = "Base"

        ft_links = net.links[
            (net.links.carrier.str.contains("Fischer-Tropsch", case=False, na=False))
            & (
                (net.links.get("p_nom_opt", 0) > 0)
                | (
                    (net.links.get("p_nom", 0) > 0)
                    & (net.links.get("p_nom_extendable", False) == False)
                )
            )
        ]
        if ft_links.empty:
            continue

        ft_link_ids = [
            l
            for l in ft_links.index
            if all(
                hasattr(net.links_t, p) and (l in getattr(net.links_t, p).columns)
                for p in ["p0", "p1", "p2", "p3"]
            )
        ]
        if not ft_link_ids:
            continue

        timestep_hours = (
            (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600
            if len(net.snapshots) > 1
            else 1.0
        )

        for link in ft_link_ids:
            try:
                region = net.buses.at[ft_links.at[link, "bus1"], "grid_region"]
            except KeyError:
                continue
            if pd.isna(region):
                continue

            try:
                elec_price = net.buses_t.marginal_price[ft_links.at[link, "bus3"]]
                h2_price = net.buses_t.marginal_price[ft_links.at[link, "bus0"]]
                co2_price = net.buses_t.marginal_price[ft_links.at[link, "bus2"]]
            except KeyError:
                continue

            p1 = -net.links_t.p1[link] * timestep_hours  # product out (MWh)
            p3 = net.links_t.p3[link] * timestep_hours  # elec in
            p0 = net.links_t.p0[link] * timestep_hours  # H2 in
            p2 = (
                net.links_t.p2[link].clip(upper=0) * timestep_hours
            )  # CO2 (t, negative)

            prod_twh = p1.sum() / 1e6
            if prod_twh < min_production_twh:
                continue
            fuel_output_safe = p1.sum()
            if fuel_output_safe <= 0:
                continue

            elec_cost = (p3 * elec_price).sum() / fuel_output_safe
            h2_cost = (p0 * h2_price).sum() / fuel_output_safe
            co2_cost = (-p2 * co2_price).sum() / fuel_output_safe
            total_cost = elec_cost + h2_cost + co2_cost
            total_cost_gallon = total_cost * MWH_PER_GALLON

            all_rows.append(
                {
                    "Scenario": scenario,
                    "Year": year,
                    "Grid Region": region,
                    "Production (TWh)": prod_twh,
                    "Electricity cost (USD/MWh e-kerosene)": elec_cost,
                    "Hydrogen cost (USD/MWh e-kerosene)": h2_cost,
                    "CO2 cost (USD/MWh e-kerosene)": co2_cost,
                    "Total production cost (USD/MWh e-kerosene)": total_cost,
                    "Total cost (USD/gallon e-kerosene)": total_cost_gallon,
                }
            )

    if not (aggregate or wide):
        # unchanged legacy behavior (only for networks with data)
        if not all_rows:
            print("No e-kerosene production found.")
            return
        df_all = pd.DataFrame(all_rows)
        for (scen, yr), df_sub in df_all.groupby(["Scenario", "Year"]):

            def wavg(g, col):
                return (g[col] * g["Production (TWh)"]).sum() / g[
                    "Production (TWh)"
                ].sum()

            grouped = df_sub.groupby("Grid Region").apply(
                lambda g: pd.Series(
                    {
                        "Production (TWh)": g["Production (TWh)"].sum(),
                        **{c: wavg(g, c) for c in cost_cols},
                    }
                )
            )
            grouped = grouped[grouped["Production (TWh)"] >= min_production_twh]
            if grouped.empty:
                continue
            print(f"\n{yr if year_title else scen + '_' + str(yr)}:\n")
            display(
                grouped.round(2)
                .style.format(
                    {"Production (TWh)": "{:,.2f}", **{c: "{:,.2f}" for c in cost_cols}}
                )
                .hide(axis="index")
            )
            total_prod = grouped["Production (TWh)"].sum()
            if total_prod > 0:
                w_cost = (
                    grouped["Total production cost (USD/MWh e-kerosene)"]
                    * grouped["Production (TWh)"]
                ).sum() / total_prod
                w_cost_gallon = w_cost * MWH_PER_GALLON
                print(
                    f"Weighted average production cost: {w_cost:.2f} USD/MWh ({w_cost_gallon:.2f} USD/gallon)"
                )
        return

    if not all_rows:
        # build empty frame with expected combos
        base = []
        for scen, yr in product(expected_scenarios, expected_years):
            base.append(
                {
                    "Scenario": scen,
                    "Year": yr,
                    "Grid Region": None,
                    "Production (TWh)": 0.0,
                    **{
                        c: (
                            fill_cost_with
                            if fill_cost_with is not None
                            else float("nan")
                        )
                        for c in cost_cols
                    },
                }
            )
        df_empty = pd.DataFrame(base)
        return df_empty if not wide else pd.DataFrame()

    df_all = pd.DataFrame(all_rows)

    # weighted aggregation per Scenario-Year-Region
    def wavg_group(g, col):
        return (g[col] * g["Production (TWh)"]).sum() / g["Production (TWh)"].sum()

    grouped = (
        df_all.groupby(["Scenario", "Year", "Grid Region"])
        .apply(
            lambda g: pd.Series(
                {
                    "Production (TWh)": g["Production (TWh)"].sum(),
                    **{c: wavg_group(g, c) for c in cost_cols},
                }
            )
        )
        .reset_index()
    )

    # Collect all grid regions encountered
    grid_regions_all = sorted(grouped["Grid Region"].unique())

    # Insert missing scenario-year-region combinations
    existing_keys = set(zip(grouped.Scenario, grouped.Year, grouped["Grid Region"]))
    missing_rows = []
    for scen, yr, reg in product(expected_scenarios, expected_years, grid_regions_all):
        if (scen, yr, reg) not in existing_keys:
            missing_rows.append(
                {
                    "Scenario": scen,
                    "Year": yr,
                    "Grid Region": reg,
                    "Production (TWh)": 0.0,
                    **{
                        c: (
                            fill_cost_with
                            if fill_cost_with is not None
                            else float("nan")
                        )
                        for c in cost_cols
                    },
                }
            )
    if missing_rows:
        grouped = pd.concat([grouped, pd.DataFrame(missing_rows)], ignore_index=True)

    # Sort
    grouped = grouped.sort_values(["Scenario", "Year", "Grid Region"])

    if wide:
        # Pivot: index -> (Grid Region, Year) if multiple years
        multi = grouped.pivot_table(
            index=["Grid Region", "Year"],
            columns="Scenario",
            values=metrics_all,
            aggfunc="first",
        )

        # Reorder columns so scenarios in expected_scenarios order
        # Current columns: (metric, scenario) -> swap if needed
        if scenario_first_level:
            # we want (Scenario, Metric)
            multi = multi.swaplevel(0, 1, axis=1)

        # Ensure all scenarios present
        # Build full column MultiIndex
        if scenario_first_level:
            full_cols = pd.MultiIndex.from_product(
                [expected_scenarios, metrics_all], names=multi.columns.names
            )
        else:
            full_cols = pd.MultiIndex.from_product(
                [metrics_all, expected_scenarios], names=multi.columns.names
            )

        multi = multi.reindex(columns=full_cols)

        # If years missing, add them
        current_years = sorted({idx[1] for idx in multi.index})
        missing_years = [y for y in expected_years if y not in current_years]
        if missing_years:
            # create empty rows for each grid region
            grids = sorted({idx[0] for idx in multi.index})
            add_index = pd.MultiIndex.from_product(
                [grids, missing_years], names=multi.index.names
            )
            empty_df = pd.DataFrame(0.0, index=add_index, columns=multi.columns)
            # For cost columns set NaN (or fill value)
            if scenario_first_level:
                for scen in expected_scenarios:
                    for cost_col in cost_cols:
                        col = (scen, cost_col)
                        if fill_cost_with is None:
                            empty_df[col] = float("nan")
                        else:
                            empty_df[col] = fill_cost_with
            else:
                for cost_col in cost_cols:
                    for scen in expected_scenarios:
                        col = (cost_col, scen)
                        if fill_cost_with is None:
                            empty_df[col] = float("nan")
                        else:
                            empty_df[col] = fill_cost_with
            multi = pd.concat([multi, empty_df]).sort_index()

        # Drop Year level if single expected year
        if len(expected_years) == 1:
            multi.index = multi.index.droplevel("Year")

        return multi if not return_weighted_average else (multi, None)

    # Long form
    return grouped if not return_weighted_average else (grouped, None)


#### VALIDATION HELPERS FUNCTIONS #####


def convert_two_country_code_to_three(country_code):
    """
    Convert a two-letter country code to a three-letter ISO country code.

    Args:
        country_code (str): Two-letter country code (ISO 3166-1 alpha-2).

    Returns:
        str: Three-letter country code (ISO 3166-1 alpha-3).
    """
    country = pycountry.countries.get(alpha_2=country_code)
    return country.alpha_3


def get_country_name(country_code):
    """Input:
        country_code - two letter code of the country
    Output:
        country.name - corresponding name of the country
        country.alpha_3 - three letter code of the country
    """
    try:
        country = pycountry.countries.get(alpha_2=country_code)
        return country.name, country.alpha_3 if country else None
    except KeyError:
        return None


def get_data_EIA(data_path, country_code, year):
    """
    Retrieves energy generation data from the EIA dataset for a specified country and year.

    Args:
        data_path (str): Path to the EIA CSV file.
        country_code (str): Two-letter or three-letter country code (ISO).
        year (int or str): Year for which energy data is requested.

    Returns:
        pd.DataFrame: DataFrame containing energy generation data for the given country and year,
                    or None if no matching country is found.
    """

    # Load EIA data from CSV file
    data = pd.read_csv(data_path)

    # Rename the second column to 'country' for consistency
    data.rename(columns={"Unnamed: 1": "country"}, inplace=True)

    # Remove leading and trailing spaces in the 'country' column
    data["country"] = data["country"].str.strip()

    # Extract the three-letter country code from the 'API' column
    data["code_3"] = data.dropna(subset=["API"])["API"].apply(
        lambda x: x.split("-")[2] if isinstance(x, str) and len(x.split("-")) > 3 else x
    )

    # Get the official country name and three-letter country code using the provided two-letter code
    country_name, country_code3 = get_country_name(country_code)

    # Check if the three-letter country code exists in the dataset
    if country_code3 and country_code3 in data.code_3.unique():
        # Retrieve the generation data for the specified year
        result = data.query("code_3 == @country_code3")[["country", str(year)]]

    # If not found by code, search by the country name
    elif country_name and country_name in data.country.unique():
        # Find the country index and retrieve generation data
        country_index = data.query("country == @country_name").index[0]
        result = data.iloc[country_index + 1 : country_index + 18][
            ["country", str(year)]
        ]

    else:
        # If no match is found, return None
        result = None

    # Convert the year column to float for numeric operations
    result[str(year)] = result[str(year)].astype(float)

    return result


def get_demand_ember(data, country_code, year):
    """
    Get the electricity demand for a given country and year from Ember data.

    Args:
        data (pd.DataFrame): Ember data.
        country_code (str): Country code (ISO 3166-1 alpha-2).
        year (int): Year of interest.

    Returns:
        float or None: Electricity demand if found, otherwise None.
    """
    demand = data[
        (data["Year"] == year)
        & (data["Country code"] == country_code)
        & (data["Category"] == "Electricity demand")
        & (data["Subcategory"] == "Demand")
    ]["Value"]

    if len(demand) != 0:
        return demand.iloc[0]
    return None


def preprocess_eia_data_detail(data):
    """
    Preprocesses the EIA energy data by renaming and filtering rows and columns.

    Args:
        data (pd.DataFrame): DataFrame containing EIA energy data.

    Returns:
        pd.DataFrame: Cleaned and preprocessed DataFrame ready for analysis.
    """

    # Strip the last 13 characters (descriptive text) from the 'country' column
    data["country"] = data["country"].apply(lambda x: x[:-13].strip())

    # Set 'country' as the index of the DataFrame
    data.set_index("country", inplace=True)

    # Rename columns to provide clarity
    data.columns = ["EIA data"]

    # Rename specific rows to match more standard terms
    data.rename(
        index={
            "Hydroelectricity": "Hydro",
            "Biomass and waste": "Biomass",
            "Hydroelectric pumped storage": "PHS",
        },
        inplace=True,
    )

    # Drop unwanted renewable energy categories
    data.drop(
        index=[
            "Fossil fuels",
            "Renewables",
            "Non-hydroelectric renewables",
            "Solar, tide, wave, fuel cell",
            "Tide and wave",
        ],
        inplace=True,
    )

    # Filter the DataFrame to only include relevant energy sources
    data = data.loc[
        [
            "Nuclear",
            "Coal",
            "Natural gas",
            "Oil",
            "Geothermal",
            "Hydro",
            "PHS",
            "Solar",
            "Wind",
            "Biomass",
        ],
        :,
    ]
    return data


def get_generation_capacity_ember_detail(data, three_country_code, year):
    """
    Get electricity generation by fuel type for a given country and year from Ember data.

    Args:
        data (pd.DataFrame): Ember data.
        three_country_code (str): Country code (ISO 3166-1 alpha-3).
        year (int): Year of interest.

    Returns:
        pd.DataFrame: Electricity generation by fuel type.
    """
    generation_ember = data[
        (data["Category"] == "Electricity generation")
        & (data["Country code"] == three_country_code)
        & (data["Year"] == year)
        & (data["Subcategory"] == "Fuel")
        & (data["Unit"] == "TWh")
    ][["Variable", "Value"]].reset_index(drop=True)

    # Drop irrelevant rows
    drop_row = ["Other Renewables"]
    generation_ember = generation_ember[~generation_ember["Variable"].isin(drop_row)]

    # Standardize fuel types
    generation_ember = generation_ember.replace(
        {
            "Gas": "Natural gas",
            "Bioenergy": "Biomass",
            # "Coal": "Fossil fuels",
            # "Other Fossil": "Fossil fuels"
        }
    )

    # Group by fuel type
    generation_ember = generation_ember.groupby("Variable").sum()
    generation_ember.loc["Load shedding"] = 0.0
    generation_ember.columns = ["Ember data"]

    return generation_ember


def get_installed_capacity_ember(data, three_country_code, year):
    """
    Get installed capacity by fuel type for a given country and year from Ember data.

    Args:
        data (pd.DataFrame): Ember data.
        three_country_code (str): Country code (ISO 3166-1 alpha-3).
        year (int): Year of interest.

    Returns:
        pd.DataFrame: Installed capacity by fuel type.
    """
    capacity_ember = data[
        (data["Country code"] == three_country_code)
        & (data["Year"] == year)
        & (data["Category"] == "Capacity")
        & (data["Subcategory"] == "Fuel")
    ][["Variable", "Value"]].reset_index(drop=True)

    # Drop irrelevant rows
    drop_row = ["Other Renewables"]
    capacity_ember = capacity_ember[~capacity_ember["Variable"].isin(drop_row)]

    # Standardize fuel types
    capacity_ember = capacity_ember.replace(
        {
            # "Gas": "Fossil fuels",
            "Bioenergy": "Biomass",
            # "Coal": "Fossil fuels",
            "Other Fossil": "Fossil fuels",
        }
    )

    capacity_ember = capacity_ember.groupby("Variable").sum()
    capacity_ember.columns = ["Ember data"]

    return capacity_ember


def preprocess_eia_data(data):
    """
    Preprocesses the EIA energy data by renaming and filtering rows and columns.

    Args:
        data (pd.DataFrame): DataFrame containing EIA energy data.

    Returns:
        pd.DataFrame: Cleaned and preprocessed DataFrame ready for analysis.
    """

    # Strip the last 13 characters (descriptive text) from the 'country' column
    data["country"] = data["country"].apply(lambda x: x[:-13].strip())

    # Set 'country' as the index of the DataFrame
    data.set_index("country", inplace=True)

    # Rename columns to provide clarity
    data.columns = ["EIA data"]

    # Rename specific rows to match more standard terms
    data.rename(
        index={
            "Hydroelectricity": "Hydro",
            "Biomass and waste": "Biomass",
            "Hydroelectric pumped storage": "PHS",
        },
        inplace=True,
    )

    # Drop unwanted renewable energy categories
    data.drop(
        index=[
            "Renewables",
            "Non-hydroelectric renewables",
            "Solar, tide, wave, fuel cell",
            "Tide and wave",
        ],
        inplace=True,
    )

    # Filter the DataFrame to only include relevant energy sources
    data = data.loc[
        [
            "Nuclear",
            "Fossil fuels",
            "Geothermal",
            "Hydro",
            "PHS",
            "Solar",
            "Wind",
            "Biomass",
        ],
        :,
    ]

    return data


def get_demand_pypsa(network):
    """
    Get the total electricity demand from the PyPSA-Earth network.

    Args:
        network (pypsa.Network): PyPSA network object.

    Returns:
        float: Total electricity demand in TWh.
    """
    demand_pypsa = (
        network.loads_t.p_set.multiply(network.snapshot_weightings.objective, axis=0)
        .sum()
        .sum()
        / 1e6
    )
    demand_pypsa = demand_pypsa.round(4)
    return demand_pypsa


def preprocess_eia_demand(path, horizon):
    statewise_df = pd.read_excel(path, sheet_name="Data")

    demand_df = statewise_df.loc[statewise_df["MSN"] == "ESTXP"]
    demand_df.set_index("State", inplace=True)

    # data is in million kWh (GWh) - hence dividing by 1e3 to get the data in TWh
    demand_df = demand_df[int(horizon)] / 1e3
    demand_df = demand_df.to_frame()
    demand_df.columns = ["EIA"]

    demand_df.drop(["US"], axis=0, inplace=True)
    return demand_df


def plot_stacked_costs_by_year_plotly(
    cost_data, cost_type_label, tech_colors=None, index="year"
):
    # Filter data
    data_filtered = cost_data[
        (cost_data["cost_type"] == cost_type_label) & (cost_data["cost_billion"] != 0)
    ].copy()

    if data_filtered.empty:
        print("No data to plot.")
        return

    # Pivot table: index x tech_label
    pivot_table = data_filtered.pivot_table(
        index=index, columns="tech_label", values="cost_billion", aggfunc="sum"
    ).fillna(0)

    # Mapping: tech_label → macro category / main category
    label_to_macro = data_filtered.set_index("tech_label")["macro_category"].to_dict()
    label_to_category = data_filtered.set_index("tech_label")["main_category"].to_dict()

    # Desired macro-category order
    desired_macro_order = [
        "Hydrogen & e-fuels",
        "Biofuels synthesis",
        "DAC",
        "End-uses",
        "Industry",
        "Power & heat generation",
        "Storage",
        "Transmission & distribution",
        "Emissions",
        "Other",
    ]
    macro_order_map = {macro: i for i, macro in enumerate(desired_macro_order)}

    # Sort tech labels by macro_category + appearance order
    all_labels = data_filtered["tech_label"].drop_duplicates().tolist()
    ordered_labels = sorted(
        all_labels,
        key=lambda lbl: (
            macro_order_map.get(label_to_macro.get(lbl, "Other"), 999),
            all_labels.index(lbl),
        ),
    )

    # Reorder pivot table
    pivot_table = pivot_table[ordered_labels[::-1]]  # reverse for stacking

    # Assign colors
    def get_color(label):
        category = label_to_category.get(label, label)
        return tech_colors.get(category, "#999999") if tech_colors else "#999999"

    color_values = {label: get_color(label) for label in pivot_table.columns}

    # Create Plotly figure
    fig = go.Figure()

    x_vals = pivot_table.index.astype(str)

    # One trace per tech — works with negative values + interactive legend
    for label in pivot_table.columns:
        y_series = pivot_table[label]
        fig.add_trace(
            go.Bar(
                x=x_vals,
                y=y_series,
                name=label,
                marker=dict(color=color_values[label]),
                hovertemplate=f"%{{x}}<br>{label}: %{{y:.2f}}B USD<extra></extra>",
            )
        )

    # Macro-category legend block (annotation)
    grouped_labels = defaultdict(list)
    for label in ordered_labels:
        macro = label_to_macro.get(label, "Other")
        grouped_labels[macro].append(label)

    legend_text = ""
    for macro in desired_macro_order:
        if macro in grouped_labels:
            legend_text += f"<b>{macro}</b><br>"
            for label in grouped_labels[macro]:
                color = color_values[label]
                legend_text += f"<span style='color:{color}'>▇</span> {label}<br>"

    # Add annotation for grouped legend
    fig.add_annotation(
        text=legend_text,
        showarrow=False,
        align="left",
        xref="paper",
        yref="paper",
        x=1.25,
        y=1,
        bordercolor="black",
        borderwidth=1,
        bgcolor="rgba(255,255,255,0.95)",
        font=dict(size=14),
    )

    # Add 0-line for clarity
    fig.add_shape(
        type="line",
        xref="paper",
        x0=0,
        x1=1,
        yref="y",
        y0=0,
        y1=0,
        line=dict(color="black", width=1),
    )

    fig.update_layout(
        barmode="relative",
        title=dict(
            text=f"{cost_type_label} - Total system costs",
            font=dict(size=16),  # Titolo del grafico
        ),
        xaxis=dict(
            title=dict(text="Years (-)", font=dict(size=12)), tickfont=dict(size=12)
        ),
        yaxis=dict(
            title=dict(text=f"{cost_type_label} (Billion USD)", font=dict(size=12)),
            tickfont=dict(size=12),
        ),
        template="plotly_white",
        width=1400,
        height=700,
        margin=dict(l=40, r=300, t=50, b=50),
        legend_title=dict(text="Technologies", font=dict(size=14)),
        legend=dict(font=dict(size=12), traceorder="reversed"),
        showlegend=False,
    )

    fig.show()


def plot_float_bar_lcoe_dispatch_ranges(
    table_df, key, nice_names, use_scenario_names=False
):
    # Extract year from the key using regex
    year_match = re.search(r"\d{4}", key)
    year_str = year_match.group() if year_match else "Year N/A"

    carrier_list = [
        "CCGT lcoe (USD/MWh)",
        "OCGT lcoe (USD/MWh)",
        "coal lcoe (USD/MWh)",
        "nuclear lcoe (USD/MWh)",
        "oil lcoe (USD/MWh)",
        "urban central gas CHP lcoe (USD/MWh)",
        "urban central solid biomass CHP lcoe (USD/MWh)",
        "biomass lcoe (USD/MWh)",
        "geothermal lcoe (USD/MWh)",
        "hydro lcoe (USD/MWh)",
        "onwind lcoe (USD/MWh)",
        "ror lcoe (USD/MWh)",
        "solar lcoe (USD/MWh)",
        "solar rooftop lcoe (USD/MWh)",
    ]

    buffer_left = 100
    buffer_right = 20

    global_min = table_df.xs("min", axis=1, level=1).min().min()
    global_max = table_df.xs("max", axis=1, level=1).max().max()

    x_min = min(-50, global_min - buffer_left)
    x_max = global_max + buffer_right

    regions = table_df.index.tolist()
    n_regions = len(regions)

    # Subplot grid size (2 columns)
    ncols = 2
    nrows = math.ceil(n_regions / ncols)

    fig, axs = plt.subplots(
        nrows=nrows, ncols=ncols, figsize=(16, nrows * 5), constrained_layout=True
    )
    axs = axs.flatten()

    for idx, region in enumerate(regions):
        ax = axs[idx]

        # Filter only available carriers
        available_carriers = [
            c for c in carrier_list if c in table_df.columns.get_level_values(0)
        ]
        if not available_carriers:
            ax.set_title(f"{region} - No carriers available", fontsize=12)
            ax.axis("off")
            continue

        table_lcoe_df = table_df[
            table_df.columns[table_df.columns.get_level_values(0).str.contains("lcoe")]
        ][available_carriers]

        table_lcoe_df_region = table_lcoe_df.loc[region, :]

        lcoe_tech_list = table_lcoe_df_region.xs("max", level=1).index

        for i, (start, end) in enumerate(
            zip(
                table_lcoe_df_region.xs("min", level=1).values,
                table_lcoe_df_region.xs("max", level=1).values,
            )
        ):
            str_attach = any(np.abs([start, end]) > 1e-3)
            width = end - start
            ax.broken_barh(
                [(start, width)], (i - 0.4, 0.8), hatch="///", edgecolor="white"
            )
            start_label = f"${round(start, 2)}" if str_attach else ""
            end_label = f"${round(start + width, 2)}" if str_attach else ""
            ax.text(start - 0.7, i, start_label, va="center", ha="right", fontsize=9)
            ax.text(
                start + width + 0.7, i, end_label, va="center", ha="left", fontsize=9
            )

        raw_labels = [
            label.replace(" lcoe", "").replace(" (USD/MWh)", "")
            for label in lcoe_tech_list
        ]
        clean_labels = [nice_names.get(lbl, lbl) for lbl in raw_labels]

        ax.set_yticks(range(len(lcoe_tech_list)))
        ax.set_yticklabels(clean_labels, fontsize=10)
        ax.set_xlabel("LCOE (USD/MWh)", fontsize=10)
        ax.set_xlim(x_min, x_max)
        ax.set_title(
            f"\n{region} - {key if use_scenario_names else year_str}", fontsize=12
        )
        ax.grid(linestyle="--", alpha=0.5)
        ax.tick_params(axis="both", labelsize=9)

    # Hide any unused axes
    for j in range(idx + 1, len(axs)):
        fig.delaxes(axs[j])

    showfig()


def compute_line_expansion_capacity(n):
    """
    Computes the line expansion capacity grouped by grid region.

    Parameters:
    n (pypsa.Network): The PyPSA network object.

    Returns:
    pd.DataFrame: DataFrame with line expansion capacity by grid region.
    """
    # Ensure 'grid_region' is present in lines
    if "grid_region" not in n.lines.columns:
        n.lines["grid_region"] = n.lines.bus0.map(n.buses.grid_region)

    # Ensure 'state' is present in lines
    if "state" not in n.lines.columns:
        n.lines["state"] = n.lines.bus0.map(n.buses.state)

    # Group by 'grid_region' and 'state' then sum the capacities
    line_exp_cap_grid = (
        n.lines.groupby("grid_region")[["s_nom", "s_nom_opt"]].sum() / 1e3
    )  # Convert to GW
    line_exp_cap_state = (
        n.lines.groupby("state")[["s_nom", "s_nom_opt"]].sum() / 1e3
    )  # Convert to GW

    return line_exp_cap_grid, line_exp_cap_state


def preprocess_res_ces_share_eia(eia_gen_data):
    eia_gen_data = eia_gen_data[eia_gen_data["YEAR"] == 2023]
    eia_gen_data = eia_gen_data[eia_gen_data["STATE"] != "US-Total"]
    eia_gen_data = eia_gen_data[
        eia_gen_data["TYPE OF PRODUCER"] == "Total Electric Power Industry"
    ]
    eia_gen_data = eia_gen_data[
        (eia_gen_data["ENERGY SOURCE"] != "Total")
        & (eia_gen_data["ENERGY SOURCE"] != "Other")
    ]
    eia_gen_data.replace(
        {
            "ENERGY SOURCE": {
                "Coal": "coal",
                "Hydroelectric Conventional": "hydro",
                "Pumped Storage": "PHS",
                "Solar Thermal and Photovoltaic": "solar",
                "Natural Gas": "gas",
                "Petroleum": "oil",
                "Wind": "wind",
                "Nuclear": "nuclear",
                "Geothermal": "geothermal",
                "Pumped Storage": "PHS",
                "Wood and Wood Derived Fuels": "biomass",
                "Other Biomass": "biomass",
                "Other Gases": "gas",
            }
        },
        inplace=True,
    )

    eia_gen_data["GENERATION (TWh)"] = eia_gen_data["GENERATION (Megawatthours)"] / 1e6
    eia_gen_data_df = (
        eia_gen_data.groupby(["STATE", "ENERGY SOURCE"])[["GENERATION (TWh)"]]
        .sum()
        .unstack(fill_value=0)
    )
    eia_gen_data_df.columns = eia_gen_data_df.columns.droplevel(0)

    eia_res_carriers = ["solar", "wind", "hydro", "geothermal", "biomass", "PHS"]
    eia_ces_carriers = eia_res_carriers + ["nuclear"]

    res_total = eia_gen_data_df[eia_res_carriers].sum(axis=1)
    ces_total = eia_gen_data_df[eia_ces_carriers].sum(axis=1)
    all_total = eia_gen_data_df[
        [
            "PHS",
            "biomass",
            "coal",
            "gas",
            "geothermal",
            "hydro",
            "nuclear",
            "oil",
            "solar",
            "wind",
        ]
    ].sum(axis=1)

    eia_gen_data_df["% Actual RES"] = (res_total / all_total) * 100
    eia_gen_data_df["% Actual CES"] = (ces_total / all_total) * 100

    return eia_gen_data_df


def compute_links_only_costs(network, name_tag):
    """
    Compute costs for power generation:
    - Links (incl. fossil power links)
    - Non-fossil generators
    - Battery storage (Store) + battery power components (Links)

    Uses network.statistics(groupby=None) to stay consistent
    with the rest of the plotting pipeline.
    """
    year_str = name_tag[-4:]
    fossil_carriers = ["coal", "gas", "oil", "biomass"]

    costs_detailed = network.statistics(groupby=None)[
        ["Capital Expenditure", "Operational Expenditure"]
    ]

    final_results = []

    # -------------------------------------------------
    # 1. LINKS (incl. battery charger/discharger)
    # -------------------------------------------------
    try:
        link_costs = costs_detailed.loc["Link"].reset_index()
        link_costs["tech_label"] = link_costs["carrier"]

        # CAPEX
        for tech, sub in link_costs.groupby("tech_label"):
            capex = sub["Capital Expenditure"].sum()
            if capex != 0:
                final_results.append(
                    {
                        "tech_label": tech,
                        "cost_type": "Capital expenditure",
                        "cost_billion": capex / 1e9,
                        "year": year_str,
                        "scenario": name_tag,
                    }
                )

        # OPEX
        for tech, sub in link_costs.groupby("tech_label"):
            opex = sub["Operational Expenditure"].sum()
            if opex != 0:
                final_results.append(
                    {
                        "tech_label": tech,
                        "cost_type": "Operational expenditure",
                        "cost_billion": opex / 1e9,
                        "year": year_str,
                        "scenario": name_tag,
                    }
                )
    except KeyError:
        pass

    # -------------------------------------------------
    # 2. NON-FOSSIL GENERATORS
    # -------------------------------------------------
    try:
        gen_costs = costs_detailed.loc["Generator"].reset_index()
        gen_costs["tech_label"] = gen_costs["carrier"]

        non_fossil = gen_costs[~gen_costs["tech_label"].isin(fossil_carriers)]

        for tech, sub in non_fossil.groupby("tech_label"):
            capex = sub["Capital Expenditure"].sum()
            if capex != 0:
                final_results.append(
                    {
                        "tech_label": tech,
                        "cost_type": "Capital expenditure",
                        "cost_billion": capex / 1e9,
                        "year": year_str,
                        "scenario": name_tag,
                    }
                )

            opex = sub["Operational Expenditure"].sum()
            if opex != 0:
                final_results.append(
                    {
                        "tech_label": tech,
                        "cost_type": "Operational expenditure",
                        "cost_billion": opex / 1e9,
                        "year": year_str,
                        "scenario": name_tag,
                    }
                )
    except KeyError:
        pass

    # -------------------------------------------------
    # 3. BATTERY STORAGE (Store)  <<< QUESTO È IL FIX
    # -------------------------------------------------
    try:
        store_costs = costs_detailed.loc["Store"].reset_index()

        # prendiamo SOLO battery
        store_costs = store_costs[store_costs["carrier"] == "battery"]

        capex = store_costs["Capital Expenditure"].sum()
        if capex != 0:
            final_results.append(
                {
                    "tech_label": "battery",  # <-- fondamentale
                    "cost_type": "Capital expenditure",
                    "cost_billion": capex / 1e9,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

        opex = store_costs["Operational Expenditure"].sum()
        if opex != 0:
            final_results.append(
                {
                    "tech_label": "battery",
                    "cost_type": "Operational expenditure",
                    "cost_billion": opex / 1e9,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )
    except KeyError:
        pass

    # -------------------------------------------------
    # 4. FUEL COST ADJUSTMENT FOR FOSSIL LINKS (COME PRIMA)
    # -------------------------------------------------
    fuel_cost_adjustments = {}

    for carrier in fossil_carriers:
        links = network.links[network.links.carrier == carrier]
        total_fuel_cost = 0.0

        for link_id in links.index:
            try:
                p0 = network.links_t.p0[link_id]
                fuel_bus = links.loc[link_id, "bus0"]
                fuel_price = network.buses_t.marginal_price[fuel_bus]
                weightings = network.snapshot_weightings["objective"]

                fuel_cost = (p0 * fuel_price * weightings).sum()
                total_fuel_cost += fuel_cost
            except KeyError:
                continue

        if total_fuel_cost > 0:
            fuel_cost_adjustments[carrier] = total_fuel_cost / 1e9

    df_results = pd.DataFrame(final_results)

    for carrier in fossil_carriers:
        if carrier in fuel_cost_adjustments:
            mask = (df_results["tech_label"] == carrier) & (
                df_results["cost_type"] == "Operational expenditure"
            )
            if mask.any():
                df_results.loc[mask, "cost_billion"] += fuel_cost_adjustments[carrier]
                df_results.loc[mask, "tech_label"] = f"{carrier} (power)"

    return df_results


def identify_power_generation_technologies(
    rename_techs_capex, rename_techs_opex, categories_capex, categories_opex
):
    """
    Identify technologies for power generation only (including (power) versions of conventional fuels)
    """
    power_gen_techs = set()

    # Check CAPEX mappings
    for original_tech, intermediate_category in rename_techs_capex.items():
        if categories_capex.get(intermediate_category) == "Power & heat generation":
            if intermediate_category != "Heating":  # Exclude heating
                # Convert conventional fuels to (power) format
                if original_tech in ["coal", "gas", "oil", "biomass"]:
                    power_gen_techs.add(f"{original_tech} (power)")
                else:
                    power_gen_techs.add(original_tech)

    # Check OPEX mappings
    for original_tech, intermediate_category in rename_techs_opex.items():
        if categories_opex.get(intermediate_category) == "Power & heat generation":
            if intermediate_category != "Heating":  # Exclude heating
                # Convert conventional fuels to (power) format
                if original_tech in ["coal", "gas", "oil", "biomass"]:
                    power_gen_techs.add(f"{original_tech} (power)")
                else:
                    power_gen_techs.add(original_tech)

    return power_gen_techs


def plot_power_generation_details(
    cost_data,
    cost_type_label,
    power_techs,
    tech_colors=None,
    nice_names=None,
    tech_order=None,
    index="year",
):
    """
    Plot interactive detailed breakdown of Power & heat generation technologies showing original tech_labels

    Parameters:
    - cost_data: DataFrame with cost data containing original tech_labels
    - cost_type_label: str, "Capital expenditure" or "Operational expenditure"
    - power_techs: set of technology names that belong to Power & heat generation
    - tech_colors: dict, mapping from original tech_labels to colors
    - nice_names: dict, mapping from original tech_labels to display names
    """

    # Filter for Power & heat generation technologies
    power_data = cost_data[
        cost_data["tech_label"].isin(power_techs)
        & (cost_data["cost_type"] == cost_type_label)
        & (cost_data["cost_billion"] != 0)
    ].copy()

    # Aggregate technologies with the same name (e.g., Offshore Wind AC + DC)
    power_data = power_data.groupby(
        ["tech_label", "cost_type", "year", "scenario"], as_index=False
    ).agg({"cost_billion": "sum"})

    if power_data.empty:
        return

    # Create pivot table: years x technologies
    pivot_table = power_data.pivot_table(
        index=index, columns="tech_label", values="cost_billion", aggfunc="sum"
    ).fillna(0)

    # Sort technologies by total cost (largest first)
    if tech_order:
        available_techs = set(pivot_table.columns)
        ordered_techs = [tech for tech in tech_order if tech in available_techs]
        remaining_techs = available_techs - set(ordered_techs)
        ordered_techs.extend(sorted(remaining_techs))
    else:
        tech_totals = pivot_table.sum().sort_values(ascending=False)
        ordered_techs = tech_totals.index.tolist()

    # Get colors for each technology
    def get_tech_color(original_tech_label):
        if tech_colors and original_tech_label in tech_colors:
            return tech_colors[original_tech_label]
        # Try some variations if exact match not found
        elif tech_colors:
            for variant in [original_tech_label.lower(), original_tech_label.title()]:
                if variant in tech_colors:
                    return tech_colors[variant]
        # Default color if not found
        import matplotlib.colors as mcolors

        colors = list(mcolors.TABLEAU_COLORS.values())
        return colors[hash(original_tech_label) % len(colors)]

    # Create interactive plotly chart
    import plotly.graph_objects as go

    fig = go.Figure()

    # Add traces for each technology (in order of importance)
    # Only exclude technologies that are ALL zeros
    for tech in ordered_techs:
        y_values = pivot_table[tech]
        # Skip only if ALL values are exactly zero
        if (y_values == 0).all():
            continue

        display_name = nice_names.get(tech, tech) if nice_names else tech
        color = get_tech_color(tech)

        fig.add_trace(
            go.Bar(
                name=display_name,
                x=pivot_table.index.astype(str),
                y=y_values,
                marker_color=color,
                hovertemplate=f"%{{x}}<br>{display_name}: %{{y:.2f}}B USD<extra></extra>",
            )
        )

    # Update layout for interactivity
    fig.update_layout(
        barmode="relative",  # Handle negative values correctly
        title=dict(text=f"Power Generation - {cost_type_label}", font=dict(size=16)),
        xaxis=dict(
            title=dict(text="Years", font=dict(size=14)), tickfont=dict(size=12)
        ),
        yaxis=dict(
            title=dict(text=f"{cost_type_label} (Billion USD)", font=dict(size=12)),
            tickfont=dict(size=12),
        ),
        template="plotly_white",
        width=1400,
        height=700,
        margin=dict(l=40, r=300, t=50, b=50),
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02,
            font=dict(size=12),
            traceorder="reversed",
        ),
    )

    # Add horizontal line at zero
    fig.add_hline(y=0, line_width=1, line_color="black")

    return fig


def compute_h2_efuels_costs(network, name_tag):
    """
    Compute costs for H2 and e-fuels technologies (links only).
    Returns values in million USD.
    """
    year_str = name_tag[-4:]
    carriers = [
        "Alkaline electrolyzer large",
        "Fischer-Tropsch",
        "PEM electrolyzer",
        "SOEC",
    ]

    # Extract component-level statistics
    stats = network.statistics()[["Capital Expenditure", "Operational Expenditure"]]
    df = stats.reset_index()
    if df.columns[0] != "component":
        df.rename(columns={df.columns[0]: "component"}, inplace=True)

    # Select only link components of interest
    df_links = df[(df["component"] == "Link") & (df["carrier"].isin(carriers))]

    records = []
    if not df_links.empty:
        # CAPEX
        cap = df_links.groupby("carrier", as_index=False)["Capital Expenditure"].sum()
        for _, r in cap.iterrows():
            records.append(
                {
                    "tech_label": r["carrier"],
                    "cost_type": "Capital expenditure",
                    "cost_million": r["Capital Expenditure"] / 1e6,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

        # OPEX
        opx = df_links.groupby("carrier", as_index=False)[
            "Operational Expenditure"
        ].sum()
        for _, r in opx.iterrows():
            records.append(
                {
                    "tech_label": r["carrier"],
                    "cost_type": "Operational expenditure",
                    "cost_million": r["Operational Expenditure"] / 1e6,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

    return pd.DataFrame(records)


def calculate_lcoh_by_region(
    networks,
    h2_carriers,
    regional_fees,
    emm_mapping,
    output_threshold=1.0,
    year_title=True,
    electricity_price="marginal",
    grid_region_lcoe=None,
    include_baseload=False,
    baseload_charge_path="./data/energy_charge_rate.csv",
    customer_charge_mw=400.0,
    demand_charge_rate=9.0,
    baseload_percentages=None,
):
    """
    Compute weighted average LCOH by grid region and year.
    All costs are returned in USD/kg H2 (LHV = 33 kWh/kg).
    """

    results = {}

    # -----------------------------
    # Conversion (LHV)
    # -----------------------------
    conv = 1000.0 / 33.0  # kg H2 per MWh H2
    suffix = "USD/kg H2"

    # -----------------------------
    # Baseload charges (optional)
    # -----------------------------
    baseload_charges = {}
    if include_baseload:
        baseload_charges = calculate_baseload_charge(
            networks=networks,
            h2_carriers=h2_carriers,
            emm_mapping=emm_mapping,
            energy_charge_path=baseload_charge_path,
            customer_charge_mw=customer_charge_mw,
            demand_charge_rate=demand_charge_rate,
            baseload_percentages=baseload_percentages,
            output_threshold=output_threshold,
            verbose=False,
        )

    # -----------------------------
    # Loop over networks
    # -----------------------------
    for year_key, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(year_key)).group())

        links = net.links[net.links.carrier.isin(h2_carriers)]
        if links.empty:
            continue

        # Flows
        p0 = net.links_t.p0[links.index]  # electricity input
        p1 = net.links_t.p1[links.index]  # H2 output (negative)
        w = net.snapshot_weightings.generators

        cons = p0.clip(lower=0).multiply(w, axis=0)  # MWh_el
        h2 = (-p1).clip(lower=0).multiply(w, axis=0)  # MWh_H2
        h2_out = h2.sum()

        valid = h2_out > output_threshold
        if valid.sum() == 0:
            continue

        out_valid = h2_out[valid]

        # -----------------------------
        # CAPEX / OPEX  (USD/MWh_H2 → USD/kg)
        # -----------------------------
        capex = links.loc[valid, "capital_cost"] * links.loc[valid, "p_nom_opt"]
        opex = links.loc[valid, "marginal_cost"] * cons.loc[:, valid].sum(axis=0)

        capex_val = capex / out_valid / conv
        opex_val = opex / out_valid / conv

        # -----------------------------
        # Electricity cost
        # -----------------------------
        elec_cost = pd.Series(0.0, index=valid.index[valid])

        if electricity_price == "marginal":
            for l in valid.index[valid]:
                bus = links.at[l, "bus0"]
                elec_cost[l] = (cons[l] * net.buses_t.marginal_price[bus]).sum()

        elif electricity_price == "LCOE":
            if grid_region_lcoe is None:
                raise ValueError(
                    "grid_region_lcoe must be provided when electricity_price='LCOE'"
                )
            for l in valid.index[valid]:
                bus = links.at[l, "bus0"]
                region = net.buses.at[bus, "grid_region"]
                elec_cost[l] = cons[l].sum() * grid_region_lcoe.get(region, np.nan)

        else:
            raise ValueError("electricity_price must be 'marginal' or 'LCOE'")

        elec_val = elec_cost / out_valid / conv

        # -----------------------------
        # Assemble dataframe
        # -----------------------------
        df = pd.DataFrame(
            {
                f"Electrolysis CAPEX ({suffix})": capex_val,
                f"Electrolysis OPEX ({suffix})": opex_val,
                f"Electricity ({suffix})": elec_val,
                "h2_out": out_valid,  # MWh H2
                "bus": links.loc[valid, "bus0"],
            }
        )

        df["grid_region"] = df["bus"].map(net.buses["grid_region"])

        # -----------------------------
        # Transmission & Distribution fees
        # -----------------------------
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year,
            ["region", "Transmission nom USD/MWh", "Distribution nom USD/MWh"],
        ].set_index("region")

        df["EMM"] = df["grid_region"].map(emm_mapping)

        fee_trans = df["EMM"].map(fee_map["Transmission nom USD/MWh"])
        fee_dist = df["EMM"].map(fee_map["Distribution nom USD/MWh"])

        elec_rate = cons.loc[:, valid].sum(axis=0) / out_valid  # MWh_el / MWh_H2

        fee_trans_val = (fee_trans * elec_rate) / conv
        fee_dist_val = (fee_dist * elec_rate) / conv

        # -----------------------------
        # LCOH components
        # -----------------------------
        df[f"LCOH (excl. T&D fees) ({suffix})"] = (
            df[f"Electrolysis CAPEX ({suffix})"]
            + df[f"Electrolysis OPEX ({suffix})"]
            + df[f"Electricity ({suffix})"]
        )

        df[f"LCOH + Transmission fees ({suffix})"] = (
            df[f"LCOH (excl. T&D fees) ({suffix})"] + fee_trans_val
        )

        df[f"LCOH + T&D fees ({suffix})"] = (
            df[f"LCOH + Transmission fees ({suffix})"] + fee_dist_val
        )

        # -----------------------------
        # Baseload correction
        # -----------------------------
        if include_baseload and scen_year in baseload_charges:
            baseload_df = baseload_charges[scen_year]

            baseload_frac = baseload_df.set_index("grid_region")["baseload_pct"] / 100.0
            baseload_cost = baseload_df.set_index("grid_region")[
                "baseload_cost_per_mwh_h2"
            ]

            activity_factor = {}
            for l in valid.index[valid]:
                p_prod = (-net.links_t.p1[l]).clip(lower=0)
                activity_factor[l] = (p_prod > 0).sum() / len(p_prod)

            df["activity_factor"] = df.index.map(activity_factor)
            df["baseload_frac"] = df["grid_region"].map(baseload_frac).fillna(0)

            df[f"Electricity ({suffix})"] *= (
                1 - df["baseload_frac"] * df["activity_factor"]
            )
            df[f"Baseload charges ({suffix})"] = (
                df["grid_region"].map(baseload_cost).fillna(0) / conv
            )

            df[f"LCOH + Transmission fees + Baseload charges ({suffix})"] = (
                df[f"LCOH + Transmission fees ({suffix})"]
                + df[f"Baseload charges ({suffix})"]
            )

            df[f"LCOH + T&D fees + Baseload charges ({suffix})"] = (
                df[f"LCOH + T&D fees ({suffix})"] + df[f"Baseload charges ({suffix})"]
            )

        else:
            df[f"Baseload charges ({suffix})"] = 0.0
            df[f"LCOH + Transmission fees + Baseload charges ({suffix})"] = df[
                f"LCOH + Transmission fees ({suffix})"
            ]
            df[f"LCOH + T&D fees + Baseload charges ({suffix})"] = df[
                f"LCOH + T&D fees ({suffix})"
            ]

        # -----------------------------
        # Aggregate by region
        # -----------------------------
        region_summary = (
            df.groupby("grid_region")
            .apply(
                lambda g: pd.Series(
                    {
                        c: (g[c] * g["h2_out"]).sum() / g["h2_out"].sum()
                        for c in g.columns
                        if suffix in c
                    }
                )
            )
            .reset_index()
            .rename(columns={"grid_region": "Grid Region"})
        )

        # Dispatch
        region_summary["Hydrogen Dispatch (GWh per region)"] = (
            df.groupby("grid_region")["h2_out"].sum() / 1000.0
        ).values

        region_summary["Hydrogen Dispatch (tons per region)"] = (
            df.groupby("grid_region")["h2_out"].sum() / conv
        ).values

        results[str(scen_year) if year_title else year_key] = region_summary.round(2)

    return results


def plot_h2_efuels_details(
    cost_data, cost_type_label, tech_colors=None, tech_order=None, index="year"
):
    """
    Plot interactive detailed breakdown of H2 and e-fuels technologies
    """

    # Filter for H2/e-fuels technologies
    h2_efuels_data = cost_data[
        (cost_data["cost_type"] == cost_type_label)
        &
        # Only values > 1 million USD
        (cost_data["cost_billion"].abs() > 0.001)
    ].copy()

    if h2_efuels_data.empty:
        print(f"No data for {cost_type_label}")
        return

    # Create pivot table: years x technologies
    pivot_table = h2_efuels_data.pivot_table(
        index=index, columns="tech_label", values="cost_billion", aggfunc="sum"
    ).fillna(0)

    # Order technologies
    if tech_order:
        available_techs = set(pivot_table.columns)
        ordered_techs = [tech for tech in tech_order if tech in available_techs]
        remaining_techs = available_techs - set(ordered_techs)
        ordered_techs.extend(sorted(remaining_techs))
    else:
        tech_totals = pivot_table.sum().sort_values(ascending=False)
        ordered_techs = tech_totals.index.tolist()

    # Get colors for each technology
    def get_tech_color(tech_label):
        if tech_colors and tech_label in tech_colors:
            return tech_colors[tech_label]
        # Default colors matching your figure + magenta for Fischer-Tropsch
        default_colors = {
            "Alkaline electrolyzer large": "#1f77b4",  # Blue
            "PEM electrolyzer": "#2ca02c",  # Green
            "SOEC": "#d62728",  # Red
            "Fischer-Tropsch": "#e81cd0",  # Magenta
        }
        return default_colors.get(tech_label, "#999999")

    # Create interactive plotly chart
    import plotly.graph_objects as go

    fig = go.Figure()

    # Add traces for each technology
    for tech in ordered_techs:
        y_values = pivot_table[tech]
        # Skip if all values are zero
        if (y_values == 0).all():
            continue

        color = get_tech_color(tech)

        fig.add_trace(
            go.Bar(
                name=tech,
                x=pivot_table.index.astype(str),
                y=y_values,
                marker_color=color,
                hovertemplate=f"%{{x}}<br>{tech}: %{{y:.2f}}B USD<extra></extra>",
            )
        )

    # Update layout
    fig.update_layout(
        barmode="relative",
        title=dict(
            text=f"H2 & e-fuels Technologies - {cost_type_label}", font=dict(size=16)
        ),
        xaxis=dict(
            title=dict(text="Years", font=dict(size=14)), tickfont=dict(size=12)
        ),
        yaxis=dict(
            title=dict(text=f"{cost_type_label} (Billion USD)", font=dict(size=14)),
            tickfont=dict(size=12),
            tickformat=".2f",  # Force decimal format, no scientific notation
        ),
        template="plotly_white",
        width=1200,
        height=600,
        margin=dict(l=40, r=200, t=50, b=50),
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02,
            font=dict(size=12),
            traceorder="reversed",
        ),
    )

    # Add horizontal line at zero
    fig.add_hline(y=0, line_width=1, line_color="black")

    return fig


def create_h2_efuels_analysis(networks, index="year"):
    """
    Create complete analysis for H2 and e-fuels technologies.
    CAPEX in billion USD, OPEX in million USD.
    Hover only, legend limited to techs with data.
    """

    all_h2_efuels_costs = []
    for name_tag, network in networks.items():
        df_costs = compute_h2_efuels_costs(network, name_tag)
        all_h2_efuels_costs.append(df_costs)

    df_h2_efuels_costs = pd.concat(all_h2_efuels_costs, ignore_index=True)

    # Add cost in billion for CAPEX plotting
    df_h2_efuels_costs["cost_billion"] = df_h2_efuels_costs.apply(
        lambda r: r["cost_million"] / 1e3
        if r["cost_type"] == "Capital expenditure"
        else r["cost_million"],
        axis=1,
    )

    h2_efuels_order = [
        "Alkaline electrolyzer large",
        "PEM electrolyzer",
        "SOEC",
        "Fischer-Tropsch",
    ]

    h2_efuels_colors = {
        "Alkaline electrolyzer large": "#1f77b4",
        "PEM electrolyzer": "#2ca02c",
        "SOEC": "#d62728",
        "Fischer-Tropsch": "#e81cd0",
    }

    # keep only techs with nonzero costs
    active_techs = df_h2_efuels_costs.groupby("tech_label")["cost_million"].sum()
    active_techs = active_techs[active_techs > 0].index.tolist()
    active_colors = {k: v for k, v in h2_efuels_colors.items() if k in active_techs}
    active_order = [t for t in h2_efuels_order if t in active_techs]

    # CAPEX in billion USD
    df_capex = df_h2_efuels_costs[
        df_h2_efuels_costs["cost_type"] == "Capital expenditure"
    ].copy()
    df_capex["plot_value"] = df_capex["cost_billion"]

    fig1 = plot_h2_efuels_details(
        df_capex,
        "Capital expenditure",
        tech_colors=active_colors,
        tech_order=active_order,
        index=index,
    )
    if fig1:
        fig1.update_traces(
            text=None, hovertemplate="%{x} – %{label}<br>%{y:.2f} billion USD"
        )
        fig1.update_yaxes(title_text="Cost (billion USD)")
        fig1.show()

    # OPEX in million USD
    df_opex = df_h2_efuels_costs[
        df_h2_efuels_costs["cost_type"] == "Operational expenditure"
    ].copy()
    df_opex["plot_value"] = df_opex["cost_million"]

    fig2 = plot_h2_efuels_details(
        df_opex,
        "Operational expenditure",
        tech_colors=active_colors,
        tech_order=active_order,
        index=index,
    )
    if fig2:
        fig2.update_traces(
            text=None, hovertemplate="%{x} – %{label}<br>%{y:.2f} million USD"
        )
        fig2.update_yaxes(title_text="Cost (million USD)")
        fig2.show()

    return df_h2_efuels_costs


def hourly_matching_plot(networks, year_title=True):
    for idx, (network_name, network) in enumerate(networks.items()):
        year_str = network_name.split("_")[-1]
        # define additionality
        additionality = True

        # calculate electrolyzers consumption
        electrolysis_carrier = [
            "H2 Electrolysis",
            "Alkaline electrolyzer large",
            "Alkaline electrolyzer medium",
            "Alkaline electrolyzer small",
            "PEM electrolyzer",
            "SOEC",
        ]

        electrolyzers = network.links[
            network.links.carrier.isin(electrolysis_carrier)
        ].index
        electrolyzers_consumption = (
            network.links_t.p0[electrolyzers]
            .multiply(network.snapshot_weightings.objective, axis=0)
            .sum(axis=1)
        )

        # calculate RES generation
        res_carriers = [
            "csp",
            "solar",
            "onwind",
            "offwind-ac",
            "offwind-dc",
            "ror",
        ]
        res_stor_techs = ["hydro"]

        # get RES generators and storage units
        res_gens = network.generators.query("carrier in @res_carriers").index
        res_storages = network.storage_units.query("carrier in @res_stor_techs").index

        if additionality:
            # get new generators and storage_units
            new_gens = network.generators.loc[
                network.generators.build_year == int(year_str)
            ].index
            new_stor = network.storage_units.loc[
                network.storage_units.build_year == int(year_str)
            ].index
            # keep only new RES generators and storage units
            res_gens = res_gens.intersection(new_gens)
            res_storages = res_storages.intersection(new_stor)

        # calculate RES generation
        res_generation = (
            network.generators_t.p[res_gens]
            .multiply(network.snapshot_weightings.objective, axis=0)
            .sum(axis=1)
        )
        res_storages_dispatch = (
            network.storage_units_t.p[res_storages]
            .multiply(network.snapshot_weightings.objective, axis=0)
            .sum(axis=1)
        )
        res_generation_total = res_generation + res_storages_dispatch

        compare_df = pd.concat(
            [res_generation_total, electrolyzers_consumption], axis=1
        )
        compare_df.rename(
            columns={0: "RES generation", 1: "Electrolyzer consumption"}, inplace=True
        )
        # compare_df = compare_df.resample("D").mean()

        fig, ax = plt.subplots(figsize=(10, 2.5))
        (
            compare_df[["RES generation", "Electrolyzer consumption"]]
            .div(network.snapshot_weightings.objective, axis=0)
            .div(1e3)
            .resample("D")
            .mean()
            .plot(ax=ax)
        )
        ax.set_ylabel("GW")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
        ax.set_xlabel(None)
        ax.set_title(year_str if year_title else network_name)
        ax.legend(loc="lower left")


def preprocess_res_ces_share_eia_region(eia_gen_data, grid_regions):
    """
    Aggregate EIA 2023 generation data by grid region.
    Uses the mapping in grid_regions_to_states.xlsx.
    Returns a DataFrame with % Actual RES and % Actual CES for each grid region.
    """

    # Filter EIA data
    eia_gen_data = eia_gen_data[eia_gen_data["YEAR"] == 2023]
    eia_gen_data = eia_gen_data[eia_gen_data["STATE"] != "US-Total"]
    eia_gen_data = eia_gen_data[
        eia_gen_data["TYPE OF PRODUCER"] == "Total Electric Power Industry"
    ]
    eia_gen_data = eia_gen_data[
        (eia_gen_data["ENERGY SOURCE"] != "Total")
        & (eia_gen_data["ENERGY SOURCE"] != "Other")
    ]

    # Normalize energy source names
    eia_gen_data.replace(
        {
            "ENERGY SOURCE": {
                "Coal": "coal",
                "Hydroelectric Conventional": "hydro",
                "Pumped Storage": "PHS",
                "Solar Thermal and Photovoltaic": "solar",
                "Natural Gas": "gas",
                "Petroleum": "oil",
                "Wind": "wind",
                "Nuclear": "nuclear",
                "Geothermal": "geothermal",
                "Wood and Wood Derived Fuels": "biomass",
                "Other Biomass": "biomass",
                "Other Gases": "gas",
            }
        },
        inplace=True,
    )

    # Convert to TWh
    eia_gen_data["GENERATION (TWh)"] = eia_gen_data["GENERATION (Megawatthours)"] / 1e6

    # Pivot: state × source
    eia_state_df = (
        eia_gen_data.groupby(["STATE", "ENERGY SOURCE"])[["GENERATION (TWh)"]]
        .sum()
        .unstack(fill_value=0)
    )
    eia_state_df.columns = eia_state_df.columns.droplevel(0)

    # Merge with grid region mapping
    eia_with_region = eia_state_df.reset_index().rename(columns={"STATE": "State"})
    grid_regions = grid_regions.copy()
    grid_regions["States"] = grid_regions["States"].str.strip().str.upper()
    eia_with_region = eia_with_region.merge(
        grid_regions, left_on="State", right_on="States"
    )

    # Aggregate by grid region
    region_agg = eia_with_region.groupby("Grid region").sum(numeric_only=True)

    # Compute % RES and % CES
    eia_res_carriers = ["solar", "wind", "hydro", "geothermal", "biomass", "PHS"]
    eia_ces_carriers = eia_res_carriers + ["nuclear"]

    total = region_agg.sum(axis=1)
    res_total = region_agg[eia_res_carriers].sum(axis=1)
    ces_total = region_agg[eia_ces_carriers].sum(axis=1)

    region_agg["% Actual RES"] = (res_total / total) * 100
    region_agg["% Actual CES"] = (ces_total / total) * 100

    return region_agg[["% Actual RES", "% Actual CES"]]


def evaluate_res_ces_by_region(networks, ces_carriers, res_carriers):
    """
    Evaluate RES and CES shares from the model, aggregated by grid region.
    Assumes `attach_grid_region_to_buses` has been called on each network.
    Returns results[(scenario, year)] = df.
    """

    results = {}

    for name, network in networks.items():
        # Extract scenario and year from network name (e.g., "scenario_01_2030" -> scenario="scenario_01", year=2030)
        match = re.search(r"(?:scenario_(\d{2})|Base)_(\d{4})", name)
        if match:
            if match.group(1):
                scenario = f"scenario_{match.group(1)}"
            else:
                scenario = "Base"
            year = int(match.group(2))
        else:
            print(
                f"Warning: Skipping network '{name}' - does not match expected format (e.g., 'scenario_01_2030' or 'Base_2023')."
            )
            continue

        snapshots = network.snapshots
        timestep_h = (snapshots[1] - snapshots[0]).total_seconds() / 3600
        snapshots_slice = slice(None)

        gen_and_sto_carriers = {
            "csp",
            "solar",
            "onwind",
            "offwind-dc",
            "offwind-ac",
            "nuclear",
            "geothermal",
            "ror",
            "hydro",
            "solar rooftop",
        }
        link_carriers = ["coal", "oil", "OCGT", "CCGT", "biomass", "lignite"]

        # Generators
        gen = network.generators[
            network.generators.carrier.isin(gen_and_sto_carriers)
        ].copy()
        gen["grid_region"] = gen["bus"].map(network.buses["grid_region"])
        gen = gen[gen["grid_region"].notna()]

        gen_p = network.generators_t.p.loc[snapshots_slice, gen.index].clip(lower=0)
        gen_energy = gen_p.multiply(timestep_h).sum()
        gen_energy = gen_energy.to_frame(name="energy_mwh")
        gen_energy["carrier"] = gen.loc[gen_energy.index, "carrier"]
        gen_energy["grid_region"] = gen.loc[gen_energy.index, "grid_region"]

        # Storage
        sto = network.storage_units[
            network.storage_units.carrier.isin(gen_and_sto_carriers)
        ].copy()
        sto["grid_region"] = sto["bus"].map(network.buses["grid_region"])
        sto = sto[sto["grid_region"].notna()]

        sto_p = network.storage_units_t.p.loc[snapshots_slice, sto.index].clip(lower=0)
        sto_energy = sto_p.multiply(timestep_h).sum()
        sto_energy = sto_energy.to_frame(name="energy_mwh")
        sto_energy["carrier"] = sto.loc[sto_energy.index, "carrier"]
        sto_energy["grid_region"] = sto.loc[sto_energy.index, "grid_region"]

        # Links
        link_data = []
        for i, link in network.links.iterrows():
            if link["carrier"] in link_carriers and pd.notna(
                network.buses.loc[link["bus1"], "grid_region"]
            ):
                p1 = -network.links_t.p1.loc[snapshots_slice, i].clip(upper=0)
                energy_mwh = p1.sum() * timestep_h
                link_data.append(
                    {
                        "carrier": link["carrier"],
                        "grid_region": network.buses.loc[link["bus1"], "grid_region"],
                        "energy_mwh": energy_mwh,
                    }
                )

        link_energy = pd.DataFrame(link_data)

        # Combine
        all_energy = pd.concat(
            [
                gen_energy[["carrier", "grid_region", "energy_mwh"]],
                sto_energy[["carrier", "grid_region", "energy_mwh"]],
                link_energy[["carrier", "grid_region", "energy_mwh"]],
            ]
        )

        # Aggregate by grid region
        region_totals = all_energy.groupby("grid_region")["energy_mwh"].sum()
        region_ces = (
            all_energy[all_energy["carrier"].isin(ces_carriers)]
            .groupby("grid_region")["energy_mwh"]
            .sum()
        )
        region_res = (
            all_energy[all_energy["carrier"].isin(res_carriers)]
            .groupby("grid_region")["energy_mwh"]
            .sum()
        )

        df = pd.DataFrame(
            {
                "Total (MWh)": region_totals,
                "CES_energy": region_ces,
                "RES_energy": region_res,
            }
        ).fillna(0)

        df["% CES"] = 100 * df["CES_energy"] / df["Total (MWh)"]
        df["% RES"] = 100 * df["RES_energy"] / df["Total (MWh)"]

        df = df[["Total (MWh)", "% RES", "% CES"]].round(2)

        results[(scenario, year)] = df.sort_index()

    return results


# Function to format values
def fmt_2dp_or_na(v):
    if isinstance(v, str) and v.strip().upper() == "N/A":
        return "N/A"
    if pd.isna(v):
        return ""
    try:
        return f"{float(v):.2f}"
    except Exception:
        return v


# Function for deviation-based coloring (2023 only)


def deviation_color(a, b):
    """
    Color by absolute percent deviation |a - b| / b:
      - Green  -> ≤ 10%
      - Yellow -> > 10% and ≤ 20%
      - Red    -> > 20%
      - None   -> if b is 'N/A', NaN, or zero
    """
    try:
        if isinstance(b, str) and b.strip().upper() == "N/A":
            return ""
        a_val = float(a)
        b_val = float(b)
        if not pd.notna(b_val) or b_val == 0:
            return ""  # avoid invalid/zero baseline
        deviation = abs((a_val - b_val) / b_val) * 100.0

        if deviation <= 10:
            return "background-color:#d4edda;"  # green
        elif deviation <= 20:
            return "background-color:#fff3cd;"  # yellow
        else:
            return "background-color:#f8d7da;"  # red
    except Exception:
        return ""


# Simple green/red for future years
def simple_color(a, b):
    """
    Returns green if a >= b, red if a < b, none if N/A or not numeric.
    """
    try:
        if isinstance(b, str) and b.strip().upper() == "N/A":
            return ""
        return (
            "background-color:#d4edda;"
            if float(a) >= float(b)
            else "background-color:#f8d7da;"
        )
    except:
        return ""


def get_us_from_eia(eia_generation_data):
    """
    Compute national (US) RES and CES shares from EIA data (2023).
    Weighted by total generation (TWh) of each state.
    """
    eia_state = preprocess_res_ces_share_eia(eia_generation_data)

    # Get total generation per state in TWh
    eia_gen_data = eia_generation_data[
        (eia_generation_data["YEAR"] == 2023)
        & (eia_generation_data["STATE"] != "US-Total")
        & (eia_generation_data["TYPE OF PRODUCER"] == "Total Electric Power Industry")
    ].copy()
    eia_gen_data["GENERATION (TWh)"] = eia_gen_data["GENERATION (Megawatthours)"] / 1e6
    total_by_state = eia_gen_data.groupby("STATE")["GENERATION (TWh)"].sum()

    # Weighted average
    us_res = (eia_state["% Actual RES"] * total_by_state).sum() / total_by_state.sum()
    us_ces = (eia_state["% Actual CES"] * total_by_state).sum() / total_by_state.sum()

    return round(us_res, 2), round(us_ces, 2)


def preprocess_res_ces_share_grid_region(
    eia_gen_data=None,
    grid_regions=None,
    file_path="./data/validation_data/generation_grid_regions.xlsx",
    sheet_name="Generation (TWh)",
):
    """
    Drop-in replacement for preprocess_res_ces_share_eia_region.
    Ignores eia_gen_data and grid_regions, and instead loads
    precomputed grid-region data in TWh from generation_grid_regions.xlsx.
    """

    import pandas as pd

    df = pd.read_excel(file_path, sheet_name=sheet_name)

    if "Region" in df.columns and "Grid Region" not in df.columns:
        df = df.rename(columns={"Region": "Grid Region"})

    res_carriers = ["Solar", "Wind", "Hydro", "Geothermal", "Biomass"]
    ces_carriers = res_carriers + ["Nuclear"]

    total = df[["Coal", "Gas", "Oil", "Nuclear", "Other"] + res_carriers].sum(axis=1)
    res_total = df[res_carriers].sum(axis=1)
    ces_total = df[ces_carriers].sum(axis=1)

    result = df[["Grid Region"]].copy()
    result["% Actual RES"] = (res_total / total) * 100
    result["% Actual CES"] = (ces_total / total) * 100

    return result.set_index("Grid Region")


def display_grid_region_results(networks, ces, res, ces_carriers, res_carriers):
    """
    Show RES/CES results for grid regions using pre-aggregated Excel file
    (generation_grid_regions.xlsx). Adds columns for absolute generation (TWh),
    regional shares (%), and a national total row (U.S.).
    """

    res_by_region = evaluate_res_ces_by_region(
        networks, ces_carriers=ces_carriers, res_carriers=res_carriers
    )

    legend_html = """
    <div style="padding:10px; margin-bottom:15px; border:1px solid #ccc; border-radius:5px; width: fit-content;">
    <strong>Legend</strong>
    <ul style="margin:5px 0; padding-left:20px;">
        <li style="background-color:#d4edda; padding:2px;">Diff. ≤ ±10%</li>
        <li style="background-color:#fff3cd; padding:2px;">10% < Diff. ≤ 20%</li>
        <li style="background-color:#f8d7da; padding:2px;">Diff. > 20%</li>
    </ul>
    </div>
    """

    html_blocks = []
    cols_per_row = 2

    for scenario, yr in sorted(res_by_region.keys()):
        df_year = res_by_region[(scenario, yr)].copy()

        if yr == 2023:
            # Load actuals from Excel
            eia_region = preprocess_res_ces_share_grid_region()

            excel_df = pd.read_excel(
                "./data/validation_data/generation_grid_regions.xlsx",
                sheet_name="Generation (TWh)",
            )
            if "Region" in excel_df.columns and "Grid Region" not in excel_df.columns:
                excel_df = excel_df.rename(columns={"Region": "Grid Region"})
            excel_df = excel_df.set_index("Grid Region")

            # Add stats total generation
            eia_region = eia_region.join(excel_df[["Net generation (TWh)"]])

            # Add model total generation (TWh)
            df_year["Model generation (TWh)"] = df_year["Total (MWh)"] / 1e6

            # Merge with stats
            df_year = df_year.merge(eia_region, left_index=True, right_index=True)
            df_year = df_year.rename(
                columns={"Net generation (TWh)": "Stats generation (TWh)"}
            )

            # Regional shares
            df_year["% Model generation share"] = (
                df_year["Model generation (TWh)"]
                / df_year["Model generation (TWh)"].sum()
                * 100
            )
            df_year["% Stats generation share"] = (
                df_year["Stats generation (TWh)"]
                / df_year["Stats generation (TWh)"].sum()
                * 100
            )

            # Add national total row (U.S.)
            totals = pd.Series(
                {
                    "% RES": (
                        df_year["% RES"] * df_year["Model generation (TWh)"]
                    ).sum()
                    / df_year["Model generation (TWh)"].sum(),
                    "% Actual RES": (
                        df_year["% Actual RES"] * df_year["Stats generation (TWh)"]
                    ).sum()
                    / df_year["Stats generation (TWh)"].sum(),
                    "% CES": (
                        df_year["% CES"] * df_year["Model generation (TWh)"]
                    ).sum()
                    / df_year["Model generation (TWh)"].sum(),
                    "% Actual CES": (
                        df_year["% Actual CES"] * df_year["Stats generation (TWh)"]
                    ).sum()
                    / df_year["Stats generation (TWh)"].sum(),
                    "Model generation (TWh)": df_year["Model generation (TWh)"].sum(),
                    "Stats generation (TWh)": df_year["Stats generation (TWh)"].sum(),
                    "% Model generation share": 100.0,
                    "% Stats generation share": 100.0,
                },
                name="U.S.",
            )
            df_year = pd.concat([df_year, totals.to_frame().T])

            df_disp = df_year[
                [
                    "% RES",
                    "% Actual RES",
                    "% CES",
                    "% Actual CES",
                    "Model generation (TWh)",
                    "Stats generation (TWh)",
                    "% Model generation share",
                    "% Stats generation share",
                ]
            ].round(2)

            df_disp = (
                df_disp.reset_index()
                .rename(columns={"index": "Grid Region"})
                .set_index("Grid Region")
            )

            def style_row(row):
                styles = []
                for col in df_disp.columns:
                    if "RES" in col:
                        styles.append(
                            deviation_color(row.get("% RES"), row.get("% Actual RES"))
                        )
                    elif "CES" in col:
                        styles.append(
                            deviation_color(row.get("% CES"), row.get("% Actual CES"))
                        )
                    elif "generation" in col or "share" in col:
                        styles.append(
                            deviation_color(
                                row.get("Model generation (TWh)"),
                                row.get("Stats generation (TWh)"),
                            )
                        )
                    else:
                        styles.append("")
                return styles

            styled_df = (
                df_disp.style.apply(style_row, axis=1)
                .format(fmt_2dp_or_na)
                .set_table_styles(
                    [{"selector": "th.row_heading", "props": "font-weight:bold;"}]
                )
            )

            # Force wide table (no wrapping)
            df_html = styled_df.to_html() + legend_html
            df_html = (
                f"<div style='overflow-x:auto; white-space:nowrap;'>{df_html}</div>"
            )

        else:
            expected_cols = ["% RES", "% RES target", "% CES", "% CES target"]
            cols_present = [c for c in expected_cols if c in df_year.columns]
            df_year = df_year.reindex(columns=cols_present).round(2)

            # Add model total generation for future years
            if "Total (MWh)" in df_year.columns:
                df_year["Model generation (TWh)"] = df_year["Total (MWh)"] / 1e6
                df_year["% Model generation share"] = (
                    df_year["Model generation (TWh)"]
                    / df_year["Model generation (TWh)"].sum()
                    * 100
                )

            totals = pd.Series(
                {c: df_year[c].mean() for c in df_year.columns}, name="U.S."
            )
            df_year = pd.concat([df_year, totals.to_frame().T])

            df_disp = (
                df_year.reset_index()
                .rename(columns={"index": "Grid Region"})
                .set_index("Grid Region")
            )

            def style_row(row):
                styles = []
                for col in df_disp.columns:
                    if col.startswith("% RES"):
                        styles.append(
                            simple_color(row.get("% RES"), row.get("% RES target"))
                        )
                    elif col.startswith("% CES"):
                        styles.append(
                            simple_color(row.get("% CES"), row.get("% CES target"))
                        )
                    else:
                        styles.append("")
                return styles

            styled_df = (
                df_disp.style.apply(style_row, axis=1)
                .format(fmt_2dp_or_na)
                .set_table_styles(
                    [{"selector": "th.row_heading", "props": "font-weight:bold;"}]
                )
            )
            df_html = styled_df.to_html()

        block = f"""
        <div style="flex:1; padding:20px; min-width:300px;">
            <h4 style="text-align:left;">Year: {yr}</h4>
            {df_html}
        </div>
        """
        html_blocks.append(block)

    rows = [
        "<div style='display:flex; gap:10px; flex-wrap:wrap;'>"
        + "".join(html_blocks[i : i + cols_per_row])
        + "</div>"
        for i in range(0, len(html_blocks), cols_per_row)
    ]

    for row in rows:
        display(HTML(row))


def attach_emm_region_to_buses(network, path_shape, distance_crs):
    """
    Attach EMM region to buses
    """
    # Read the shapefile using geopandas
    shape = gpd.read_file(path_shape, crs=distance_crs)
    # shape.rename(columns={"GRID_REGIO": "region"}, inplace=True)

    ac_dc_carriers = ["AC", "DC"]
    location_mapping = network.buses.query("carrier in @ac_dc_carriers")[["x", "y"]]

    network.buses["x"] = network.buses["location"].map(location_mapping["x"]).fillna(0)
    network.buses["y"] = network.buses["location"].map(location_mapping["y"]).fillna(0)

    pypsa_gpd = gpd.GeoDataFrame(
        network.buses,
        geometry=gpd.points_from_xy(network.buses.x, network.buses.y),
        crs=4326,
    )

    network_columns = network.buses.columns
    bus_cols = [*network_columns, "subregion"]

    st_buses = gpd.sjoin_nearest(shape, pypsa_gpd, how="right")[bus_cols]

    network.buses["emm_region"] = st_buses["subregion"]


def compute_ekerosene_by_region(
    networks: dict,
    regional_fees: pd.DataFrame,
    year_title: bool = True,
    p_nom_threshold: float = 1e-3,
    unit: str = "MWh",  # "MWh" or "gal"
):
    """
    Compute input rates, input prices, T&D fees and total unit cost for
    Fischer-Tropsch e-kerosene plants by grid region.
    """

    MWH_PER_GAL = (34.0 / 3600.0) * 3.78541  # MWh per gallon
    conv = 1.0 if unit == "MWh" else MWH_PER_GAL
    suffix = f"USD/{unit} e-ker"

    for name, net in networks.items():
        m = re.search(r"\d{4}", str(name))
        scen_year = int(m.group()) if m else 2030

        emm_mapping = dict(zip(net.buses.grid_region, net.buses.emm_region))

        ft = net.links[
            net.links.carrier.str.contains("Fischer-Tropsch", case=False, na=False)
        ].copy()
        if ft.empty:
            continue

        cap_series = np.where(
            ft.get("p_nom_extendable", False),
            ft.get("p_nom_opt", 0.0),
            ft.get("p_nom", 0.0),
        )
        ft = ft[pd.Series(cap_series, index=ft.index) > p_nom_threshold]
        if ft.empty:
            continue

        needed_cols = ("p0", "p1", "p2", "p3")
        ft_ids = [
            l
            for l in ft.index
            if all(l in getattr(net.links_t, c).columns for c in needed_cols)
        ]
        if not ft_ids:
            continue
        ft = ft.loc[ft_ids]

        dt_h = (
            (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600.0
            if len(net.snapshots) > 1
            else 1.0
        )
        rows = []

        for link in ft.index:
            try:
                region = net.buses.at[ft.at[link, "bus1"], "grid_region"]
            except KeyError:
                continue

            bus_map = {}
            for i in range(4):
                b = ft.at[link, f"bus{i}"]
                if b not in net.buses.index:
                    continue
                carrier = net.buses.at[b, "carrier"]
                bus_map[carrier] = i

            out_MWh = (-net.links_t.p1[link] * dt_h).sum()
            if out_MWh <= 0:
                continue

            def get_flow_and_price(carrier_key):
                if carrier_key not in bus_map:
                    return 0.0, None
                idx = bus_map[carrier_key]
                flow = (net.links_t[f"p{idx}"][link] * dt_h).sum()
                price = net.buses_t.marginal_price[ft.at[link, f"bus{idx}"]]
                return flow, price

            elec_in, p_elec = get_flow_and_price("AC")
            h2_in, p_h2 = get_flow_and_price("grid H2")
            co2_in, p_co2 = get_flow_and_price("co2 stored")

            r_elec = elec_in / out_MWh if out_MWh > 0 else 0.0
            r_h2 = h2_in / out_MWh if out_MWh > 0 else 0.0
            r_co2 = co2_in / out_MWh if out_MWh > 0 else 0.0

            def avg_price(flow_series, price_series, total_flow):
                return (
                    (flow_series * price_series * dt_h).sum() / total_flow
                    if total_flow > 0
                    else 0.0
                )

            avg_p_elec = (
                avg_price(net.links_t[f"p{bus_map['AC']}"][link], p_elec, elec_in)
                if "AC" in bus_map
                else 0.0
            )
            avg_p_h2 = (
                avg_price(net.links_t[f"p{bus_map['grid H2']}"][link], p_h2, h2_in)
                if "grid H2" in bus_map
                else 0.0
            )
            avg_p_co2 = (
                avg_price(net.links_t[f"p{bus_map['co2 stored']}"][link], p_co2, co2_in)
                if "co2 stored" in bus_map
                else 0.0
            )

            c_elec = avg_p_elec * r_elec / conv
            c_h2 = avg_p_h2 * r_h2 / conv
            c_co2 = avg_p_co2 * r_co2 / conv

            rows.append(
                {
                    "Grid Region": region,
                    "Production (TWh)": out_MWh / 1e6,
                    f"Electricity rate (MWh el / MWh e-ker)": r_elec,
                    f"H2 rate (MWh H2 / MWh e-ker)": r_h2,
                    f"CO2 rate (tCO2 / MWh e-ker)": r_co2,
                    f"Electricity cost ({suffix})": c_elec,
                    f"Hydrogen cost ({suffix})": c_h2,
                    f"CO2 cost ({suffix})": c_co2,
                }
            )

        if not rows:
            continue

        df = pd.DataFrame(rows)

        def wavg(group, col):
            return (group[col] * group["Production (TWh)"]).sum() / group[
                "Production (TWh)"
            ].sum()

        g = (
            df.groupby("Grid Region")
            .apply(
                lambda x: pd.Series(
                    {
                        "Production (TWh)": x["Production (TWh)"].sum(),
                        f"Electricity rate (MWh el / MWh e-ker)": wavg(
                            x, f"Electricity rate (MWh el / MWh e-ker)"
                        ),
                        f"H2 rate (MWh H2 / MWh e-ker)": wavg(
                            x, f"H2 rate (MWh H2 / MWh e-ker)"
                        ),
                        f"CO2 rate (tCO2 / MWh e-ker)": wavg(
                            x, f"CO2 rate (tCO2 / MWh e-ker)"
                        ),
                        f"Electricity cost ({suffix})": wavg(
                            x, f"Electricity cost ({suffix})"
                        ),
                        f"Hydrogen cost ({suffix})": wavg(
                            x, f"Hydrogen cost ({suffix})"
                        ),
                        f"CO2 cost ({suffix})": wavg(x, f"CO2 cost ({suffix})"),
                    }
                )
            )
            .reset_index()  # keep Grid Region, drop numeric index
        )

        g["EMM Region"] = g["Grid Region"].map(emm_mapping)
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year,
            ["region", "Transmission nom USD/MWh", "Distribution nom USD/MWh"],
        ].set_index("region")

        g[f"Transmission fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Transmission nom USD/MWh"])
            * g[f"Electricity rate (MWh el / MWh e-ker)"]
            / conv
        )
        g[f"Distribution fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Distribution nom USD/MWh"])
            * g[f"Electricity rate (MWh el / MWh e-ker)"]
            / conv
        )
        g.drop(columns=["EMM Region"], inplace=True)

        g[f"LCO e-kerosene incl. T&D ({suffix})"] = (
            g[f"Electricity cost ({suffix})"]
            + g[f"Hydrogen cost ({suffix})"]
            + g[f"CO2 cost ({suffix})"]
            + g[f"Transmission fees ({suffix})"]
            + g[f"Distribution fees ({suffix})"]
        )

        title = (
            re.search(r"\d{4}", str(name)).group()
            if year_title and re.search(r"\d{4}", str(name))
            else str(name)
        )
        print(f"\n{title} ({unit} view):")
        display(g.round(3))


def compute_ekerosene_by_region(
    networks: dict,
    regional_fees: pd.DataFrame,
    year_title: bool = True,
    p_nom_threshold: float = 1e-3,
    unit: str = "MWh",  # "MWh" or "gal"
):
    """
    Compute input rates, input prices, T&D fees and total unit cost for
    Fischer-Tropsch e-kerosene plants by grid region.
    """

    MWH_PER_GAL = (34.0 / 3600.0) * 3.78541  # MWh per gallon
    conv = 1.0 if unit == "MWh" else MWH_PER_GAL
    suffix = f"USD/{unit} e-ker"

    for name, net in networks.items():
        m = re.search(r"\d{4}", str(name))
        scen_year = int(m.group()) if m else 2030

        emm_mapping = dict(zip(net.buses.grid_region, net.buses.emm_region))

        ft = net.links[
            net.links.carrier.str.contains("Fischer-Tropsch", case=False, na=False)
        ].copy()
        if ft.empty:
            continue

        cap_series = np.where(
            ft.get("p_nom_extendable", False),
            ft.get("p_nom_opt", 0.0),
            ft.get("p_nom", 0.0),
        )
        ft = ft[pd.Series(cap_series, index=ft.index) > p_nom_threshold]
        if ft.empty:
            continue

        needed_cols = ("p0", "p1", "p2", "p3")
        ft_ids = [
            l
            for l in ft.index
            if all(l in getattr(net.links_t, c).columns for c in needed_cols)
        ]
        if not ft_ids:
            continue
        ft = ft.loc[ft_ids]

        dt_h = (
            (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600.0
            if len(net.snapshots) > 1
            else 1.0
        )
        rows = []

        for link in ft.index:
            try:
                region = net.buses.at[ft.at[link, "bus1"], "grid_region"]
            except KeyError:
                continue

            bus_map = {}
            for i in range(4):
                b = ft.at[link, f"bus{i}"]
                if b not in net.buses.index:
                    continue
                carrier = net.buses.at[b, "carrier"]
                bus_map[carrier] = i

            out_MWh = (-net.links_t.p1[link] * dt_h).sum()
            if out_MWh <= 0:
                continue

            def get_flow_and_price(carrier_key):
                if carrier_key not in bus_map:
                    return 0.0, None
                idx = bus_map[carrier_key]
                flow = (net.links_t[f"p{idx}"][link] * dt_h).sum()
                price = net.buses_t.marginal_price[ft.at[link, f"bus{idx}"]]
                return flow, price

            elec_in, p_elec = get_flow_and_price("AC")
            h2_in, p_h2 = get_flow_and_price("grid H2")
            co2_in, p_co2 = get_flow_and_price("co2 stored")

            r_elec = elec_in / out_MWh if out_MWh > 0 else 0.0
            r_h2 = h2_in / out_MWh if out_MWh > 0 else 0.0
            r_co2 = co2_in / out_MWh if out_MWh > 0 else 0.0

            def avg_price(flow_series, price_series, total_flow):
                return (
                    (flow_series * price_series * dt_h).sum() / total_flow
                    if total_flow > 0
                    else 0.0
                )

            avg_p_elec = (
                avg_price(net.links_t[f"p{bus_map['AC']}"][link], p_elec, elec_in)
                if "AC" in bus_map
                else 0.0
            )
            avg_p_h2 = (
                avg_price(net.links_t[f"p{bus_map['grid H2']}"][link], p_h2, h2_in)
                if "grid H2" in bus_map
                else 0.0
            )
            avg_p_co2 = (
                avg_price(net.links_t[f"p{bus_map['co2 stored']}"][link], p_co2, co2_in)
                if "co2 stored" in bus_map
                else 0.0
            )

            c_elec = avg_p_elec * r_elec / conv
            c_h2 = avg_p_h2 * r_h2 / conv
            c_co2 = avg_p_co2 * r_co2 / conv

            rows.append(
                {
                    "Grid Region": region,
                    "Production (TWh)": out_MWh / 1e6,
                    f"Electricity rate (MWh el / MWh e-ker)": r_elec,
                    f"H2 rate (MWh H2 / MWh e-ker)": r_h2,
                    f"CO2 rate (tCO2 / MWh e-ker)": r_co2,
                    f"Electricity cost ({suffix})": c_elec,
                    f"Hydrogen cost ({suffix})": c_h2,
                    f"CO2 cost ({suffix})": c_co2,
                }
            )

        if not rows:
            continue

        df = pd.DataFrame(rows)

        def wavg(group, col):
            return (group[col] * group["Production (TWh)"]).sum() / group[
                "Production (TWh)"
            ].sum()

        g = (
            df.groupby("Grid Region")
            .apply(
                lambda x: pd.Series(
                    {
                        "Production (TWh)": x["Production (TWh)"].sum(),
                        f"Electricity rate (MWh el / MWh e-ker)": wavg(
                            x, f"Electricity rate (MWh el / MWh e-ker)"
                        ),
                        f"H2 rate (MWh H2 / MWh e-ker)": wavg(
                            x, f"H2 rate (MWh H2 / MWh e-ker)"
                        ),
                        f"CO2 rate (tCO2 / MWh e-ker)": wavg(
                            x, f"CO2 rate (tCO2 / MWh e-ker)"
                        ),
                        f"Electricity cost ({suffix})": wavg(
                            x, f"Electricity cost ({suffix})"
                        ),
                        f"Hydrogen cost ({suffix})": wavg(
                            x, f"Hydrogen cost ({suffix})"
                        ),
                        f"CO2 cost ({suffix})": wavg(x, f"CO2 cost ({suffix})"),
                    }
                )
            )
            .reset_index()  # keep Grid Region, drop numeric index
        )

        g["EMM Region"] = g["Grid Region"].map(emm_mapping)
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year,
            ["region", "Transmission nom USD/MWh", "Distribution nom USD/MWh"],
        ].set_index("region")

        g[f"Transmission fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Transmission nom USD/MWh"])
            * g[f"Electricity rate (MWh el / MWh e-ker)"]
            / conv
        )
        g[f"Distribution fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Distribution nom USD/MWh"])
            * g[f"Electricity rate (MWh el / MWh e-ker)"]
            / conv
        )
        g.drop(columns=["EMM Region"], inplace=True)

        g[f"LCO e-kerosene incl. T&D ({suffix})"] = (
            g[f"Electricity cost ({suffix})"]
            + g[f"Hydrogen cost ({suffix})"]
            + g[f"CO2 cost ({suffix})"]
            + g[f"Transmission fees ({suffix})"]
            + g[f"Distribution fees ({suffix})"]
        )

        title = (
            re.search(r"\d{4}", str(name)).group()
            if year_title and re.search(r"\d{4}", str(name))
            else str(name)
        )
        print(f"\n{title} ({unit} view):")
        display(g.round(3))


def compute_ft_capacity_factor(
    networks: dict,
    carrier_regex: str = "Fischer-Tropsch",
    p_nom_threshold: float = 1.0,  # MW
    year_title: bool = True,
    round_digits: int = 2,
    output_threshold: float = 1.0,  # MWh
):
    """
    Compute annual capacity factor of Fischer-Tropsch plants by grid region.

    Definition:
    - Capacity factor (%) = H2 input at bus0 (MWh) /
                            (Installed capacity at bus0 (MW) * total weighted hours)

    Notes:
    - p_nom / p_nom_opt always refer to bus0 (PyPSA convention).
    - p0 (bus0) = H2 input (MWh).
    - p1 (bus1) = fuel output (MWh).
    """

    results = {}

    for year_key, n in networks.items():
        # extract year
        m = re.search(r"\d{4}", str(year_key))
        if not m:
            continue
        scen_year = int(m.group())
        key = str(scen_year) if year_title else year_key

        # select FT links
        ft = n.links[
            n.links.carrier.str.contains(carrier_regex, case=False, na=False)
        ].copy()
        if ft.empty:
            continue

        # capacity series (MW, bus0 reference)
        cap_series = pd.Series(
            np.where(
                ft.get("p_nom_extendable", False),
                ft.get("p_nom_opt", 0.0),
                ft.get("p_nom", 0.0),
            ),
            index=ft.index,
        ).astype(float)

        # filter by capacity threshold
        ft = ft[cap_series > p_nom_threshold]
        cap_series = cap_series.loc[ft.index]
        if ft.empty:
            continue

        # snapshot weighting (hours per snapshot)
        w = n.snapshot_weightings.generators
        total_hours = float(w.sum())

        # H2 input (bus0, MWh)
        p0 = n.links_t.p0[ft.index]
        energy_h2_in = (p0).clip(lower=0).multiply(w, axis=0).sum(axis=0)

        # Fuel output (bus1, MWh, for reporting only)
        p1 = n.links_t.p1[ft.index]
        energy_out = (-p1).clip(lower=0).multiply(w, axis=0).sum(axis=0)

        # add region info
        ft["grid_region"] = ft["bus1"].map(n.buses["grid_region"])
        df_links = pd.DataFrame(
            {
                "Link": ft.index,
                "Grid Region": ft["grid_region"],
                "Capacity (MW)": cap_series,
                "Hydrogen input (MWh)": energy_h2_in,
                "Fuel output (MWh)": energy_out,
            }
        ).dropna()

        # aggregate by region
        region_summary = (
            df_links.groupby("Grid Region")
            .apply(
                lambda g: pd.Series(
                    {
                        "Capacity (GW input H2)": g["Capacity (MW)"].sum() / 1e3,
                        "Hydrogen input (MWh)": g["Hydrogen input (MWh)"].sum(),
                        "Fuel output (MWh)": g["Fuel output (MWh)"].sum(),
                        "Capacity factor (%)": (
                            g["Hydrogen input (MWh)"].sum()
                            / (g["Capacity (MW)"].sum() * total_hours)
                            * 100
                        ),
                    }
                )
            )
            .reset_index()
        )

        # skip if negligible production
        if region_summary["Fuel output (MWh)"].sum() < output_threshold:
            continue

        results[key] = region_summary.round(round_digits)

    return results


def compute_LCO_ekerosene_by_region(
    networks: dict,
    fx_2020: float,
    fx_recent: float,
    regional_fees: pd.DataFrame,
    emm_mapping: dict,
    unit: str = "gal",  # "gal" or "MWh"
    year_title: bool = True,
    p_nom_threshold: float = 1e-3,
    electricity_price: str = "marginal",  # "marginal" or "lcoe"
    hydrogen_price: str = "marginal",  # "marginal" or "lcoh"
    co2_price: str = "marginal",  # "marginal" or "lcoc"
    lcoe_by_region=None,  # Series or dict
    lcoh_by_region: dict = None,  # required if hydrogen_price in {"lcoh", "marginal"} for fees
    lcoc_by_region: dict = None,  # required if co2_price="lcoc"
    verbose=True,
):
    """
    Levelized cost of e-kerosene by grid region (USD/gal or USD/MWh e-ker).

    Convention:
    - Input rates = consumed energy / MWh e-ker (production-weighted).
    - Input prices = consumption-weighted marginal price (or LCOE/LCOH/LCOC if requested).
    - Autodetects flow sign per port (robust to sign conventions).

    IMPORTANT:
    - If hydrogen_price == "marginal", we still ADD H2 transmission + baseload EX-POST,
      using the SAME deltas as in calculate_lcoh_by_region, i.e. from lcoh_by_region:
        (LCOH + Trans + Baseload) - (LCOH excl. T&D)  [USD/kg] -> [USD/MWh] via *1000/33
    - No new arguments were added to the signature.
    """

    def get_year(data_dict, scen_year: int) -> str:
        """Return scen_year if available, otherwise fallback (only 2023→2020)."""
        if not data_dict:
            raise ValueError("Dataset is empty, cannot determine year.")
        if str(scen_year) in data_dict:
            return str(scen_year)
        if scen_year == 2023 and "2020" in data_dict:
            print("Year 2023 not found in dataset, using 2020 instead.")
            return "2020"
        raise KeyError(
            f"Year {scen_year} not found and no fallback available. "
            f"Keys: {list(data_dict.keys())}"
        )

    # Conversion MWh ↔ gallon
    MWH_PER_GALLON = (34.0 / 3600.0) * 3.78541
    conv = MWH_PER_GALLON if unit == "gal" else 1.0
    suffix = f"USD/{unit} e-ker"

    results = {}

    # VOM trajectory (EUR)
    vom_eur_points = {
        2020: 5.6360,
        2025: 5.0512,
        2030: 4.4663,
        2035: 3.9346,
        2040: 3.4029,
    }
    years_sorted = np.array(sorted(vom_eur_points.keys()))
    values_sorted = np.array([vom_eur_points[y] for y in years_sorted])

    def vom_usd_for_year(y: int) -> float:
        vom_eur = float(np.interp(y, years_sorted, values_sorted))
        fx = fx_2020 if y == 2020 else fx_recent
        return vom_eur * fx

    for name, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(name)).group())

        # Select Fischer-Tropsch links
        ft = net.links[
            net.links.carrier.str.contains("Fischer-Tropsch", case=False, na=False)
        ].copy()
        if ft.empty:
            continue

        # Capacity filter
        cap_series = np.where(
            ft.get("p_nom_extendable", False),
            ft.get("p_nom_opt", 0.0),
            ft.get("p_nom", 0.0),
        )
        ft = ft[pd.Series(cap_series, index=ft.index) > p_nom_threshold]
        if ft.empty:
            continue

        # Snapshot duration [h]
        if len(net.snapshots) > 1:
            dt_h = (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600.0
        else:
            dt_h = 1.0

        def energy_in(series):
            """MWh consumed at that port (sign-robust)."""
            e_pos = (series.clip(lower=0) * dt_h).sum()
            e_neg = ((-series).clip(lower=0) * dt_h).sum()
            return (
                (-series).clip(lower=0) * dt_h
                if e_neg >= e_pos
                else series.clip(lower=0) * dt_h
            )

        def energy_out(series):
            """MWh produced at that port (sign-robust)."""
            e_pos = (series.clip(lower=0) * dt_h).sum()
            e_neg = ((-series).clip(lower=0) * dt_h).sum()
            return (
                series.clip(lower=0) * dt_h
                if e_pos >= e_neg
                else (-series).clip(lower=0) * dt_h
            )

        rows = []

        for link in ft.index:
            try:
                region = net.buses.at[ft.at[link, "bus1"], "grid_region"]
            except KeyError:
                continue

            out_MWh = energy_out(net.links_t.p1[link]).sum()
            if out_MWh <= 0:
                continue

            elec_cons = energy_in(net.links_t.p3[link])  # electricity
            h2_cons = energy_in(net.links_t.p0[link])  # hydrogen
            co2_cons = energy_in(net.links_t.p2[link])  # CO2

            r_elec = elec_cons.sum() / out_MWh
            r_h2 = h2_cons.sum() / out_MWh
            r_co2 = co2_cons.sum() / out_MWh

            # --- electricity price ---
            if electricity_price == "marginal":
                p_elec = net.buses_t.marginal_price[ft.at[link, "bus3"]]
                avg_p_elec = (
                    (elec_cons * p_elec).sum() / elec_cons.sum()
                    if elec_cons.sum() > 0
                    else 0.0
                )
            elif electricity_price == "lcoe":
                if isinstance(lcoe_by_region, dict):
                    if not lcoe_by_region:
                        avg_p_elec = 0.0
                    else:
                        year = get_year(lcoe_by_region, scen_year)
                        avg_p_elec = lcoe_by_region[year].loc[region]
                else:
                    avg_p_elec = lcoe_by_region.loc[region]
            else:
                raise ValueError("electricity_price must be 'marginal' or 'lcoe'")

            # --- hydrogen price ---
            if hydrogen_price == "marginal":
                # Base marginal H2 price (USD/MWh_H2) EXCL. fees
                p_h2 = net.buses_t.marginal_price[ft.at[link, "bus0"]]
                avg_p_h2 = (
                    (h2_cons * p_h2).sum() / h2_cons.sum() if h2_cons.sum() > 0 else 0.0
                )

                # Add H2 fees EX-POST (transmission + baseload), same as LCOH accounting
                if lcoh_by_region:
                    y = get_year(lcoh_by_region, scen_year)
                    lcoh_df = lcoh_by_region[y].set_index("Grid Region")

                    col_excl = "LCOH (excl. T&D fees) (USD/kg H2)"
                    col_incl = "LCOH + Transmission fees + Baseload charges (USD/kg H2)"

                    if (
                        region in lcoh_df.index
                        and col_excl in lcoh_df.columns
                        and col_incl in lcoh_df.columns
                    ):
                        delta_h2_fees_usd_per_mwh = (
                            (
                                lcoh_df.at[region, col_incl]
                                - lcoh_df.at[region, col_excl]
                            )
                            * 1000.0
                            / 33.0
                        )
                        avg_p_h2 += delta_h2_fees_usd_per_mwh

            elif hydrogen_price == "lcoh":
                if not lcoh_by_region:
                    avg_p_h2 = 0.0
                else:
                    year = get_year(lcoh_by_region, scen_year)
                    avg_p_h2 = (
                        lcoh_by_region[year]
                        .set_index("Grid Region")
                        .at[
                            region,
                            "LCOH + Transmission fees + Baseload charges (USD/kg H2)",
                        ]
                        * 1000.0
                        / 33.0
                    )
            else:
                raise ValueError("hydrogen_price must be 'marginal' or 'lcoh'")

            # --- CO2 price ---
            if co2_price == "marginal":
                p_co2 = net.buses_t.marginal_price[ft.at[link, "bus2"]]
                avg_p_co2 = (
                    (co2_cons * p_co2).sum() / co2_cons.sum()
                    if co2_cons.sum() > 0
                    else 0.0
                )
            elif co2_price == "lcoc":
                if not lcoc_by_region or scen_year == 2023:
                    avg_p_co2 = 0.0
                else:
                    try:
                        lcoc_df = lcoc_by_region[str(scen_year)].set_index(
                            "Grid Region"
                        )
                    except KeyError:
                        if verbose:
                            print("Skipping scenario (no LCOC for year)")
                        rows = []
                        break
                    avg_p_co2 = (
                        lcoc_df.at[region, "LCOC incl. T&D fees (USD/tCO2)"]
                        if region in lcoc_df.index
                        else 0.0
                    )
            else:
                raise ValueError("co2_price must be 'marginal' or 'lcoc'")

            # --- cost components ---
            c_elec = avg_p_elec * r_elec * conv
            c_h2 = avg_p_h2 * r_h2 * conv
            c_co2 = avg_p_co2 * r_co2 * conv

            cap_cost = float(ft.at[link, "capital_cost"])
            cap_MW = float(
                ft.at[link, "p_nom_opt"]
                if ft.at[link, "p_nom_extendable"]
                else ft.at[link, "p_nom"]
            )
            c_capex = ((cap_cost * cap_MW) / out_MWh) * conv
            c_vom = vom_usd_for_year(scen_year) * conv

            lco_excl_TD = c_elec + c_h2 + c_co2 + c_capex + c_vom

            rows.append(
                {
                    "Grid Region": region,
                    "Production (TWh)": out_MWh / 1e6,
                    "Electricity rate (MWh el / MWh e-ker)": r_elec,
                    "H2 rate (MWh H2 / MWh e-ker)": r_h2,
                    "CO2 rate (tCO2 / MWh e-ker)": r_co2,
                    "Electricity price (USD/MWh el)": avg_p_elec,
                    "Hydrogen price (USD/MWh H2)": avg_p_h2,
                    "CO2 price (USD/tCO2)": avg_p_co2,
                    f"Electricity cost ({suffix})": c_elec,
                    f"Hydrogen cost ({suffix})": c_h2,
                    f"CO2 cost ({suffix})": c_co2,
                    f"CAPEX ({suffix})": c_capex,
                    f"VOM ({suffix})": c_vom,
                    f"LCO e-kerosene (excl. T&D fees) ({suffix})": lco_excl_TD,
                }
            )

        if not rows:
            continue

        df = pd.DataFrame(rows)
        if df["Production (TWh)"].sum() <= 1e-3:
            continue

        def wavg(group, col):
            return (group[col] * group["Production (TWh)"]).sum() / group[
                "Production (TWh)"
            ].sum()

        g = (
            df.groupby("Grid Region")
            .apply(
                lambda x: pd.Series(
                    {
                        "Production (TWh)": x["Production (TWh)"].sum(),
                        "Electricity rate (MWh el / MWh e-ker)": wavg(
                            x, "Electricity rate (MWh el / MWh e-ker)"
                        ),
                        "H2 rate (MWh H2 / MWh e-ker)": wavg(
                            x, "H2 rate (MWh H2 / MWh e-ker)"
                        ),
                        "CO2 rate (tCO2 / MWh e-ker)": wavg(
                            x, "CO2 rate (tCO2 / MWh e-ker)"
                        ),
                        "Electricity price (USD/MWh el)": wavg(
                            x, "Electricity price (USD/MWh el)"
                        ),
                        "Hydrogen price (USD/MWh H2)": wavg(
                            x, "Hydrogen price (USD/MWh H2)"
                        ),
                        "CO2 price (USD/tCO2)": wavg(x, "CO2 price (USD/tCO2)"),
                        f"Electricity cost ({suffix})": wavg(
                            x, f"Electricity cost ({suffix})"
                        ),
                        f"Hydrogen cost ({suffix})": wavg(
                            x, f"Hydrogen cost ({suffix})"
                        ),
                        f"CO2 cost ({suffix})": wavg(x, f"CO2 cost ({suffix})"),
                        f"CAPEX ({suffix})": wavg(x, f"CAPEX ({suffix})"),
                        f"VOM ({suffix})": wavg(x, f"VOM ({suffix})"),
                        f"LCO e-kerosene (excl. T&D fees) ({suffix})": wavg(
                            x, f"LCO e-kerosene (excl. T&D fees) ({suffix})"
                        ),
                    }
                )
            )
            .reset_index()
        )

        g = g[g["Production (TWh)"] > 0].copy()

        g["EMM Region"] = g["Grid Region"].map(emm_mapping)
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year,
            ["region", "Transmission nom USD/MWh", "Distribution nom USD/MWh"],
        ].set_index("region")

        g[f"Transmission fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Transmission nom USD/MWh"])
            * g["Electricity rate (MWh el / MWh e-ker)"]
            * conv
        )
        g[f"Distribution fees ({suffix})"] = (
            g["EMM Region"].map(fee_map["Distribution nom USD/MWh"])
            * g["Electricity rate (MWh el / MWh e-ker)"]
            * conv
        )

        g.drop(columns=["EMM Region"], inplace=True)

        g[f"LCO e-kerosene (incl. T&D fees) ({suffix})"] = (
            g[f"LCO e-kerosene (excl. T&D fees) ({suffix})"]
            + g[f"Transmission fees ({suffix})"]
            + g[f"Distribution fees ({suffix})"]
        )

        results[f"{scen_year if year_title else str(name)}"] = g

        if verbose:
            tot_prod = g["Production (TWh)"].sum()
            wavg_cost = (
                g[f"LCO e-kerosene (incl. T&D fees) ({suffix})"] * g["Production (TWh)"]
            ).sum() / tot_prod

            title = re.search(r"\d{4}", str(name)).group() if year_title else str(name)
            print(f"\n{title}:")
            print(
                f"Weighted average LCO e-kerosene (incl. T&D): {wavg_cost:.2f} {suffix}"
            )
            print(f"Total production: {tot_prod:.2f} TWh\n")

            numeric_cols = g.select_dtypes(include="number").columns
            fmt = {col: "{:.2f}" for col in numeric_cols}
            display(g.style.format(fmt).hide(axis="index"))

    return results


def compute_LCOC_by_region(
    networks: dict,
    regional_fees: pd.DataFrame,
    emm_mapping: dict,
    electricity_price: str = "marginal",  # or "lcoe"
    lcoe_by_region: pd.Series = None,  # required if electricity_price="lcoe"
    year_title: bool = True,
    captured_threshold_mt: float = 1e-6,
    verbose: bool = True,
):
    """
    Compute Levelized Cost of CO2 Capture (LCOC) by grid region.
    - Filters out CCS/DAC links with negligible electricity use.
    - Distinguishes DAC vs CCS flow directions.
    - Aggregates by grid region, weighted by captured CO2.
    - Units: USD/tCO2.
    """

    results = {}
    cc_carriers = {
        "ethanol from starch CC",
        "SMR CC",
        "DRI CC",
        "BF-BOF CC",
        "dry clinker CC",
        "DAC",
    }

    for name, net in networks.items():
        scen_year = int(re.search(r"\d{4}", str(name)).group())
        ccs = net.links[net.links.carrier.isin(cc_carriers)].copy()
        if ccs.empty:
            continue

        rows = []
        dt_h = (
            (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600.0
            if len(net.snapshots) > 1
            else 1.0
        )

        for link in ccs.index:
            carrier = str(ccs.at[link, "carrier"]).upper()
            captured = 0.0
            elec_series, elec_bus = None, None

            for j in range(6):
                col = f"p{j}"
                if col not in net.links_t or link not in net.links_t[col]:
                    continue
                series = net.links_t[col][link]
                bus = net.links.at[link, f"bus{j}"]

                # CO2 accounting
                if "co2" in bus.lower():
                    flow = series.sum()
                    if flow < 0:
                        captured += -flow * dt_h
                    continue

                # Electricity usage on AC buses
                if bus in net.buses.index and str(net.buses.at[bus, "carrier"]) == "AC":
                    if "DAC" in carrier:
                        cons = (-series).clip(lower=0) * dt_h
                    else:
                        cons = (series).clip(lower=0) * dt_h
                    if cons.sum() > 1e-9:  # electricity consumption threshold (MWh)
                        elec_series = cons
                        elec_bus = bus

            if captured <= 0:
                continue

            # Skip links with negligible electricity usage
            if elec_series is None or elec_series.sum() <= 1e-3:
                continue

            try:
                region = net.buses.at[ccs.at[link, "bus0"], "grid_region"]
            except KeyError:
                continue

            # CAPEX per tCO2
            cap_cost = float(ccs.at[link, "capital_cost"])
            cap_mw = float(
                ccs.at[link, "p_nom_opt"]
                if ccs.at[link, "p_nom_extendable"]
                else ccs.at[link, "p_nom"]
            )
            c_capex = (cap_cost * cap_mw) / captured

            # Electricity cost per tCO2
            elec_rate = elec_series.sum() / captured
            if electricity_price == "marginal":
                p_elec = net.buses_t.marginal_price[elec_bus]
                avg_p_elec = (elec_series * p_elec).sum() / elec_series.sum()
            elif electricity_price == "lcoe":
                avg_p_elec = lcoe_by_region.loc[region]
            else:
                raise ValueError("electricity_price must be 'marginal' or 'lcoe'")

            c_elec = avg_p_elec * elec_rate
            lco_excl = c_capex + c_elec

            rows.append(
                {
                    "Grid Region": region,
                    "Captured CO2 (Mt)": captured / 1e6,
                    "CAPEX (USD/tCO2)": c_capex,
                    "Electricity rate (MWh el / tCO2)": elec_rate,
                    "Electricity price (USD/MWh el)": avg_p_elec,
                    "Electricity cost (USD/tCO2)": c_elec,
                    "LCOC excl. T&D fees (USD/tCO2)": lco_excl,
                }
            )

        if not rows:
            continue

        df = pd.DataFrame(rows)

        def wavg(group, col):
            return (group[col] * group["Captured CO2 (Mt)"]).sum() / group[
                "Captured CO2 (Mt)"
            ].sum()

        g = (
            df.groupby("Grid Region")
            .apply(
                lambda x: pd.Series(
                    {
                        "Captured CO2 (Mt)": x["Captured CO2 (Mt)"].sum(),
                        "CAPEX (USD/tCO2)": wavg(x, "CAPEX (USD/tCO2)"),
                        "Electricity rate (MWh el / tCO2)": wavg(
                            x, "Electricity rate (MWh el / tCO2)"
                        ),
                        "Electricity price (USD/MWh el)": wavg(
                            x, "Electricity price (USD/MWh el)"
                        ),
                        "Electricity cost (USD/tCO2)": wavg(
                            x, "Electricity cost (USD/tCO2)"
                        ),
                        "LCOC excl. T&D fees (USD/tCO2)": wavg(
                            x, "LCOC excl. T&D fees (USD/tCO2)"
                        ),
                    }
                )
            )
            .reset_index()
        )

        g = g[g["Captured CO2 (Mt)"] > captured_threshold_mt]
        if g.empty:
            continue

        g["EMM Region"] = g["Grid Region"].map(emm_mapping)
        fee_map = regional_fees.loc[
            regional_fees["Year"] == scen_year,
            ["region", "Transmission nom USD/MWh", "Distribution nom USD/MWh"],
        ].set_index("region")

        g["Transmission fee (USD/MWh)"] = g["EMM Region"].map(
            fee_map["Transmission nom USD/MWh"]
        )
        g["Distribution fee (USD/MWh)"] = g["EMM Region"].map(
            fee_map["Distribution nom USD/MWh"]
        )

        g["Transmission cost (USD/tCO2)"] = (
            g["Transmission fee (USD/MWh)"] * g["Electricity rate (MWh el / tCO2)"]
        )
        g["Distribution cost (USD/tCO2)"] = (
            g["Distribution fee (USD/MWh)"] * g["Electricity rate (MWh el / tCO2)"]
        )

        g["LCOC incl. T&D fees (USD/tCO2)"] = (
            g["LCOC excl. T&D fees (USD/tCO2)"]
            + g["Transmission cost (USD/tCO2)"]
            + g["Distribution cost (USD/tCO2)"]
        )

        g.drop(columns=["EMM Region"], inplace=True)
        results[f"{scen_year if year_title else str(name)}"] = g

        if verbose:
            tot_captured = g["Captured CO2 (Mt)"].sum()
            wavg_cost = (
                g["LCOC incl. T&D fees (USD/tCO2)"] * g["Captured CO2 (Mt)"]
            ).sum() / tot_captured
            title = re.search(r"\d{4}", str(name)).group() if year_title else str(name)
            print(f"\nYear: {title}")
            print(f"Total captured CO2: {tot_captured:.2f} Mt")
            print(f"Weighted average LCOC (incl. T&D fees): {wavg_cost:.2f} USD/tCO2\n")
            numeric_cols = g.select_dtypes(include="number").columns
            fmt = {col: "{:.2f}" for col in numeric_cols}
            display(g.style.format(fmt).hide(axis="index"))

    return results


def calculate_LCOC_by_region(
    networks: dict,
    regional_fees: pd.DataFrame,
    emm_mapping: dict,
    electricity_price: str = "marginal",  # or "lcoe"
    lcoe_by_region: pd.Series = None,  # required if electricity_price="lcoe"
    year_title: bool = True,
    captured_threshold_mt: float = 1e-6,  # MtCO2 threshold
    verbose: bool = False,
) -> dict:
    """
    Lightweight wrapper around compute_LCOC_by_region.
    Returns the results dictionary {year: DataFrame}.

    If verbose=True, compute_LCOC_by_region will also print summaries.
    """
    return compute_LCOC_by_region(
        networks=networks,
        regional_fees=regional_fees,
        emm_mapping=emm_mapping,
        electricity_price=electricity_price,
        lcoe_by_region=lcoe_by_region,
        year_title=year_title,
        captured_threshold_mt=captured_threshold_mt,
        verbose=verbose,
    )


def save_to_excel_with_formatting(
    df, sheet_name, title, excel_file_path, freeze_pane="B3"
):
    # local import to parse column letters
    from openpyxl.utils import column_index_from_string

    with pd.ExcelWriter(
        excel_file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=1)

        # Get the worksheet for formatting
        worksheet = writer.sheets[sheet_name]

        # Add a title for df summary
        worksheet["A1"] = title
        worksheet["A1"].font = Font(bold=True, size=14, color="2F4F4F")
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        extra_col = df.index.nlevels
        max_col = len(df.columns) + extra_col  # include index columns

        if max_col > 1:
            worksheet.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=max_col
            )

        # Format headers (row 2: MultiIndex headers)
        header_fill = PatternFill(
            start_color="2F4F4F", end_color="2F4F4F", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_thin

        # Format data cells
        extra_row = getattr(df.columns, "nlevels", 1)
        max_row = len(df) + extra_row + 2  # +2 for title and headers
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="center", vertical="center")

        for row in range(3, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_thin

        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(2, min(max_row + 1, 100)):  # Sample first 100 rows
                try:
                    cell_value = str(worksheet.cell(row=row, column=col).value)
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Derive frozen rows/columns from freeze_pane (e.g., "B3" -> rows 1..2 and col 1 frozen)
        try:
            m = re.match(r"([A-Za-z]+)(\d+)", str(freeze_pane))
            freeze_col_idx = column_index_from_string(m.group(1)) if m else 2
            freeze_row_idx = int(m.group(2)) if m else 3
        except Exception:
            freeze_col_idx, freeze_row_idx = 2, 3  # sensible fallback for "B3"

        # Bold the frozen header rows (above the horizontal split), preserving existing header colors
        for r in range(1, max(1, freeze_row_idx)):
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                try:
                    cell.font = cell.font.copy(bold=True)
                except Exception:
                    cell.font = Font(
                        bold=True, size=(cell.font.sz if cell.font else 10)
                    )

        # Bold the frozen index columns (to the left of the vertical split) for data rows
        for r in range(3, max_row + 1):  # data area; header rows already bold
            for c in range(1, max(1, freeze_col_idx)):
                cell = worksheet.cell(row=r, column=c)
                try:
                    cell.font = cell.font.copy(bold=True)
                except Exception:
                    cell.font = Font(bold=True, size=data_font.size)

        # Freeze panes for better navigation
        worksheet.freeze_panes = worksheet[freeze_pane]


def compare_h2_kerosene_production(
    network, plot=True, network_name="Network", plot_threshold_gw=1e-3
):
    """
    Compare H2 and e-kerosene production from a PyPSA network.

    - Production time series are expressed in GW (daily average power).
    - Installed capacity is expressed in GW.
    - Total production is in MWh, computed using snapshot_weightings (accounts for non-hourly resolution).
    - The plot is only shown if at least one daily average exceeds plot_threshold_gw (GW).
    """

    # Define hydrogen carriers
    h2_carriers = [
        "Alkaline electrolyzer large",
        "Alkaline electrolyzer medium",
        "Alkaline electrolyzer small",
        "PEM electrolyzer",
        "SOEC",
    ]

    # FT (e-kerosene) links
    ft_links = network.links[
        network.links["carrier"].str.contains(
            "FT|Fischer|Tropsch", case=False, na=False
        )
        | network.links.index.str.contains("FT|Fischer|Tropsch", case=False, na=False)
    ].copy()

    # H2 links
    h2_links = network.links[network.links.carrier.isin(h2_carriers)].copy()

    # Helper: sum p1 across given links, robust to empty indices
    def sum_p1(idx):
        if len(idx) == 0:
            try:
                base = network.links_t.p1.iloc[:, :1].copy()
                base.iloc[:, 0] = 0.0
                return base.iloc[:, 0]
            except Exception:
                return pd.Series(dtype=float, index=network.snapshots)
        return np.multiply(-1, network.links_t.p1[idx]).sum(axis=1)

    # Productions in MW (per snapshot)
    h2_prod_mw = sum_p1(h2_links.index)
    kerosene_prod_mw = sum_p1(ft_links.index)

    # Daily averages in GW
    h2_prod_gw = h2_prod_mw.resample("D").mean() / 1e3
    kerosene_prod_gw = kerosene_prod_mw.resample("D").mean() / 1e3

    # Snapshot durations (in hours)
    weights = network.snapshot_weightings["generators"]

    # Summaries: total production in MWh, installed capacity in GW
    h2_summary = {
        "total_production_MWh": (h2_prod_mw * weights).sum(),
        "installed_capacity_GW": (
            h2_links.p_nom_opt.sum()
            if "p_nom_opt" in h2_links
            else h2_links.p_nom.sum()
        )
        / 1e3,
    }
    kerosene_summary = {
        "total_production_MWh": (kerosene_prod_mw * weights).sum(),
        "installed_capacity_GW": (
            ft_links.p_nom_opt.sum()
            if "p_nom_opt" in ft_links
            else ft_links.p_nom.sum()
        )
        / 1e3,
    }

    # Comparison table
    comparison_data = {
        "Metric": ["Total Production (MWh)", "Installed Capacity (GW)"],
        "Hydrogen": [
            h2_summary["total_production_MWh"],
            h2_summary["installed_capacity_GW"],
        ],
        "E-Kerosene": [
            kerosene_summary["total_production_MWh"],
            kerosene_summary["installed_capacity_GW"],
        ],
    }
    comparison_table = pd.DataFrame(comparison_data)

    # Plot in GW if above threshold
    if plot:
        avg_h2_gw = h2_prod_gw.mean()
        avg_kerosene_gw = kerosene_prod_gw.mean()
        if (avg_h2_gw > plot_threshold_gw) or (avg_kerosene_gw > plot_threshold_gw):
            fig, ax = plt.subplots(figsize=(15, 5))
            h2_prod_gw.plot(ax=ax, label="Hydrogen production", alpha=0.8)
            kerosene_prod_gw.plot(ax=ax, label="E-Kerosene production", alpha=0.8)

            ax.set_title(f"Hydrogen vs e-kerosene Production (GW) - {network_name}")
            ax.set_xlabel("")
            ax.set_ylabel("Production (GW)")

            # x-axis formatting: monthly ticks with abbreviated month names
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))

            # legend outside
            ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", borderaxespad=0.0)
            ax.grid(True, alpha=0.3)
            plt.tight_layout(rect=[0, 0, 0.85, 1])  # leave space for legend
            showfig()
        else:
            threshold_mw = plot_threshold_gw * 1e3
            print(
                f"\nSkipped {network_name}: both daily-average productions are below {threshold_mw:.2f} MW (threshold = {plot_threshold_gw:.3e} GW).\n"
            )

    return {
        # Series in GW (daily average)
        "h2_production": h2_prod_gw,
        # Series in GW (daily average)
        "kerosene_production": kerosene_prod_gw,
        "h2_summary": h2_summary,  # Totals in MWh, capacity in GW
        "kerosene_summary": kerosene_summary,  # Totals in MWh, capacity in GW
        "comparison_table": comparison_table,
    }


def compute_capacity_factor_electrolysis(
    networks: dict,
    p_nom_threshold: float = 1.0,  # MW
    year_title: bool = True,
    round_digits: int = 2,
    output_threshold: float = 1.0,  # MWh
):
    """
    Compute annual capacity factor of H2 electrolysers by grid region.

    Definition:
    - Capacity factor (%) = Electricity input at bus0 (MWh) /
                            (Installed capacity at bus0 (MW) * total weighted hours)

    Notes:
    - p_nom / p_nom_opt always refer to bus0 (PyPSA convention).
    - p0 (bus0) = electricity input (MW, positive).
    - p1 (bus1) = hydrogen output (MW, negative) → reported only.
    """

    # Identify electrolyser carriers automatically
    h2_carriers = [
        c
        for c in pd.unique(pd.concat([n.links.carrier for n in networks.values()]))
        if "electrolyzer" in c.lower() or "soec" in c.lower()
    ]

    results = {}

    for year_key, n in networks.items():
        # extract year
        m = re.search(r"\d{4}", str(year_key))
        if not m:
            continue
        scen_year = int(m.group())
        key = str(scen_year) if year_title else year_key

        # select electrolysers
        links = n.links[n.links.carrier.isin(h2_carriers)].copy()
        if links.empty:
            continue

        # installed capacity (MW, bus0 reference)
        cap_series = pd.Series(
            np.where(
                links.get("p_nom_extendable", False),
                links.get("p_nom_opt", 0.0),
                links.get("p_nom", 0.0),
            ),
            index=links.index,
        ).astype(float)

        # filter by capacity threshold
        links = links[cap_series > p_nom_threshold]
        cap_series = cap_series.loc[links.index]
        if links.empty:
            continue

        # snapshot weighting
        w = n.snapshot_weightings.generators
        total_hours = float(w.sum())

        # electricity input (MWh, bus0)
        p0 = n.links_t.p0[links.index]
        energy_in = (p0).clip(lower=0).multiply(w, axis=0).sum(axis=0)

        # hydrogen output (MWh, bus1, for reporting only)
        p1 = n.links_t.p1[links.index]
        energy_out = (-p1).clip(lower=0).multiply(w, axis=0).sum(axis=0)

        # add grid region info
        links["grid_region"] = links["bus1"].map(n.buses["grid_region"])
        df_links = pd.DataFrame(
            {
                "Link": links.index,
                "Grid Region": links["grid_region"],
                "Capacity (MW)": cap_series,
                "Electricity input (MWh)": energy_in,
                "Hydrogen output (MWh)": energy_out,
            }
        ).dropna()

        # compute capacity factor per link
        df_links["Capacity factor (%)"] = (
            df_links["Electricity input (MWh)"]
            / (df_links["Capacity (MW)"] * total_hours)
            * 100
        )

        # aggregate by region
        region_summary = (
            df_links.groupby("Grid Region")
            .apply(
                lambda g: pd.Series(
                    {
                        "Capacity (GW input electricity)": g["Capacity (MW)"].sum()
                        / 1e3,
                        "Electricity input (MWh)": g["Electricity input (MWh)"].sum(),
                        "Hydrogen output (MWh)": g["Hydrogen output (MWh)"].sum(),
                        "Capacity factor (%)": (
                            g["Electricity input (MWh)"].sum()
                            / (g["Capacity (MW)"].sum() * total_hours)
                            * 100
                        ),
                    }
                )
            )
            .reset_index()
        )

        # skip if negligible production
        if region_summary["Hydrogen output (MWh)"].sum() < output_threshold:
            continue

        results[key] = {
            "links": df_links.round(round_digits),
            "regions": region_summary.round(round_digits),
        }

    return results


def _get_snapshot_weighting_series(network):
    """Return a snapshot weighting series aligned with the network snapshots."""

    weightings = getattr(network, "snapshot_weightings", None)
    if weightings is None:
        return pd.Series(1.0, index=network.snapshots)

    if isinstance(weightings, pd.Series):
        return weightings.reindex(network.snapshots, fill_value=0.0)

    for column in ["objective", "generators", "stores", "weights"]:
        if column in weightings.columns:
            return weightings[column].reindex(network.snapshots, fill_value=0.0)

    return weightings.iloc[:, 0].reindex(network.snapshots, fill_value=0.0)


def _get_aviation_loads_with_energy(network):
    """Collect aviation loads enriched with geographic metadata and annual energy (TWh)."""

    if "carrier" not in network.loads:
        return pd.DataFrame()

    aviation_mask = network.loads["carrier"].str.contains(
        "kerosene for aviation", case=False, na=False
    )
    aviation_loads = network.loads[aviation_mask].copy()

    if aviation_loads.empty:
        return pd.DataFrame()

    weights = _get_snapshot_weighting_series(network)
    loads_p = network.loads_t.p[aviation_loads.index]
    loads_p = loads_p.reindex(weights.index).fillna(0.0)

    energy_mwh = loads_p.mul(weights, axis=0).sum()
    average_mw = energy_mwh / weights.sum() if weights.sum() else energy_mwh
    peak_mw = loads_p.max()

    buses_info = network.buses[["state", "grid_region", "x", "y"]]
    aviation_loads = aviation_loads.join(buses_info, on="bus", how="left")
    aviation_loads["energy_MWh"] = energy_mwh
    aviation_loads["energy_TWh"] = aviation_loads["energy_MWh"] / 1e6

    return aviation_loads


def _aggregate_aviation_demand(network, level):
    """Aggregate aviation demand by the requested spatial level (state or grid_region)."""

    loads = _get_aviation_loads_with_energy(network)
    if loads.empty or level not in loads.columns:
        return pd.DataFrame(columns=[level, "energy_TWh", "lon", "lat", "share_pct"])

    loads = loads.dropna(subset=[level])
    if loads.empty:
        return pd.DataFrame(columns=[level, "energy_TWh", "lon", "lat", "share_pct"])

    aggregated = (
        loads.groupby(level)
        .agg(
            energy_TWh=("energy_TWh", "sum"),
            lon=("x", "mean"),
            lat=("y", "mean"),
        )
        .reset_index()
    )

    total_twh = aggregated["energy_TWh"].sum()
    if total_twh:
        aggregated["share_pct"] = aggregated["energy_TWh"] / total_twh * 100
    else:
        aggregated["share_pct"] = 0.0

    aggregated = aggregated.sort_values("energy_TWh", ascending=False)
    return aggregated


def compute_aviation_demand_table(network, level="state"):
    """Return a tidy aviation demand table for the requested aggregation level."""

    aggregation = _aggregate_aviation_demand(network, level)

    label = {"state": "State", "grid_region": "Grid Region"}.get(level, level.title())

    if aggregation.empty:
        return pd.DataFrame(columns=[label, "Fuel demand (TWh)", "Share (%)"])

    table = aggregation[[level, "energy_TWh", "share_pct"]].copy()
    table = table.rename(
        columns={
            level: label,
            "energy_TWh": "Fuel demand (TWh)",
            "share_pct": "Share (%)",
        }
    )

    return table


def create_aviation_demand_by_state_map(
    network,
    path_shapes,
    network_name="Network",
    distance_crs=4326,
    min_demand_twh=1.0,
    year_title=True,
):
    """Plot aviation demand aggregated by state (TWh/year) as scaled circles.

    Alaska and Hawaii remain in the returned tables, but they are omitted from the map.
    """

    state_df = _aggregate_aviation_demand(network, "state")
    if state_df.empty:
        print(f"No aviation loads found in the network: {network_name}")
        return None, None, state_df

    states_to_plot = state_df[state_df["energy_TWh"] >= min_demand_twh]
    if states_to_plot.empty:
        print(
            f"Aviation demand below {min_demand_twh} TWh for all states in {network_name}."
        )
        return None, None, state_df

    contiguous_mask = states_to_plot["lon"].between(
        -130, -65, inclusive="both"
    ) & states_to_plot["lat"].between(20, 50, inclusive="both")
    states_to_plot_contiguous = states_to_plot[contiguous_mask]
    if states_to_plot_contiguous.empty:
        print(
            "No contiguous US states meet the plotting threshold; tables still include all states."
        )
        return None, None, state_df

    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes = shapes.to_crs(epsg=4326)
    bbox = box(-130, 20, -65, 50)
    shapes_clip = shapes.clip(bbox)

    fig, ax = plt.subplots(
        figsize=(12, 6), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    shapes_clip.plot(
        ax=ax, facecolor="whitesmoke", edgecolor="gray", alpha=0.7, linewidth=0.5
    )

    pie_scale = 0.02
    min_radius = 0.1
    max_radius = 3.5

    for _, row in states_to_plot_contiguous.iterrows():
        x, y = row["lon"], row["lat"]
        if pd.isna(x) or pd.isna(y):
            continue

        radius = np.clip(row["energy_TWh"] * pie_scale, min_radius, max_radius)
        circle = plt.Circle(
            (x, y),
            radius,
            facecolor="#1f77b4",
            edgecolor="gray",
            alpha=0.65,
            linewidth=1,
            transform=ccrs.PlateCarree(),
            zorder=4,
        )
        ax.add_patch(circle)

        ax.text(
            x,
            y - radius - 0.3,
            f"{row['energy_TWh']:.1f} TWh",
            ha="center",
            va="top",
            fontsize=9,
            bbox=dict(facecolor="white", edgecolor="gray", boxstyle="round,pad=0.2"),
            transform=ccrs.PlateCarree(),
        )

    year_match = re.search(r"\d{4}", network_name)
    year_str = f" – {year_match.group()}" if year_match else ""

    ax.set_extent([-130, -65, 20, 50], crs=ccrs.PlateCarree())
    ax.set_title(
        f"Aviation fuel demand by State (TWh/year){year_str if year_title else f' – {network_name}'}",
        fontsize=12,
        pad=20,
    )
    ax.axis("off")
    plt.tight_layout()

    return fig, ax, state_df


def create_aviation_demand_by_grid_region_map(
    network,
    path_shapes,
    network_name="Network",
    distance_crs=4326,
    min_demand_twh=5.0,
    year_title=True,
):
    """Plot aviation demand aggregated by grid region (TWh/year) as scaled circles."""

    region_df = _aggregate_aviation_demand(network, "grid_region")
    if region_df.empty:
        print(
            f"No aviation loads with grid region found in the network: {network_name}"
        )
        return None, None, region_df

    regions_to_plot = region_df[region_df["energy_TWh"] >= min_demand_twh]
    if regions_to_plot.empty:
        print(
            f"Aviation demand below {min_demand_twh} TWh for all regions in {network_name}."
        )
        return None, None, region_df

    contiguous_mask = regions_to_plot["lon"].between(
        -130, -65, inclusive="both"
    ) & regions_to_plot["lat"].between(20, 50, inclusive="both")
    regions_to_plot_contiguous = regions_to_plot[contiguous_mask]
    if regions_to_plot_contiguous.empty:
        print(
            "No contiguous US grid regions meet the plotting threshold; tables still include all regions."
        )
        return None, None, region_df

    shapes = gpd.read_file(path_shapes, crs=distance_crs)
    shapes = shapes.to_crs(epsg=4326)
    bbox = box(-130, 20, -60, 50)
    shapes_clip = shapes.clip(bbox)

    fig, ax = plt.subplots(
        figsize=(12, 6), subplot_kw={"projection": ccrs.PlateCarree()}
    )
    shapes_clip.plot(
        ax=ax, facecolor="whitesmoke", edgecolor="gray", alpha=0.7, linewidth=0.5
    )

    pie_scale = 0.03
    min_radius = 0.15
    max_radius = 3.8

    for _, row in regions_to_plot_contiguous.iterrows():
        x, y = row["lon"], row["lat"]
        if pd.isna(x) or pd.isna(y):
            continue

        radius = np.clip(row["energy_TWh"] * pie_scale, min_radius, max_radius)
        circle = plt.Circle(
            (x, y),
            radius,
            facecolor="#1f77b4",
            edgecolor="gray",
            alpha=0.65,
            linewidth=1,
            transform=ccrs.PlateCarree(),
            zorder=4,
        )
        ax.add_patch(circle)

        ax.text(
            x,
            y - radius - 0.3,
            f"{row['energy_TWh']:.1f} TWh",
            ha="center",
            va="top",
            fontsize=9,
            bbox=dict(facecolor="white", edgecolor="gray", boxstyle="round,pad=0.2"),
            transform=ccrs.PlateCarree(),
        )

    year_match = re.search(r"\d{4}", network_name)
    year_str = f" – {year_match.group()}" if year_match else ""

    ax.set_extent([-130, -65, 20, 50], crs=ccrs.PlateCarree())
    ax.set_title(
        f"Aviation fuel demand by Grid Region (TWh/year){year_str if year_title else f' – {network_name}'}",
        fontsize=12,
        pad=20,
    )
    ax.axis("off")
    plt.tight_layout()

    return fig, ax, region_df


def calculate_demand_profile(network):
    """Calculate total electricity demand profile from network."""
    # Industrial processes consuming AC electricity
    target_processes = [
        "SMR CC",
        "Haber-Bosch",
        "ethanol from starch",
        "ethanol from starch CC",
        "DRI",
        "DRI CC",
        "DRI H2",
        "BF-BOF",
        "BF-BOF CC",
        "EAF",
        "dry clinker",
        "cement finishing",
        "dry clinker CC",
    ]

    # Static and dynamic loads
    static_load_carriers = [
        "rail transport electricity",
        "agriculture electricity",
        "industry electricity",
    ]

    # Static loads (constant profile)
    static_totals = (
        network.loads.groupby("carrier")
        .sum()
        .p_set.reindex(static_load_carriers)
        .fillna(0)
    )
    static_sum = static_totals.sum()  # MW
    static_profile = pd.Series(static_sum, index=network.snapshots)

    # Industrial AC consumption
    process_links = network.links[network.links.carrier.isin(target_processes)]
    ac_input_links = process_links[
        process_links.bus0.map(network.buses.carrier) == "AC"
    ].index
    ind_ac_profile = (
        network.links_t.p0[ac_input_links].sum(axis=1) if len(ac_input_links) > 0 else 0
    )

    # Non-industrial AC loads
    ac_loads = network.loads[network.loads.carrier == "AC"]
    industrial_ac_buses = (
        network.links.loc[ac_input_links, "bus0"].unique()
        if len(ac_input_links) > 0
        else []
    )
    ac_non_ind_idx = ac_loads[~ac_loads.bus.isin(industrial_ac_buses)].index
    ac_profile = network.loads_t.p_set[
        ac_non_ind_idx.intersection(network.loads_t.p_set.columns)
    ].sum(axis=1)

    # Services and EVs
    serv_idx = [
        i
        for i in network.loads[network.loads.carrier == "services electricity"].index
        if i in network.loads_t.p_set.columns
    ]
    ev_idx = [
        i
        for i in network.loads[network.loads.carrier == "land transport EV"].index
        if i in network.loads_t.p_set.columns
    ]
    serv_profile = network.loads_t.p_set[serv_idx].sum(axis=1) if serv_idx else 0
    ev_profile = network.loads_t.p_set[ev_idx].sum(axis=1) if ev_idx else 0

    # Data centers (constant profile)
    data_center_sum = network.loads.loc[
        network.loads.carrier == "data center", "p_set"
    ].sum()
    dc_profile = pd.Series(data_center_sum, index=network.snapshots)

    # Other electricity
    other_idx = [
        i
        for i in network.loads[network.loads.carrier == "other electricity"].index
        if i in network.loads_t.p_set.columns
    ]
    other_profile = network.loads_t.p_set[other_idx].sum(axis=1) if other_idx else 0

    # Total demand profile (convert to GW, keep positive for plotting)
    total_demand = (
        static_profile
        + abs(ind_ac_profile)
        + ac_profile
        + serv_profile
        + ev_profile
        + dc_profile
        + other_profile
    ) / 1000

    return total_demand


def plot_electricity_dispatch(
    networks, tech_colors, nice_names, title_year=True, return_data=False
):
    """
    Plot electricity dispatch with demand for multiple networks with stacked area charts.

    Parameters:
    -----------
    networks : dict
        Dictionary of network names to PyPSA network objects
    tech_colors : dict
        Technology color mapping
    nice_names : dict
        Technology name mapping for legend
    title_year : bool, default True
        If True, extract year from key for title. If False, use full key name
    return_data : bool, default False
        Whether to return the collected dispatch tables and demand data

    Returns:
    --------
    dict or None
        If return_data=True, returns dict with 'dispatch' and 'demand' keys
    """

    collected_dispatch_tables = {}
    collected_demand_tables = {}
    summary_list = []
    max_y = 0

    # Calculate dispatch and demand for all networks and find global max
    for key, n in networks.items():
        total_gwh, supply_gw = calculate_dispatch(n)
        demand_profile = calculate_demand_profile(n)

        summary_list.append({"Network": key, "Total Dispatch (GWh)": total_gwh})
        max_y = max(max_y, supply_gw.sum(axis=1).max(), demand_profile.max())

    # Add some margin
    ymax = max_y * 1.05

    # Create subplots
    fig, axes = plt.subplots(len(networks), 1, figsize=(22, 5 * len(networks)))

    # Handle single network case
    if len(networks) == 1:
        axes = [axes]

    # Define technology order
    ordered_columns = [
        "nuclear",
        "coal",
        "biomass",
        "CCGT",
        "OCGT",
        "oil",
        "hydro",
        "ror",
        "geothermal",
        "gas CHP",
        "biomass CHP",
        "solar",
        "solar rooftop",
        "csp",
        "onwind",
        "offwind-ac",
        "offwind-dc",
        "battery discharger",
    ]

    # Plot each network
    for ax, (key, n) in zip(axes, networks.items()):
        # Calculate dispatch
        _, supply_gw = calculate_dispatch(n)
        supply_gw.index = pd.to_datetime(supply_gw.index)
        supply_gw = supply_gw.resample("24H").mean()

        # Calculate demand
        demand_profile = calculate_demand_profile(n)
        demand_profile.index = pd.to_datetime(demand_profile.index)
        demand_daily = demand_profile.resample("24H").mean()

        # Filter and order columns
        supply_gw = supply_gw[[c for c in ordered_columns if c in supply_gw.columns]]
        collected_dispatch_tables[key] = supply_gw
        collected_demand_tables[key] = demand_daily

        # Create stacked area plot for generation
        supply_gw.plot.area(
            ax=ax,
            stacked=True,
            linewidth=0,
            color=[tech_colors.get(c, "gray") for c in supply_gw.columns],
            legend=False,
        )

        # Plot demand as line (positive values)
        demand_daily.plot(
            ax=ax,
            color="red",
            linewidth=2,
            linestyle="-",
            label="Total Demand",
            alpha=0.8,
        )

        # Add horizontal line at zero
        ax.axhline(y=0, color="black", linewidth=1.5, linestyle="-", alpha=0.8)

        # Set title based on title_year parameter
        if title_year:
            year = key[-4:]  # Extract the year
            title = f"Electricity Dispatch & Demand – {year}"
        else:
            title = f"Electricity Dispatch & Demand – {key}"

        ax.set_title(title)
        ax.set_ylabel("Power (GW)")
        ax.set_ylim(0, ymax)
        ax.grid(True, alpha=0.3)

        # Set x-axis formatting
        start = supply_gw.index.min().replace(day=1)
        end = supply_gw.index.max()
        month_starts = pd.date_range(start=start, end=end, freq="MS")

        ax.set_xlim(start, end)
        ax.set_xticks(month_starts)
        ax.set_xticklabels(month_starts.strftime("%b"))
        ax.tick_params(axis="x", which="both", labelbottom=True)

        # Create legend for technologies with non-zero values + demand
        handles, labels = ax.get_legend_handles_labels()
        sums = supply_gw.sum()

        # Filter out zero generation technologies but keep demand
        filtered = [
            (h, l)
            for h, l in zip(handles, labels)
            if sums.get(l, 0) > 0 or l == "Total Demand"
        ]

        if filtered:
            handles, labels = zip(*filtered)
            pretty_labels = [
                nice_names.get(label, label) if label != "Total Demand" else label
                for label in labels
            ]

            ax.legend(
                handles,
                pretty_labels,
                loc="center left",
                bbox_to_anchor=(1.02, 0.5),
                title="Technology",
                fontsize="small",
                title_fontsize="medium",
            )

    # Set x-label for bottom subplot
    axes[-1].set_xlabel("Time (months)")
    plt.tight_layout(rect=[0, 0.05, 0.80, 1])
    showfig()

    # Return data if requested
    if return_data:
        return {
            "dispatch": collected_dispatch_tables,
            "demand": collected_demand_tables,
        }


def extract_marginal_price_by_grid_region_weighted(n):
    """
    Extract load-weighted marginal prices per grid_region for electricity buses.
    """
    # 1. Filter electricity buses
    elec_buses = n.buses[n.buses.carrier == "AC"].index

    # 2. Pre‐extract DataFrames
    prices = n.buses_t.marginal_price[elec_buses]
    loads = n.loads_t.p[elec_buses]
    bus2region = n.buses.loc[elec_buses, "grid_region"]

    # 3. Prepare output
    snapshots = prices.index
    regions = bus2region.unique()
    df = pd.DataFrame(index=snapshots, columns=regions, dtype=float)

    # 4. Loop snapshots × region and compute weighted mean
    for snap in snapshots:
        p = prices.loc[snap]
        l = loads.loc[snap].fillna(0.0)
        for region in regions:
            buses = bus2region[bus2region == region].index
            pr = p[buses]
            ld = l[buses]
            if ld.sum() > 0:
                df.at[snap, region] = (pr * ld).sum() / ld.sum()
            else:
                df.at[snap, region] = pr.mean()
    return df


def plot_marginal_prices_by_region_weighted(
    region_prices_df,
    network,
    plot=True,
    network_name="Network",
    demand_threshold=0.01,
    year_title=True,
):
    """
    Plot load-weighted marginal prices by grid region (daily averages),
    excluding regions whose total annual demand is below threshold.
    """
    if plot:
        # Compute total annual demand per region
        elec_buses = network.buses[network.buses.carrier == "AC"].index
        loads = network.loads_t.p[elec_buses]
        bus2region = network.buses.loc[elec_buses, "grid_region"]
        demand_by_region = (
            loads.groupby(bus2region, axis=1)
            .sum()  # snapshots×regions
            .sum()  # annual total per region
        )
        system_demand = demand_by_region.sum()
        keep = demand_by_region >= demand_threshold * system_demand
        df = region_prices_df.loc[:, keep.index[keep]]

        # Resample to daily mean
        df_daily = df.resample("D").mean()

        # Plot
        fig, ax = plt.subplots(figsize=(18, 6))
        colors = [
            "#1f77b4",
            "#ff7f0e",
            "#2ca02c",
            "#d62728",
            "#9467bd",
            "#8c564b",
            "#e377c2",
            "#7f7f7f",
            "#bcbd22",
            "#17becf",
            "#aec7e8",
            "#ffbb78",
            "#98df8a",
        ]
        for i, region in enumerate(df_daily.columns):
            ax.plot(
                df_daily.index,
                df_daily[region],
                label=region,
                color=colors[i % len(colors)],
                alpha=0.6,
                linewidth=1.2,
            )

        # Title with year
        year_match = re.search(r"\d{4}", network_name)
        year = year_match.group() if year_match else network_name
        ax.set_title(
            f"Electricity Marginal Prices by Grid Region (USD/MWh) - {year if year_title else network_name}"
        )
        ax.set_ylabel("Marginal Price (USD/MWh)")
        ax.set_xlabel("")

        # One tick per month
        locator = mdates.MonthLocator(bymonthday=1)
        formatter = mdates.DateFormatter("%b")
        ax.xaxis.set_major_locator(locator)
        ax.xaxis.set_major_formatter(formatter)
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right")

        # Limit axes
        ax.set_xlim(df_daily.index[0], df_daily.index[-1])
        y0 = df_daily.min().min() * 0.95
        y1 = df_daily.max().max() * 1.05
        ax.set_ylim(y0, y1)

        # Legend & grid
        ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=10)
        ax.grid(alpha=0.3)
        plt.tight_layout(rect=[0, 0, 0.85, 1])
        showfig()

    return df_daily


def calculate_baseload_charge(
    networks: dict,
    h2_carriers: list,
    emm_mapping: dict,
    energy_charge_path: str = "./data/energy_charge_rate.csv",
    customer_charge_mw: float = 400.0,
    demand_charge_rate: float = 9.0,
    baseload_percentages: dict = None,
    output_threshold: float = 0.01,
    verbose: bool = True,
    year_title: bool = True,
    customer_charge_usd: float = 1000.0,
):
    """
    Calculate baseload charges for hydrogen production by aggregating all carriers at regional level.

    Parameters
    ----------
    networks : dict
        Dictionary of PyPSA networks
    h2_carriers : list
        List of hydrogen production carrier names
    emm_mapping : dict
        Mapping from grid regions to EMM regions
    energy_charge_path : str
        Path to CSV with energy rates
    customer_charge_mw : float
        Representative plant capacity (MW)
    demand_charge_rate : float
        Demand charge rate ($/kW)
    baseload_percentages : dict
        Baseload % by electrolyzer type
    output_threshold : float
        Minimum H2 output (TWh) to include
    verbose : bool
        Print progress information
    year_title : bool, default True
        If True, group results by extracted year (e.g., 2030).
        If False, group results by full network key name.
    """

    if baseload_percentages is None:
        baseload_percentages = {
            "Alkaline electrolyzer large": 10.0,
            "Alkaline electrolyzer medium": 10.0,
            "Alkaline electrolyzer small": 10.0,
            "PEM electrolyzer": 10.0,
            "SOEC": 3.0,
        }

    # Convert cents/kWh to $/MWh and adjust for inflation (2021 to 2020)
    # 2021 cents/kWh * 10 = 2021 $/MWh
    # Divide by 1.053 to adjust from 2021 to 2020 (5.3% inflation)
    energy_rates = pd.read_csv(energy_charge_path)
    energy_rates["Year"] = pd.to_numeric(energy_rates["Year"], errors="coerce")
    energy_rates["Generation USD/MWh 2020"] = (
        energy_rates["Generation 2021 cents/kWh"] * 10 / 1.053
    )

    results = {}

    for net_key, net in networks.items():
        year_match = re.search(r"\d{4}", net_key)
        if not year_match:
            continue
        year = int(year_match.group())

        # Determine the key for results based on year_title
        key = year if year_title else net_key

        if verbose:
            print(f"\nProcessing {net_key} (Year: {year})")

        w = net.snapshot_weightings.generators
        hours_per_month = 730

        # Filter H2 production links
        h2_links = net.links[net.links.carrier.isin(h2_carriers)].copy()

        if h2_links.empty:
            continue

        h2_links["grid_region"] = h2_links.bus0.map(net.buses.grid_region)
        h2_links["emm_region"] = h2_links.grid_region.map(emm_mapping)

        # Calculate electricity input and H2 output
        p0 = net.links_t.p0[h2_links.index].clip(lower=0)
        p1 = -net.links_t.p1[h2_links.index].clip(upper=0)

        elec_in_mwh = p0.multiply(w, axis=0).sum()
        h2_out_mwh = p1.multiply(w, axis=0).sum()
        h2_out_twh = h2_out_mwh / 1e6

        region_results = []

        # Process each region
        for region in h2_links.grid_region.unique():
            region_links = h2_links[h2_links.grid_region == region]

            # Validate EMM mapping
            emm_region = emm_mapping.get(region, None)
            if emm_region is None:
                if verbose:
                    print(f"{region}: No EMM mapping")
                continue

            # Get energy rate
            rate_row = energy_rates[
                (energy_rates["Year"] == year) & (energy_rates["region"] == emm_region)
            ]

            if rate_row.empty:
                if verbose:
                    print(f"{region}: No energy rate for {emm_region}")
                continue

            energy_rate_usd_mwh = rate_row["Generation USD/MWh 2020"].iloc[0]

            # Aggregate all carriers in the region
            total_capacity_mw = region_links.p_nom_opt.sum()
            annual_h2_twh = h2_out_twh[region_links.index].sum()

            # Apply thresholds
            if annual_h2_twh < output_threshold:
                if verbose:
                    print(
                        f"{region}: H2 output {annual_h2_twh:.3f} TWh < {output_threshold} TWh"
                    )
                continue

            n_plants = int(np.floor(total_capacity_mw / customer_charge_mw))

            if n_plants == 0:
                if verbose:
                    print(
                        f"{region}: Total capacity {total_capacity_mw:.1f} MW < {customer_charge_mw} MW"
                    )
                continue

            # Calculate weighted average baseload percentage
            carrier_capacities = region_links.groupby("carrier").p_nom_opt.sum()
            weighted_baseload_pct = (
                sum(
                    baseload_percentages.get(c, 10.0) * carrier_capacities[c]
                    for c in carrier_capacities.index
                )
                / total_capacity_mw
                / 100.0
            )

            baseload_power_mw = total_capacity_mw * weighted_baseload_pct
            baseload_power_kw = baseload_power_mw * 1000

            # Calculate charges
            customer_charge_monthly = customer_charge_usd * n_plants
            demand_charge_monthly = demand_charge_rate * baseload_power_kw
            monthly_baseload_consumption_mwh = baseload_power_mw * hours_per_month
            energy_charge_monthly = (
                monthly_baseload_consumption_mwh * energy_rate_usd_mwh
            )

            total_monthly_charge = (
                customer_charge_monthly + demand_charge_monthly + energy_charge_monthly
            )

            total_annual_charge = total_monthly_charge * 12
            baseload_cost_per_mwh_h2 = total_annual_charge / (annual_h2_twh * 1e6)

            if verbose:
                print(
                    f"{region}: {n_plants} plants, "
                    f"{annual_h2_twh:.2f} TWh H2, "
                    f"${baseload_cost_per_mwh_h2:.2f}/MWh H2"
                )

            region_results.append(
                {
                    "grid_region": region,
                    "emm_region": emm_region,
                    "carrier": "All carriers (aggregated)",
                    "capacity_mw": total_capacity_mw,
                    "n_plants": n_plants,
                    "baseload_pct": weighted_baseload_pct * 100,
                    "baseload_power_mw": baseload_power_mw,
                    "annual_h2_output_twh": annual_h2_twh,
                    "energy_rate_usd_mwh": energy_rate_usd_mwh,
                    "customer_charge_monthly": customer_charge_monthly,
                    "demand_charge_monthly": demand_charge_monthly,
                    "energy_charge_monthly": energy_charge_monthly,
                    "total_monthly_charge": total_monthly_charge,
                    "total_annual_charge": total_annual_charge,
                    "baseload_cost_per_mwh_h2": baseload_cost_per_mwh_h2,
                }
            )

        if region_results:
            df = pd.DataFrame(region_results)
            results[key] = df

            if verbose:
                print(f"\n{key} Summary:")
                print(f"  • Total regions included: {df.grid_region.nunique()}")
                print(f"  • Total plants: {df.n_plants.sum():.0f}")
                print(
                    f"  • Total annual baseload charges: ${df.total_annual_charge.sum():,.0f}"
                )

    return results


def compute_power_opex_with_tax_credits(network, name_tag):
    """
    Compute OPEX for power generation, including tax credits and input costs.
    Returns a DataFrame with:
    - tech_label
    - without_tax_credits_billion
    - tax_credits_billion
    - with_tax_credits_billion
    - year, scenario
    """

    year_str = name_tag[-4:]
    fossil_carriers = ["coal", "gas", "oil", "biomass"]
    results = []
    w = network.snapshot_weightings["objective"]

    # ---------------------------------------------------
    # 1. GENERATORS (renewables, nuclear, geothermal...)
    # ---------------------------------------------------
    for carrier in network.generators.carrier.unique():
        if carrier in fossil_carriers:
            continue

        gens = network.generators[network.generators.carrier == carrier]
        opex_no_tc = opex_with_tc = 0

        for gen_name in gens.index:
            if gen_name not in network.generators_t.p.columns:
                continue

            gen = gens.loc[gen_name]
            gen_output = network.generators_t.p[gen_name]

            # Original and current marginal cost
            mc_original = (
                gen["_marginal_cost_original"]
                if "_marginal_cost_original" in network.generators.columns
                else gen["marginal_cost"]
            )
            mc_current = gen["marginal_cost"]

            opex_no_tc += (gen_output * mc_original * w).sum()
            opex_with_tc += (gen_output * mc_current * w).sum()

        if opex_no_tc != 0 or opex_with_tc != 0:
            tax_credit = opex_with_tc - opex_no_tc
            results.append(
                {
                    "tech_label": carrier,
                    "without_tax_credits_billion": opex_no_tc / 1e9,
                    "tax_credits_billion": tax_credit / 1e9,
                    "with_tax_credits_billion": opex_with_tc / 1e9,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

    # ---------------------------------------------------
    # 2. LINKS (fossil and industrial power processes)
    # ---------------------------------------------------
    bus_cols = [c for c in network.links.columns if c.startswith("bus")]

    for carrier in network.links.carrier.unique():
        links = network.links[network.links.carrier == carrier]
        opex_no_tc = opex_with_tc = fuel_cost = 0

        for link_id, link in links.iterrows():
            if link_id not in network.links_t.p0.columns:
                continue

            dispatch = network.links_t.p0[link_id]

            # Marginal costs
            mc_original = (
                link["_marginal_cost_original"]
                if "_marginal_cost_original" in network.links.columns
                else link["marginal_cost"]
            )
            mc_current = link["marginal_cost"]

            opex_no_tc += (dispatch * mc_original * w).sum()
            opex_with_tc += (dispatch * mc_current * w).sum()

            # ---- Additional fuel/feedstock input costs ----
            for bcol in bus_cols:
                pcol = f"p{bcol[3:]}"
                if (
                    pcol not in network.links_t
                    or link_id not in network.links_t[pcol].columns
                ):
                    continue

                eff_key = (
                    f"efficiency{bcol[3:]}"
                    if f"efficiency{bcol[3:]}" in link
                    else "efficiency"
                )
                eff = link.get(eff_key, np.nan)
                if not pd.notna(eff) or eff <= 0:
                    continue  # only positive efficiencies → inputs

                bus = link[bcol]
                if bus not in network.buses_t.marginal_price.columns:
                    continue

                flows = network.links_t[pcol][link_id]
                inflow = -flows.clip(upper=0.0)
                if inflow.abs().sum() == 0:
                    continue

                price = network.buses_t.marginal_price[bus]
                input_cost = (inflow * price * w).sum()
                fuel_cost += input_cost

        # Fuel/feedstock cost not affected by tax credits
        opex_no_tc += fuel_cost
        opex_with_tc += fuel_cost

        if opex_no_tc != 0 or opex_with_tc != 0:
            tax_credit = opex_with_tc - opex_no_tc
            tech_name = f"{carrier} (power)" if carrier in fossil_carriers else carrier
            results.append(
                {
                    "tech_label": tech_name,
                    "without_tax_credits_billion": opex_no_tc / 1e9,
                    "tax_credits_billion": tax_credit / 1e9,
                    "with_tax_credits_billion": opex_with_tc / 1e9,
                    "year": year_str,
                    "scenario": name_tag,
                }
            )

    return pd.DataFrame(results)


def compute_power_capex_with_tax_credits(network, name_tag):
    """
    Compute power CAPEX with and without tax credits.

    Definitions (FINAL):
    - Battery = Store (battery) + battery charger + battery discharger (+ inverter if present)
    - ITC (30%) applies ONLY to Store (battery)
    - Links are NEVER subsidized but MUST be included in totals
    - Uses network.statistics() as single source of truth (annualized CAPEX)

    Output columns:
    - tech_label
    - with_tax_credits_billion
    - without_tax_credits_billion
    - tax_credits_billion
    - year
    - scenario
    """

    year_str = name_tag[-4:]
    ITC_RATE = 0.30
    fossil_carriers = ["coal", "gas", "oil", "biomass"]

    stats = network.statistics()
    results = []

    # -------------------------------------------------
    # 1. GENERATORS (no tax credits)
    # -------------------------------------------------
    for carrier in network.generators.carrier.unique():
        if carrier in fossil_carriers:
            continue

        key = ("Generator", carrier)
        if key not in stats.index:
            continue

        capex = stats.loc[key, "Capital Expenditure"]
        if capex == 0:
            continue

        results.append(
            {
                "tech_label": carrier,
                "with_tax_credits_billion": capex / 1e9,
                "without_tax_credits_billion": capex / 1e9,
                "tax_credits_billion": 0.0,
                "year": year_str,
                "scenario": name_tag,
            }
        )

    # -------------------------------------------------
    # 2. LINKS (power technologies, no tax credits)
    # -------------------------------------------------
    for carrier in network.links.carrier.unique():
        key = ("Link", carrier)
        if key not in stats.index:
            continue

        capex = stats.loc[key, "Capital Expenditure"]
        if capex == 0:
            continue

        tech_name = f"{carrier} (power)" if carrier in fossil_carriers else carrier

        results.append(
            {
                "tech_label": tech_name,
                "with_tax_credits_billion": capex / 1e9,
                "without_tax_credits_billion": capex / 1e9,
                "tax_credits_billion": 0.0,
                "year": year_str,
                "scenario": name_tag,
            }
        )

    # -------------------------------------------------
    # 3. BATTERY (Store + power components)
    # -------------------------------------------------
    battery_link_carriers = [
        "battery charger",
        "battery discharger",
        "battery inverter",
    ]

    battery_link_capex = 0.0

    for carrier in battery_link_carriers:
        key = ("Link", carrier)
        if key in stats.index:
            battery_link_capex += stats.loc[key, "Capital Expenditure"]

    store_capex_with_tc = 0.0
    store_capex_without_tc = 0.0

    if ("Store", "battery") in stats.index:
        store_capex_with_tc = stats.loc[("Store", "battery"), "Capital Expenditure"]
        store_capex_without_tc = store_capex_with_tc / (1 - ITC_RATE)

    battery_with_tc = store_capex_with_tc + battery_link_capex
    battery_without_tc = store_capex_without_tc + battery_link_capex
    battery_tax_credit = battery_with_tc - battery_without_tc

    if battery_with_tc != 0 or battery_without_tc != 0:
        results.append(
            {
                "tech_label": "battery",
                "with_tax_credits_billion": battery_with_tc / 1e9,
                "without_tax_credits_billion": battery_without_tc / 1e9,
                "tax_credits_billion": battery_tax_credit / 1e9,
                "year": year_str,
                "scenario": name_tag,
            }
        )

    return pd.DataFrame(results)


def plot_tax_credit_cluster_bars(
    total_tax_credit_df,
    tech_power_color,
    nice_names_power,
    title="OPEX With vs. Without Tax Credits by Technology",
    width=1400,
    height=700,
    right_margin=260,
    unit="billion",  # "billion" or "million"
    cost_type="OPEX",  # "OPEX" or "CAPEX"
    index="year",
):
    """
    Plot clustered stacked bar chart for tax credit analysis (OPEX or CAPEX).
    Order: With Tax Credits → Without Tax Credits → Tax Credits
    """
    if total_tax_credit_df.empty:
        raise ValueError("total_tax_credit_df is empty; nothing to plot.")

    if unit not in ["billion", "million"]:
        raise ValueError("unit must be either 'billion' or 'million'")
    if cost_type not in ["OPEX", "CAPEX"]:
        raise ValueError("cost_type must be either 'OPEX' or 'CAPEX'")

    if unit == "million":
        conversion_factor = 1000
        y_label = f"{cost_type} (Million USD)"
        precision = ":.2f"
    else:
        conversion_factor = 1
        y_label = f"{cost_type} (Billion USD)"
        precision = ":.3f"

    df_plot = total_tax_credit_df.pivot_table(
        index=index,
        columns="tech_label",
        values=[
            "without_tax_credits_billion",
            "with_tax_credits_billion",
            "tax_credits_billion",
        ],
    )
    df_plot = df_plot * conversion_factor

    years = df_plot.index.tolist()
    if not years:
        raise ValueError("No years found in total_tax_credit_df.")

    year_positions = np.arange(len(years), dtype=float)
    tech_labels = total_tax_credit_df["tech_label"].unique()

    color_lookup = {
        nice_names_power.get(raw_name, raw_name): hex_color
        for raw_name, hex_color in tech_power_color.items()
    }

    # Order changed: With → Without → Tax Credits
    cluster_specs = [
        ("with_tax_credits_billion", f"{cost_type} With Tax Credits", -0.28, 0.85, "/"),
        (
            "without_tax_credits_billion",
            f"{cost_type} Without Tax Credits",
            0.0,
            1.0,
            "",
        ),
        ("tax_credits_billion", "Tax Credits", 0.28, 0.7, "x"),
    ]

    fig = go.Figure()

    for column, label, offset, opacity, pattern_shape in cluster_specs:
        scenario_df = df_plot[column]
        x_vals = year_positions + offset
        for tech in tech_labels:
            if tech in scenario_df.columns:
                values = scenario_df[tech].fillna(0).values
                color = color_lookup.get(tech, tech_power_color.get(tech, "#9E9E9E"))
                marker_kwargs = {"color": color}
                if pattern_shape:
                    marker_kwargs["pattern"] = dict(
                        shape=pattern_shape,
                        fgcolor="#424242",
                        size=6,
                        solidity=0.35,
                    )
                fig.add_bar(
                    x=x_vals,
                    y=values,
                    name=tech if offset == cluster_specs[0][2] else None,
                    legendgroup=tech,
                    showlegend=(offset == cluster_specs[0][2]),
                    marker=marker_kwargs,
                    opacity=opacity,
                    width=0.25,
                    customdata=np.array(years),
                    hovertemplate=(
                        f"<b>{tech}</b><br>{label}<br>"
                        f"Year: %{{customdata}}<br>"
                        f"Value: %{{y{precision}}} {unit.capitalize()} USD<extra></extra>"
                    ),
                )

    style_legend_specs = [
        (f"{cost_type} With Tax Credits", "/", 0.85),
        ("Tax Credits", "x", 0.7),
    ]
    for legend_label, pattern_shape, opacity in style_legend_specs:
        fig.add_bar(
            x=[years[0]],
            y=[0],
            name=legend_label,
            marker=dict(
                color="#BDBDBD",
                pattern=dict(
                    shape=pattern_shape, fgcolor="#424242", size=6, solidity=0.35
                ),
            ),
            opacity=opacity,
            legendgroup=f"style_{pattern_shape}",
            showlegend=True,
            hoverinfo="skip",
            visible="legendonly",
        )

    padding = 0.5 if len(years) > 1 else 0.6
    fig.update_layout(
        title=title,
        barmode="relative",  # negatives appear below zero
        bargap=0.15,
        xaxis=dict(
            title="Year",
            tickmode="array",
            tickvals=year_positions,
            ticktext=years,
            range=[
                year_positions.min() - (padding + 0.2),
                year_positions.max() + padding,
            ],
        ),
        yaxis=dict(title=y_label),
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02,
            font=dict(size=12),
            traceorder="normal",
        ),
        template="plotly_white",
        width=width,
        height=height,
        margin=dict(l=80, r=right_margin, t=50, b=50),
    )

    fig.add_hline(y=0, line_width=1, line_color="black")
    return fig


def plot_geostorage_daily(networks, plot=True, year_title=True):
    """
    For each network extract daily-mean CO2 flows and stock for stores with carrier == "co2 stored".
    Returns:
      combined_df: pd.DataFrame with MultiIndex columns (scenario -> ['flows_MtCO2','stock_MtCO2'])
      scenario_tables: dict mapping scenario -> per-day DataFrame
    If plot=True the original two-panel plot per scenario is shown.
    """

    scenario_tables = {}

    for key, n in networks.items():
        geo_mask = n.stores.carrier == "co2 stored"
        geo_stores = n.stores.index[geo_mask]
        if len(geo_stores) == 0:
            print(f"No permanent CO2 storage in {key}")
            continue

        # time series (MtCO2)
        flows = n.stores_t.p[geo_stores].sum(axis=1) / 1e6
        stock = n.stores_t.e[geo_stores].sum(axis=1) / 1e6

        # ensure DatetimeIndex
        if not isinstance(flows.index, pd.DatetimeIndex):
            flows.index = pd.to_datetime(flows.index)
            stock.index = pd.to_datetime(stock.index)

        # daily mean
        flows_daily = flows.resample("D").mean()
        stock_daily = stock.resample("D").mean()

        # per-scenario table
        df = pd.DataFrame({"flows_MtCO2": flows_daily, "stock_MtCO2": stock_daily})
        scenario_tables[key] = df

        if plot:
            year = key[-4:]
            fig, axes = plt.subplots(2, 1, figsize=(15, 8), sharex=True)
            df["flows_MtCO2"].plot(ax=axes[0], color="tab:blue", lw=1.2)
            axes[0].axhline(0, color="k", lw=0.8)
            axes[0].set_ylabel("Flusso CO₂ [Mt]")
            axes[0].set_title(f"CO2 permanent storage – {year if year_title else key}")

            df["stock_MtCO2"].plot(ax=axes[1], color="tab:green", lw=1.5)
            axes[1].set_ylabel("Stock CO2 (Mt)")
            axes[1].set_xlabel("Time")

            start = df.index.min().replace(day=1)
            end = df.index.max()
            month_starts = pd.date_range(start=start, end=end, freq="MS")
            axes[1].set_xlim(start, end)
            axes[1].set_xticks(month_starts)
            axes[1].set_xticklabels(month_starts.strftime("%b"))

            for ax in axes:
                ax.grid(True)

            plt.tight_layout()
            showfig()

    if not scenario_tables:
        print("No CO2 geostorage data collected from networks.")
        return pd.DataFrame(), {}

    # concat into MultiColumn DataFrame with scenarios as top-level keys
    combined_df = pd.concat(scenario_tables, axis=1, sort=False)
    # ensure column order subcolumns consistent
    # (some scenarios might miss one of the two columns)
    # reorder subcolumns for each top-level scenario to ['flows_MtCO2','stock_MtCO2'] when present
    top_keys = combined_df.columns.levels[0]
    new_cols = []
    for tk in top_keys:
        for sub in ["flows_MtCO2", "stock_MtCO2"]:
            if (tk, sub) in combined_df.columns:
                new_cols.append((tk, sub))
    combined_df = combined_df.reindex(columns=pd.MultiIndex.from_tuples(new_cols))

    return combined_df, scenario_tables


def compute_h2_balance_tables(networks, energy_carriers=None, unit="MWh", plot=False):
    """
    For each scenario/network create:
      - a per-scenario summary DataFrame (index = carriers, columns = [production, consumption, net])
      - a per-scenario details dict (per carrier a DataFrame of per-process values and direction)
    Then collect all per-scenario summaries into a single MultiColumn DataFrame with
    scenarios as the top-level columns.

    Returns:
      combined_summary: pd.DataFrame with MultiIndex columns (scenario -> [production_MWh, consumption_MWh, net_MWh])
                        index = carriers
      per_scenario_tables: dict mapping scenario_name -> {"summary": DataFrame, "details": {carrier: DataFrame}}
    """

    if energy_carriers is None:
        energy_carriers = ["H2", "grid H2", "e-kerosene", "NH3"]

    per_scenario_tables = {}

    for scen_name, net in networks.items():
        # attempt to read the full energy_balance table once
        try:
            eb_all = net.statistics.energy_balance()
        except Exception:
            eb_all = pd.DataFrame()

        # Prepare summary table for this scenario
        summary = pd.DataFrame(
            index=energy_carriers,
            columns=["production_" + unit, "consumption_" + unit, "net_" + unit],
            dtype=float,
        ).fillna(0.0)

        details = {}  # per-carrier detailed breakdown (process -> value)

        for carrier in energy_carriers:
            # default if carrier not present
            if eb_all.empty:
                details[carrier] = pd.DataFrame(columns=["value_" + unit, "direction"])
                continue

            try:
                part = eb_all.xs(carrier, level=2)
            except (KeyError, IndexError):
                # carrier not present in this network
                details[carrier] = pd.DataFrame(columns=["value_" + unit, "direction"])
                continue

            # normalize to a Series of values per process
            if isinstance(part, pd.DataFrame):
                if part.shape[1] == 1:
                    s = part.iloc[:, 0].astype(float)
                else:
                    # if multiple columns aggregate by sum across columns
                    s = part.sum(axis=1).astype(float)
            else:
                s = pd.Series(part).astype(float)

            # convert MultiIndex index entries to readable labels
            def label(idx):
                if isinstance(idx, tuple):
                    return ": ".join([str(x) for x in idx if pd.notna(x) and x != ""])
                return str(idx)

            s.index = s.index.map(label)

            production = s[s > 0].sum()
            consumption = -s[s < 0].sum()  # make positive
            net_val = production - consumption

            summary.loc[carrier, "production_" + unit] = production
            summary.loc[carrier, "consumption_" + unit] = consumption
            summary.loc[carrier, "net_" + unit] = net_val

            # detail DataFrame with direction label
            detail_df = pd.DataFrame(
                {
                    "value_" + unit: s,
                    "direction": np.where(
                        s > 0, "production", np.where(s < 0, "consumption", "zero")
                    ),
                }
            ).sort_values(by="value_" + unit, ascending=False)
            details[carrier] = detail_df

        per_scenario_tables[scen_name] = {"summary": summary, "details": details}

        # Optional plotting (replicates previous visual style but only if requested)
        if plot:
            import matplotlib.gridspec as gridspec

            height_ratios = [max(1, len(details[c])) for c in energy_carriers]
            fig = plt.figure(figsize=(10, sum(height_ratios) * 0.25 + 2))
            gs = gridspec.GridSpec(
                len(energy_carriers), 1, height_ratios=height_ratios, hspace=0.35
            )

            for i, carrier in enumerate(energy_carriers):
                ax = fig.add_subplot(gs[i])
                d = details[carrier]
                if d.empty:
                    ax.text(
                        0.5,
                        0.5,
                        f"No data for {carrier}",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                    )
                    ax.set_title(f"{carrier}: {scen_name}")
                    ax.set_ylabel(None)
                    ax.set_yticks([])
                    continue

                # plot horizontal bar for absolute values, color by direction
                colors = (
                    d["direction"]
                    .map(
                        {
                            "production": "tab:blue",
                            "consumption": "tab:orange",
                            "zero": "grey",
                        }
                    )
                    .tolist()
                )
                d["value_abs"] = d["value_" + unit].abs()
                d["value_abs"].plot(kind="barh", ax=ax, color=colors, edgecolor="k")
                ax.set_title(f"{carrier}: {scen_name}")
                ax.set_ylabel(None)
                ax.grid(True, axis="x")

            plt.tight_layout()
            showfig()

    # Build combined DataFrame: top-level = scenario, second-level = metrics (production, consumption, net)
    summaries = [per_scenario_tables[s]["summary"] for s in per_scenario_tables.keys()]
    combined_summary = pd.concat(
        summaries, axis=1, keys=list(per_scenario_tables.keys()), sort=False
    )

    return combined_summary, per_scenario_tables


def display_state_results(
    networks, eia_generation_data_df, ces, res, ces_carriers, res_carriers
):
    """
    Display RES/CES results aggregated by state in a single MultiIndex table:
    - Index: States.
    - Columns: MultiIndex (scenario, year, metric), e.g., (scenario_01, 2030, % RES).
    - For 2023: Includes EIA comparison with deviation colors.
    - For future years: Only model results with targets.
    - Returns the final DataFrame for reuse (e.g., export or further analysis).
    """

    # Legend for 2023 deviation colors
    legend_html = """
    <div style="padding:10px; margin-bottom:15px; border:1px solid #ccc; border-radius:5px; width: fit-content;">
    <strong>Legend (2023 Only)</strong>
    <ul style="margin:5px 0; padding-left:20px;">
        <li style="background-color:#d4edda; padding:2px;">Diff. ≤ ±5%</li>
        <li style="background-color:#fff3cd; padding:2px;">Diff. > ±5% and ≤ ±10%</li>
        <li style="background-color:#ffe5b4; padding:2px;">Diff. > ±10% and ≤ ±15%</li>
        <li style="background-color:#f8d7da; padding:2px;">Diff. > ±15%</li>
    </ul>
    </div>
    """

    # Get RES/CES data by year
    res_by_network = evaluate_res_ces_by_state(
        networks, ces=ces, res=res, ces_carriers=ces_carriers, res_carriers=res_carriers
    )

    # Collect all data into a big DataFrame
    # {(state): {(scenario, year, metric): value}}
    all_data = defaultdict(dict)
    scenarios_years = set()  # Track all (scenario, year) combinations

    for network_key in sorted(networks.keys()):
        try:
            year = int(network_key[-4:])
        except ValueError:
            continue

        if year in res_by_network:
            df_year = res_by_network[year].copy()
            # Extract scenario
            match = re.search(r"(?:scenario_(\d{2})|Base)_(\d{4})", network_key)
            if match:
                if match.group(1):
                    scenario = f"scenario_{match.group(1)}"
                else:
                    scenario = "Base"
            else:
                continue

            for state in df_year.index:
                # Model values
                all_data[state][(scenario, year, "% RES")] = df_year.at[state, "% RES"]
                all_data[state][(scenario, year, "% CES")] = df_year.at[state, "% CES"]

                # For 2023, add EIA values and targets
                if year == 2023:
                    if state in eia_generation_data_df.index:
                        all_data[state][(scenario, year, "% Actual RES")] = (
                            eia_generation_data_df.at[state, "% Actual RES"]
                        )
                        all_data[state][(scenario, year, "% Actual CES")] = (
                            eia_generation_data_df.at[state, "% Actual CES"]
                        )
                else:
                    # For future years, add targets
                    all_data[state][(scenario, year, "% RES target")] = df_year.at[
                        state, "% RES target"
                    ]
                    all_data[state][(scenario, year, "% CES target")] = df_year.at[
                        state, "% CES target"
                    ]

            scenarios_years.add((scenario, year))

    # Build the MultiIndex DataFrame
    states_list = sorted(all_data.keys())
    sorted_scenarios_years = sorted(scenarios_years, key=lambda x: (x[0], x[1]))

    # Create columns: (scenario, year, metric)
    columns = []
    for scenario, year in sorted_scenarios_years:
        if year == 2023:
            columns.extend(
                [
                    (scenario, year, "% RES"),
                    (scenario, year, "% Actual RES"),
                    (scenario, year, "% CES"),
                    (scenario, year, "% Actual CES"),
                ]
            )
        else:
            columns.extend(
                [
                    (scenario, year, "% RES"),
                    (scenario, year, "% RES target"),
                    (scenario, year, "% CES"),
                    (scenario, year, "% CES target"),
                ]
            )

    # Create the DataFrame
    df_final = pd.DataFrame(
        index=states_list, columns=pd.MultiIndex.from_tuples(columns)
    )

    # Populate the DataFrame
    for state, data_dict in all_data.items():
        for key, value in data_dict.items():
            df_final.at[state, key] = value

    # Styling function for deviation colors (only for 2023) and simple colors for targets
    def style_row(row):
        styles = []
        for col in df_final.columns:
            if col[2] == "% RES" and col[1] == 2023:
                actual_col = (col[0], col[1], "% Actual RES")
                if actual_col in df_final.columns and not pd.isna(row[actual_col]):
                    styles.append(deviation_color(row[col], row[actual_col]))
                else:
                    styles.append("")
            elif col[2] == "% CES" and col[1] == 2023:
                actual_col = (col[0], col[1], "% Actual CES")
                if actual_col in df_final.columns and not pd.isna(row[actual_col]):
                    styles.append(deviation_color(row[col], row[actual_col]))
                else:
                    styles.append("")
            elif col[2] == "% RES target":
                model_col = (col[0], col[1], "% RES")
                if model_col in df_final.columns and not pd.isna(row[model_col]):
                    styles.append(simple_color(row[model_col], row[col]))
                else:
                    styles.append("")
            elif col[2] == "% CES target":
                model_col = (col[0], col[1], "% CES")
                if model_col in df_final.columns and not pd.isna(row[model_col]):
                    styles.append(simple_color(row[model_col], row[col]))
                else:
                    styles.append("")
            else:
                styles.append("")
        return styles

    # Apply styling
    styled_df = df_final.style.format(fmt_2dp_or_na, na_rep="N/A").apply(
        style_row, axis=1
    )

    # Display the legend and table
    display(HTML(legend_html))
    display(styled_df)

    # Return the DataFrame for reuse
    return df_final


def display_grid_region_results_multiple_scenario(
    networks, ces, res, ces_carriers, res_carriers
):
    """
    Collect RES/CES tables for all (scenario, year) and return a single MultiColumn DataFrame
    with scenario names as top-level columns. For each scenario the second level is year and
    the third level are the metrics (e.g. '% RES', '% Actual RES', ...).

    Returns:
      combined_df: pd.DataFrame with MultiIndex columns (scenario, year, metric)
      per_year_dfs: dict mapping (scenario, year) -> per-region DataFrame (raw numeric, not styled)
    """
    res_by_region = evaluate_res_ces_by_region(
        networks, ces_carriers=ces_carriers, res_carriers=res_carriers
    )

    per_year_dfs = {}  # (scenario, year) -> df_disp (index = Grid Region)
    for (scenario, yr), df_year in sorted(res_by_region.items()):
        df_year = df_year.copy()

        if yr == 2023:
            # Load actuals from Excel / helper
            eia_region = preprocess_res_ces_share_grid_region()

            excel_df = pd.read_excel(
                "./data/validation_data/generation_grid_regions.xlsx",
                sheet_name="Generation (TWh)",
            )
            if "Region" in excel_df.columns and "Grid Region" not in excel_df.columns:
                excel_df = excel_df.rename(columns={"Region": "Grid Region"})
            excel_df = excel_df.set_index("Grid Region")

            # Add stats total generation
            eia_region = eia_region.join(excel_df[["Net generation (TWh)"]])

            # Model generation (TWh)
            if "Total (MWh)" in df_year.columns:
                df_year["Model generation (TWh)"] = df_year["Total (MWh)"] / 1e6
            else:
                df_year["Model generation (TWh)"] = np.nan

            # Merge with stats
            df_year = df_year.merge(
                eia_region, left_index=True, right_index=True, how="left"
            )
            df_year = df_year.rename(
                columns={"Net generation (TWh)": "Stats generation (TWh)"}
            )

            # Regional shares (guard against zero)
            if df_year["Model generation (TWh)"].sum() > 0:
                df_year["% Model generation share"] = (
                    df_year["Model generation (TWh)"]
                    / df_year["Model generation (TWh)"].sum()
                    * 100
                )
            else:
                df_year["% Model generation share"] = np.nan

            if df_year["Stats generation (TWh)"].sum() > 0:
                df_year["% Stats generation share"] = (
                    df_year["Stats generation (TWh)"]
                    / df_year["Stats generation (TWh)"].sum()
                    * 100
                )
            else:
                df_year["% Stats generation share"] = np.nan

            # U.S. totals row
            totals = pd.Series(
                {
                    "% RES": (
                        df_year["% RES"] * df_year["Model generation (TWh)"]
                    ).sum()
                    / df_year["Model generation (TWh)"].sum()
                    if df_year["Model generation (TWh)"].sum() > 0
                    else np.nan,
                    "% Actual RES": (
                        df_year.get("% Actual RES", pd.Series(dtype=float))
                        * df_year["Stats generation (TWh)"]
                    ).sum()
                    / df_year["Stats generation (TWh)"].sum()
                    if df_year["Stats generation (TWh)"].sum() > 0
                    else np.nan,
                    "% CES": (
                        df_year["% CES"] * df_year["Model generation (TWh)"]
                    ).sum()
                    / df_year["Model generation (TWh)"].sum()
                    if df_year["Model generation (TWh)"].sum() > 0
                    else np.nan,
                    "% Actual CES": (
                        df_year.get("% Actual CES", pd.Series(dtype=float))
                        * df_year["Stats generation (TWh)"]
                    ).sum()
                    / df_year["Stats generation (TWh)"].sum()
                    if df_year["Stats generation (TWh)"].sum() > 0
                    else np.nan,
                    "Model generation (TWh)": df_year["Model generation (TWh)"].sum(),
                    "Stats generation (TWh)": df_year["Stats generation (TWh)"].sum()
                    if "Stats generation (TWh)" in df_year.columns
                    else np.nan,
                    "% Model generation share": 100.0,
                    "% Stats generation share": 100.0,
                },
                name="U.S.",
            )
            df_year = pd.concat([df_year, totals.to_frame().T])

            df_disp = (
                df_year[
                    [
                        "% RES",
                        "% Actual RES",
                        "% CES",
                        "% Actual CES",
                        "Model generation (TWh)",
                        "Stats generation (TWh)",
                        "% Model generation share",
                        "% Stats generation share",
                    ]
                ]
                .round(2)
                .copy()
            )

        else:
            expected_cols = ["% RES", "% RES target", "% CES", "% CES target"]
            cols_present = [c for c in expected_cols if c in df_year.columns]
            df_year = df_year.reindex(columns=cols_present).round(2)

            # Add model total generation for future years if available
            if "Total (MWh)" in df_year.columns:
                df_year["Model generation (TWh)"] = df_year["Total (MWh)"] / 1e6
                if df_year["Model generation (TWh)"].sum() > 0:
                    df_year["% Model generation share"] = (
                        df_year["Model generation (TWh)"]
                        / df_year["Model generation (TWh)"].sum()
                        * 100
                    )
                else:
                    df_year["% Model generation share"] = np.nan

            # national averages (U.S.)
            totals = pd.Series(
                {c: df_year[c].mean() for c in df_year.columns}, name="U.S."
            )
            df_year = pd.concat([df_year, totals.to_frame().T])

            # keep whatever columns are present + derived
            df_disp = df_year.copy().round(2)

        # normalize index name and ensure Grid Region index
        df_disp = (
            df_disp.reset_index()
            .rename(columns={"index": "Grid Region"})
            .set_index("Grid Region")
        )
        per_year_dfs[(scenario, yr)] = df_disp

    # Build per-scenario frames: second level = year, third level = metric
    scenarios = sorted(set(s for s, _ in per_year_dfs.keys()))
    scenario_frames = {}
    for scen in scenarios:
        years = sorted([yr for (s, yr) in per_year_dfs.keys() if s == scen])
        if not years:
            continue
        dfs = []
        keys = []
        for yr in years:
            df = per_year_dfs.get((scen, yr))
            if df is None:
                continue
            # ensure consistent row index across concatenation later -> leave as-is, will outer-join
            dfs.append(df)
            keys.append(yr)
        # concat year-level frames into one frame where columns are (year, metric)
        if dfs:
            scen_df = pd.concat(dfs, axis=1, keys=keys, sort=False)
        else:
            scen_df = pd.DataFrame()
        scenario_frames[scen] = scen_df

    # Combine scenario frames into final combined_df with top-level = scenario
    if scenario_frames:
        combined_df = pd.concat(scenario_frames, axis=1, sort=False)
        # ensure column names: (scenario, year, metric) -> if deeper nesting produce 3-level columns
        # leave as-is (top-level = scenario) which satisfies requirement
    else:
        combined_df = pd.DataFrame()

    # Return combined dataframe and the raw per-(scenario,year) tables for reuse
    return combined_df, per_year_dfs


def plot_renewable_potential(
    renewable_profile_path, title=None, vmax=0.15, cmap="YlOrRd"
):
    """
    Plot renewable energy potential from Atlite weather data.

    Parameters:
    -----------
    renewable_profile_path : str
        Path to the renewable profile NetCDF file
    title : str, optional
        Custom title for the plot
    vmax : float, optional
        Maximum value for color scale
    exclude_alaska_hawaii : bool, optional
        If True, focus on continental US only
    cmap : str, optional
        Colormap to use for the plot

    Returns:
    --------
    fig, ax : matplotlib objects
        The figure and axis objects
    renewable_ds : xarray.Dataset
        The loaded renewable dataset
    """

    # Load renewable profile data
    renewable_ds = xr.open_dataset(renewable_profile_path)
    renewable_potential = renewable_ds["potential"]

    # Try using imshow with proper extent instead
    fig = plt.figure(figsize=(20, 10))
    ax = plt.axes(projection=ccrs.PlateCarree())

    # Set extent to continental US
    ax.set_extent([-125, -66, 24, 50], crs=ccrs.PlateCarree())

    # Mask zero values
    renewable_potential_masked = renewable_ds["potential"].where(
        renewable_ds["potential"] > 0
    )

    # Get the extent from the coordinates
    x_min, x_max = float(renewable_ds["x"].min()), float(renewable_ds["x"].max())
    y_min, y_max = float(renewable_ds["y"].min()), float(renewable_ds["y"].max())

    # Use imshow with explicit extent
    im = ax.imshow(
        renewable_potential_masked,
        origin="lower",
        extent=[x_min, x_max, y_min, y_max],
        transform=ccrs.PlateCarree(),
        cmap=cmap,
        vmin=0,
        vmax=vmax,
        aspect="auto",
        interpolation="nearest",
    )

    # Add map features
    ax.add_feature(cfeature.STATES, edgecolor="black", linewidth=0.8, alpha=0.7)
    ax.add_feature(cfeature.COASTLINE, linewidth=1.0)
    ax.add_feature(cfeature.BORDERS, linewidth=0.5, linestyle="--", alpha=0.5)

    # Add gridlines with labels
    gl = ax.gridlines(draw_labels=True, alpha=0.3, linestyle="--", linewidth=0.5)
    gl.top_labels = False
    gl.right_labels = False

    # Add colorbar
    cbar = plt.colorbar(im, ax=ax, orientation="vertical", pad=0.02, shrink=0.7)
    cbar.set_label(f"{title} Capacity Factor (Mean Annual)", fontsize=13, weight="bold")

    # Title
    ax.set_title(f"{title} Energy Potential", fontsize=15, weight="bold", pad=20)

    return fig, ax, renewable_ds


def compute_captured_co2_by_tech_from_network(net):
    """
    Compute captured CO2 per real CC technologies only.
    """

    # allowed CC technologies
    cc_techs = {
        "BF-BOF CC",
        "DRI CC",
        "SMR CC",
        "dry clinker CC",
        "ethanol from starch CC",
        "DAC",
    }

    weights = net.snapshot_weightings["objective"]
    bus_cols = [c for c in net.links.columns if c.startswith("bus")]
    p_cols = [f"p{i}" for i in range(len(bus_cols))]

    co2_cap = {}

    for link_name, row in net.links.iterrows():
        tech = row["carrier"]

        # skip everything that's not a CC technology
        if tech not in cc_techs:
            continue

        captured_tons = 0.0

        for j, bus_col in enumerate(bus_cols):
            bus = str(row[bus_col]).lower()
            pcol = p_cols[j]

            if pcol not in net.links_t or link_name not in net.links_t[pcol]:
                continue

            flow = (net.links_t[pcol][link_name] * weights).sum()

            # capture = CO2 flowing **into** the capture link
            if "buffer" in bus or "co2 stored" in bus:
                captured_tons += max(-flow, 0)

        if captured_tons > 0:
            co2_cap[tech] = co2_cap.get(tech, 0) + captured_tons / 1e6

    return co2_cap


def compute_captured_co2_by_tech_from_network(net):
    """
    Compute captured CO2 per real CC technologies only.
    """

    # allowed CC technologies
    cc_techs = {
        "BF-BOF CC",
        "DRI CC",
        "SMR CC",
        "dry clinker CC",
        "ethanol from starch CC",
        "DAC",
    }

    weights = net.snapshot_weightings["objective"]
    bus_cols = [c for c in net.links.columns if c.startswith("bus")]
    p_cols = [f"p{i}" for i in range(len(bus_cols))]

    co2_cap = {}

    for link_name, row in net.links.iterrows():
        tech = row["carrier"]

        # skip everything that's not a CC technology
        if tech not in cc_techs:
            continue

        captured_tons = 0.0

        for j, bus_col in enumerate(bus_cols):
            bus = str(row[bus_col]).lower()
            pcol = p_cols[j]

            if pcol not in net.links_t or link_name not in net.links_t[pcol]:
                continue

            flow = (net.links_t[pcol][link_name] * weights).sum()

            # capture = CO2 flowing **into** the capture link
            if "buffer" in bus or "co2 stored" in bus:
                captured_tons += max(-flow, 0)

        if captured_tons > 0:
            co2_cap[tech] = co2_cap.get(tech, 0) + captured_tons / 1e6

    return co2_cap


def compute_cc_energy_costs(net, year, co2_captured_dict):
    """
    Compute for each CC technology:
        - gas_input_MWh
        - elec_input_MWh
        - cost_gas_USD
        - cost_elec_USD
        - capex_annual_USD
        - co2_cost_USD_tCO2 ( (capex+opex)/captured )
        - co2_captured_Mt
    """

    weights = net.snapshot_weightings["objective"]

    # link structure
    bus_cols = [c for c in net.links.columns if c.startswith("bus")]
    p_cols = [f"p{i}" for i in range(len(bus_cols))]

    # GAS price map
    gas_buses = [b for b in net.buses.index if b.endswith(" gas")]
    gas_price_map = {b: float(net.buses_t.marginal_price[b].mean()) for b in gas_buses}

    # ELECTRICITY price map (AC buses only)
    ac_buses = [b for b in net.buses.index if net.buses.at[b, "carrier"] == "AC"]
    elec_price_map = {b: float(net.buses_t.marginal_price[b].mean()) for b in ac_buses}

    results = []

    for tech in co2_captured_dict.keys():
        links = net.links[net.links.carrier == tech]
        if links.empty:
            continue

        gas_MWh = 0.0
        elec_MWh = 0.0
        cost_gas = 0.0
        cost_elec = 0.0
        capex_annual = 0.0

        for link_name, row in links.iterrows():
            # CAPEX already annualised by PyPSA
            capital_cost = row.get("capital_cost", 0.0)
            p_nom_opt = row.get("p_nom_opt", 0.0)
            capex_annual += capital_cost * p_nom_opt

            # ENERGY INPUTS
            for j, bus_col in enumerate(bus_cols):
                bus = str(row[bus_col])
                pcol = p_cols[j]

                if pcol not in net.links_t or link_name not in net.links_t[pcol]:
                    continue

                flow = (net.links_t[pcol][link_name] * weights).sum()
                f_in = max(flow, 0)

                # GAS
                if bus.endswith(" gas"):
                    gas_MWh += f_in
                    if bus in gas_price_map:
                        cost_gas += f_in * gas_price_map[bus]

                # ELECTRICITY
                elif bus in elec_price_map:
                    elec_MWh += f_in
                    cost_elec += f_in * elec_price_map[bus]

        # CO2 captured
        co2_Mt = co2_captured_dict[tech]
        co2_t = co2_Mt * 1e6

        if co2_t > 0:
            total_cost = capex_annual + cost_gas + cost_elec
            co2_cost = total_cost / co2_t
        else:
            co2_cost = 0.0

        results.append(
            {
                "Technology": tech,
                "co2_captured_Mt": co2_Mt,
                "gas_input_MWh": gas_MWh,
                "elec_input_MWh": elec_MWh,
                "cost_gas_USD": cost_gas,
                "cost_elec_USD": cost_elec,
                "capex_annual_USD": capex_annual,
                "co2_cost_USD_tCO2": co2_cost,
            }
        )

    return pd.DataFrame(results)


def display_cc_summary(df):
    """
    Format and display the CC summary:
      - gas/electricity input in TWh
      - costs in MUSD
      - CO2 captured in Mt
      - Cost per tCO2 renamed to CO2 cost
      - two decimals
      - remove rows with zero captured CO2
      - sorted by CO2 capture cost (cheapest first)
    """

    if df.empty:
        print("No CC data available.")
        return

    # Remove zero-capture rows
    df = df[df["co2_captured_Mt"] > 0].copy()
    if df.empty:
        print("No technologies with CO2 capture.")
        return

    # Unit conversions
    df["Gas input (TWh)"] = df["gas_input_MWh"] / 1e6
    df["Electricity input (TWh)"] = df["elec_input_MWh"] / 1e6
    df["Gas cost (MUSD)"] = df["cost_gas_USD"] / 1e6
    df["Electricity cost (MUSD)"] = df["cost_elec_USD"] / 1e6
    df["CAPEX expenditure (MUSD)"] = df["capex_annual_USD"] / 1e6
    df["Captured CO2 (Mt)"] = df["co2_captured_Mt"]
    df["CO2 capture cost (USD/tCO2)"] = df["co2_cost_USD_tCO2"]

    cols_show = [
        "Technology",
        "Captured CO2 (Mt)",
        "Gas input (TWh)",
        "Electricity input (TWh)",
        "Gas cost (MUSD)",
        "Electricity cost (MUSD)",
        "CAPEX expenditure (MUSD)",
        "CO2 capture cost (USD/tCO2)",
    ]

    df = df[cols_show].copy()

    # Sort: cheapest to most expensive
    df = df.sort_values(by="CO2 capture cost (USD/tCO2)", ascending=True)

    sty = df.style.hide(axis="index").format(
        {
            "Captured CO2 (Mt)": "{:,.2e}",
            "Gas input (TWh)": "{:,.2f}",
            "Electricity input (TWh)": "{:,.2f}",
            "Gas cost (MUSD)": "{:,.2f}",
            "Electricity cost (MUSD)": "{:,.2f}",
            "CAPEX expenditure (MUSD)": "{:,.2f}",
            "CO2 capture cost (USD/tCO2)": "{:,.2f}",
        }
    )

    display(sty)


def compute_aviation_shares(network, level="state"):
    """Compute kerosene and e-kerosene demand and shares by region."""

    carrier_e = "e-kerosene for aviation"
    carrier_f = "kerosene for aviation"

    # Use the same snapshot weights as compute_aviation_fuel_demand()
    weights = network.snapshot_weightings.generators

    buses = network.buses[["state", "grid_region"]]
    all_levels = buses[level].dropna().unique()

    # --- Helper: compute annual regional energy for a given carrier ---
    def get_energy(carrier_name):
        mask = network.loads["carrier"].eq(carrier_name)
        loads = network.loads[mask].copy()

        if loads.empty:
            return pd.DataFrame(
                {level: all_levels, "energy_TWh": np.zeros(len(all_levels))}
            )

        # Time series for these loads (aligned, no reindexing that creates zeros)
        loads_p = network.loads_t.p[loads.index]
        # exact alignment with snapshot weighting
        loads_p = loads_p.loc[weights.index]

        # Apply snapshot weights and sum across time → MWh per load
        e_mwh = loads_p.mul(weights, axis=0).sum()

        # Merge load metadata
        loads = loads.join(buses, on="bus", how="left")
        loads["energy_TWh"] = e_mwh / 1e6

        # Aggregate by region
        df = loads.groupby(level)["energy_TWh"].sum().reset_index()

        # Ensure all regions appear
        df_full = pd.DataFrame({level: all_levels})
        df_full = df_full.merge(df, on=level, how="left").fillna(0.0)
        return df_full

    # Fossil and synthetic kerosene by region
    df_f = get_energy(carrier_f)
    df_e = get_energy(carrier_e)

    df = df_f.merge(df_e, on=level, suffixes=("_kero", "_ekero"))

    # Shares
    total = df["energy_TWh_kero"] + df["energy_TWh_ekero"]
    df["Kerosene share (%)"] = np.where(
        total > 0, df["energy_TWh_kero"] / total * 100, 0
    )
    df["e-kerosene share (%)"] = np.where(
        total > 0, df["energy_TWh_ekero"] / total * 100, 0
    )

    # Rename columns
    df = df.rename(
        columns={
            level: "State" if level == "state" else "Grid region",
            "energy_TWh_kero": "Kerosene cons. (TWh)",
            "energy_TWh_ekero": "e-kerosene cons. (TWh)",
        }
    )

    # Round
    df["Kerosene cons. (TWh)"] = df["Kerosene cons. (TWh)"].round(2)
    df["e-kerosene cons. (TWh)"] = df["e-kerosene cons. (TWh)"].round(2)
    df["Kerosene share (%)"] = df["Kerosene share (%)"].round(2)
    df["e-kerosene share (%)"] = df["e-kerosene share (%)"].round(2)

    # Clean table: index = region
    df = df.set_index(df.columns[0])

    return df


def compute_additionality_compliance_data(
    network,
    region: Optional[str] = None,
    year: Optional[int] = None,
    additionality: bool = True,
    res_carriers: Optional[List[str]] = None,
    res_stor_techs: Optional[List[str]] = None,
    electrolysis_carriers: Optional[List[str]] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> pd.DataFrame:
    """
    Compute RES generation and electrolyzer consumption for additionality compliance analysis.

    For additionality plots, this function ensures that total electrolyzer consumption
    is always less than or equal to eligible RES generation by:
    - Including ALL electrolyzers in the region (all build years)
    - Including RES with build_year >= oldest electrolyzer build year
    This provides a conservative check that sufficient new RES exists to power all electrolyzers.

    Parameters
    ----------
    network : pypsa.Network
        The PyPSA network to analyze
    region : str, optional
        Grid region to filter (e.g., "Plains", "Southeast"). If None, analyze whole country.
    year : int, optional
        Fallback year for additionality filtering if no electrolyzers exist in region.
        If additionality=True, RES with build_year >= oldest electrolyzer (or this year) are included.
    additionality : bool, default True
        If True, filter RES to only include those built >= oldest electrolyzer build year.
        This ensures electrolyzer consumption ≤ eligible RES generation.
    res_carriers : list of str, optional
        List of RES carrier names. Defaults to standard list.
    res_stor_techs : list of str, optional
        List of RES storage tech names. Defaults to ["hydro"].
    electrolysis_carriers : list of str, optional
        List of electrolyzer carrier names. Defaults to standard list.
    start_date : str, optional
        Start date for filtering snapshots (format: 'MM-DD', e.g., '06-01' for June 1st).
        Year is automatically taken from network snapshots. If None, uses first snapshot.
    end_date : str, optional
        End date for filtering snapshots (format: 'MM-DD', e.g., '08-31' for August 31st).
        Year is automatically taken from network snapshots. If None, uses last snapshot.

    Returns
    -------
    pd.DataFrame
        DataFrame with columns for each RES carrier and 'Electrolyzer consumption' (in MWh)
    """
    # Default carrier lists
    if res_carriers is None:
        res_carriers = [
            "csp",
            "solar",
            "onwind",
            "offwind-ac",
            "ror",
            "nuclear",
        ]

    if res_stor_techs is None:
        res_stor_techs = ["hydro"]

    if electrolysis_carriers is None:
        electrolysis_carriers = [
            "H2 Electrolysis",
            "Alkaline electrolyzer large",
            "Alkaline electrolyzer medium",
            "Alkaline electrolyzer small",
            "PEM electrolyzer",
            "SOEC",
        ]

    # Filter by region if specified
    if region is not None:
        region_buses = network.buses.query(f"grid_region == '{region}'").index
        region_gens = network.generators[
            network.generators.bus.isin(region_buses)
        ].index
        region_stor = network.storage_units[
            network.storage_units.bus.isin(region_buses)
        ].index
        region_links = network.links[network.links.bus0.isin(region_buses)].index
    else:
        # Whole country: use all buses/generators/storage
        region_gens = network.generators.index
        region_stor = network.storage_units.index
        region_links = network.links.index

    # Get all electrolyzers in the region
    region_electrolyzers = network.links[
        network.links.carrier.isin(electrolysis_carriers)
        & network.links.index.isin(region_links)
    ]

    filtered_electrolyzers = region_electrolyzers.index

    electrolyzers_consumption = (
        network.links_t.p0[filtered_electrolyzers]
        .multiply(network.snapshot_weightings.objective, axis=0)
        .sum(axis=1)
    )

    # Build dataframe with separate columns per carrier
    res_by_carrier = {}

    for carrier in res_carriers:
        carrier_gens = network.generators.query(
            "carrier == @carrier and index in @region_gens"
        ).index

        if len(carrier_gens) > 0:
            carrier_gen = (
                network.generators_t.p[carrier_gens]
                .multiply(network.snapshot_weightings.objective, axis=0)
                .sum(axis=1)
            )
            res_by_carrier[carrier] = carrier_gen
        else:
            res_by_carrier[carrier] = pd.Series(0, index=network.snapshots)

    # Add storage
    res_storages = network.storage_units.query(
        "carrier in @res_stor_techs and index in @region_stor"
    ).index

    if len(res_storages) > 0:
        res_storages_dispatch = (
            network.storage_units_t.p[res_storages]
            .multiply(network.snapshot_weightings.objective, axis=0)
            .sum(axis=1)
        )
        res_by_carrier["hydro"] = res_storages_dispatch
    else:
        res_by_carrier["hydro"] = pd.Series(0, index=network.snapshots)

    # Create DataFrame with all carriers
    plot_df = pd.DataFrame(res_by_carrier)

    # Add electrolyzer consumption
    plot_df["Electrolyzer consumption"] = electrolyzers_consumption

    # Filter by date range if specified (using month-day format)
    if start_date is not None or end_date is not None:
        # Get the year from the network snapshots
        snapshot_year = plot_df.index[0].year if len(plot_df) > 0 else None

        if snapshot_year is not None:
            if start_date is not None:
                # Parse MM-DD format and add the year from snapshots
                start_timestamp = pd.Timestamp(f"{snapshot_year}-{start_date}")
                plot_df = plot_df[plot_df.index >= start_timestamp]
            if end_date is not None:
                # Parse MM-DD format and add the year from snapshots
                end_timestamp = pd.Timestamp(f"{snapshot_year}-{end_date}")
                plot_df = plot_df[plot_df.index <= end_timestamp]

    return plot_df


def apply_nice_names_and_resample(
    df: pd.DataFrame,
    nice_names_power: Dict[str, str],
    nice_names: Dict[str, str],
    resample_freq: str = "D",
) -> pd.DataFrame:
    """
    Apply nice names to carriers and resample to daily (or other) frequency.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with raw carrier names as columns
    nice_names_power : dict
        Primary mapping for power carrier names
    nice_names : dict
        Fallback mapping for carrier names
    resample_freq : str, default "D"
        Resampling frequency (e.g., "D" for daily, "W" for weekly)

    Returns
    -------
    pd.DataFrame
        Resampled DataFrame with nice names (in GW)
    """
    # Convert to GW and resample
    plot_df = df.div(1e3).resample(resample_freq).mean()

    # Apply nice names
    carrier_rename = {}
    for carrier in df.columns:
        if carrier == "Electrolyzer consumption":
            continue
        nice_carrier = nice_names_power.get(carrier, nice_names.get(carrier, carrier))
        carrier_rename[carrier] = nice_carrier

    plot_df.rename(columns=carrier_rename, inplace=True)

    return plot_df


def plot_additionality_compliance(
    plot_df: pd.DataFrame,
    tech_power_color: Dict[str, str],
    tech_colors: Dict[str, str],
    region: Optional[str] = None,
    scenario_name: Optional[str] = None,
    year: Optional[int] = None,
    additionality: bool = True,
    show_scenario_name: bool = True,
    show_year: bool = True,
    ylim: Optional[Tuple[float, float]] = None,
    figsize: Tuple[float, float] = (18, 4),
    show_plot: bool = True,
) -> plt.Figure:
    """
    Create an area plot showing additionality-compliant RES generation vs electrolyzer consumption.

    Parameters
    ----------
    plot_df : pd.DataFrame
        DataFrame with RES carriers and 'Electrolyzer consumption' column (in GW)
    tech_power_color : dict
        Color mapping for power technologies
    tech_colors : dict
        Fallback color mapping
    region : str, optional
        Grid region name for title. If None, assumes whole country.
    scenario_name : str, optional
        Scenario name for title
    year : int, optional
        Year for title
    additionality : bool, default True
        Whether additionality was applied (for title)
    show_scenario_name : bool, default True
        Whether to include scenario name in title
    show_year : bool, default True
        Whether to include year in title
    ylim : tuple of (float, float), optional
        Y-axis limits (min, max). If None, auto-scale.
    figsize : tuple, default (18, 4)
        Figure size
    show_plot : bool, default True
        Whether to call showfig()

    Returns
    -------
    matplotlib.figure.Figure
        The figure object
    """
    # Separate RES columns and electrolyzer
    electrolyzer_col = "Electrolyzer consumption"
    res_cols = [col for col in plot_df.columns if col != electrolyzer_col]

    # Create color list based on the renamed column names in res_cols
    color_list = []
    for col in res_cols:
        color = tech_power_color.get(col, tech_colors.get(col, "gray"))
        color_list.append(color)

    # Create area plot
    fig, ax = plt.subplots(figsize=figsize)

    # Plot stacked areas for RES carriers
    plot_df[res_cols].plot.area(ax=ax, stacked=True, alpha=0.7, color=color_list)

    # Plot electrolyzer consumption as line on top
    plot_df[electrolyzer_col].plot(
        ax=ax, linewidth=2, color="black", linestyle="-", label=electrolyzer_col
    )

    ax.set_ylabel("(GW)")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
    ax.set_xlabel(None)
    ax.grid(alpha=0.3, zorder=0)

    if ylim is not None:
        ax.set_ylim(ylim)

    # Build title
    title_parts = [
        "Temporal matching and additionality-compliant electricity generation VS Electrolyzer consumption"
    ]

    if region:
        title_parts.append(f"{region} region")
    else:
        title_parts.append("")

    # Add scenario name and/or year based on flags
    subtitle = []
    if show_scenario_name and scenario_name:
        subtitle.append(scenario_name)
    elif show_year and year:
        subtitle.append(year)

    if subtitle:
        title_parts.append(str(subtitle[0]))

    # if additionality:
    #     title_parts.append(
    #         "[RES + nuclear generation (additionality-compliant)]")

    ax.set_title(" ".join(title_parts))

    ax.legend(
        loc="center left",
        bbox_to_anchor=(1.05, 0.5),
        frameon=True,
        fancybox=True,
        shadow=False,
    )

    plt.tight_layout()

    if show_plot:
        showfig()

    return fig


def analyze_additionality_single_scenario(
    network,
    network_name: str,
    region: Optional[str] = None,
    additionality: bool = True,
    nice_names_power: Optional[Dict[str, str]] = None,
    nice_names: Optional[Dict[str, str]] = None,
    tech_power_color: Optional[Dict[str, str]] = None,
    tech_colors: Optional[Dict[str, str]] = None,
    show_scenario_name: bool = True,
    show_year: bool = True,
    show_plot: bool = True,
    **kwargs,
) -> Tuple[pd.DataFrame, plt.Figure]:
    """
    Analyze additionality compliance for a single scenario: compute data, apply names, and create plot.

    Parameters
    ----------
    network : pypsa.Network
        The PyPSA network
    network_name : str
        Network name (used to extract year and for title)
    region : str, optional
        Grid region to analyze. If None, analyze whole country.
    additionality : bool, default True
        Apply additionality filtering
    nice_names_power : dict, optional
        Power carrier name mapping
    nice_names : dict, optional
        Fallback carrier name mapping
    tech_power_color : dict, optional
        Color mapping for power technologies
    tech_colors : dict, optional
        Fallback color mapping
    show_scenario_name : bool, default True
        Whether to include scenario name in plot title
    show_year : bool, default True
        Whether to include year in plot title
    show_plot : bool, default True
        Whether to display the plot
    **kwargs
        Additional arguments passed to compute_additionality_compliance_data

    Returns
    -------
    tuple of (pd.DataFrame, matplotlib.figure.Figure)
        The processed data and the figure object
    """
    # Extract year from network name (assumes format like "scenario_02_2030")
    match = re.search(r"(20\d{2})", network_name)
    year = int(match.group(1)) if match else None

    # Extract scenario name (everything except the year)
    scenario_name = network_name.rsplit("_", 1)[0] if year else network_name

    # Compute raw data
    raw_df = compute_additionality_compliance_data(
        network, region=region, year=year, additionality=additionality, **kwargs
    )

    # Apply nice names and resample
    if nice_names_power is None:
        nice_names_power = {}
    if nice_names is None:
        nice_names = {}

    plot_df = apply_nice_names_and_resample(raw_df, nice_names_power, nice_names)

    # Create plot
    if tech_power_color is None:
        tech_power_color = {}
    if tech_colors is None:
        tech_colors = {}

    fig = plot_additionality_compliance(
        plot_df,
        tech_power_color,
        tech_colors,
        region=region,
        scenario_name=scenario_name,
        year=year,
        additionality=additionality,
        show_scenario_name=show_scenario_name,
        show_year=show_year,
        show_plot=show_plot,
    )

    return plot_df, fig


def analyze_additionality_multiple_networks(
    networks: Dict[str, any],
    regions: Optional[List[str]] = None,
    additionality: bool = True,
    nice_names_power: Optional[Dict[str, str]] = None,
    nice_names: Optional[Dict[str, str]] = None,
    tech_power_color: Optional[Dict[str, str]] = None,
    tech_colors: Optional[Dict[str, str]] = None,
    show_scenario_name: bool = True,
    show_year: bool = True,
    show_plots: bool = True,
    skip_years: Optional[List[int]] = None,
    **kwargs,
) -> Dict[str, Dict[str, Tuple[pd.DataFrame, plt.Figure]]]:
    """
    Analyze additionality compliance across multiple networks and regions, storing results.

    Parameters
    ----------
    networks : dict of {str: pypsa.Network}
        Dictionary mapping network names to network objects
    regions : list of str, optional
        List of grid regions to analyze. If None, only analyze whole country.
        To analyze whole country AND regions, include None in the list.
    additionality : bool, default True
        Apply additionality filtering
    nice_names_power : dict, optional
        Power carrier name mapping
    nice_names : dict, optional
        Fallback carrier name mapping
    tech_power_color : dict, optional
        Color mapping for power technologies
    tech_colors : dict, optional
        Fallback color mapping
    show_scenario_name : bool, default True
        Whether to include scenario name in plot titles
    show_year : bool, default True
        Whether to include year in plot titles
    show_plots : bool, default True
        Whether to display plots
    skip_years : list of int, optional
        Years to skip (e.g., [2023] if hourly matching not implemented)
    **kwargs
        Additional arguments passed to compute_additionality_compliance_data

    Returns
    -------
    dict
        Nested dictionary: {network_name: {region: (dataframe, figure)}}
        where region can be a region name or "whole_country"
    """
    if regions is None:
        regions = [None]  # Analyze whole country by default

    if skip_years is None:
        skip_years = []

    results = {}

    for network_name, network in networks.items():
        # Extract year
        import re

        match = re.search(r"(20\d{2})", network_name)
        year = int(match.group(1)) if match else None

        # Skip if in skip list
        if year and year in skip_years:
            print(f"Skipping {network_name}: year {year} in skip list")
            continue

        results[network_name] = {}

        for region in regions:
            region_key = region if region else "whole_country"

            plot_df, fig = analyze_additionality_single_scenario(
                network,
                network_name,
                region=region,
                additionality=additionality,
                nice_names_power=nice_names_power,
                nice_names=nice_names,
                tech_power_color=tech_power_color,
                tech_colors=tech_colors,
                show_scenario_name=show_scenario_name,
                show_year=show_year,
                show_plot=show_plots,
                **kwargs,
            )

            results[network_name][region_key] = (plot_df, fig)

    return results


def get_all_regions(network) -> List[str]:
    """
    Get all unique grid regions from a network.

    Parameters
    ----------
    network : pypsa.Network
        The PyPSA network

    Returns
    -------
    list of str
        Sorted list of unique region names
    """
    if "grid_region" not in network.buses.columns:
        raise ValueError("Network buses do not have 'grid_region' column")

    return sorted(network.buses["grid_region"].dropna().unique())


def plot_additionality_regions_subplots(
    network,
    network_name: str,
    regions: List[str],
    include_whole_country: bool = True,
    additionality: bool = True,
    skip_no_electrolyzer: bool = True,
    min_electrolyzer_mw: float = 1.0,
    nice_names_power: Optional[Dict[str, str]] = None,
    nice_names: Optional[Dict[str, str]] = None,
    tech_power_color: Optional[Dict[str, str]] = None,
    tech_colors: Optional[Dict[str, str]] = None,
    show_scenario_name: bool = True,
    show_year: bool = True,
    shared_ylim: bool = True,
    ylim: Optional[Tuple[float, float]] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    ncols: int = 2,
    subplot_height: float = 4,
    subplot_width: float = 9,
    show_plot: bool = True,
    **kwargs,
) -> Tuple[Dict[str, pd.DataFrame], plt.Figure]:
    """
    Create subplots showing additionality compliance for all regions (and optionally whole country) for a single scenario.

    Parameters
    ----------
    network : pypsa.Network
        The PyPSA network
    network_name : str
        Network name (used to extract year and for title)
    regions : list of str
        List of grid regions to plot
    include_whole_country : bool, default True
        Whether to include a subplot for the whole country
    additionality : bool, default True
        Apply additionality filtering
    skip_no_electrolyzer : bool, default True
        If True, skip regions with no or negligible electrolyzer consumption
    min_electrolyzer_mw : float, default 1.0
        Minimum average electrolyzer consumption (MW) for a region to be included.
        Only used if skip_no_electrolyzer=True.
    nice_names_power : dict, optional
        Power carrier name mapping
    nice_names : dict, optional
        Fallback carrier name mapping
    tech_power_color : dict, optional
        Color mapping for power technologies
    tech_colors : dict, optional
        Fallback color mapping
    show_scenario_name : bool, default True
        Whether to include scenario name in overall title
    show_year : bool, default True
        Whether to include year in overall title
    shared_ylim : bool, default True
        Whether to use the same y-axis scale across all subplots.
        If True, automatically calculates max value across all regions.
    ylim : tuple of (float, float), optional
        Explicitly set y-axis limits (min, max) for all subplots. If provided, overrides shared_ylim.
    start_date : str, optional
        Start date for plotting (format: 'MM-DD', e.g., '06-01' for June 1st).
        Year is automatically taken from network snapshots. If None, plots from the beginning of the year.
    end_date : str, optional
        End date for plotting (format: 'MM-DD', e.g., '08-31' for August 31st).
        Year is automatically taken from network snapshots. If None, plots to the end of the year.
    ncols : int, default 2
        Number of columns in subplot grid
    subplot_height : float, default 4
        Height of each subplot
    subplot_width : float, default 9
        Width of each subplot
    show_plot : bool, default True
        Whether to display the plot
    **kwargs
        Additional arguments passed to compute_additionality_compliance_data

    Returns
    -------
    tuple of (dict, matplotlib.figure.Figure)
        Dictionary mapping region names to dataframes, and the figure object
    """

    # Extract year from network name
    match = re.search(r"(20\d{2})", network_name)
    year = int(match.group(1)) if match else None
    scenario_name = network_name.rsplit("_", 1)[0] if year else network_name

    # Set defaults
    if nice_names_power is None:
        nice_names_power = {}
    if nice_names is None:
        nice_names = {}
    if tech_power_color is None:
        tech_power_color = {}
    if tech_colors is None:
        tech_colors = {}

    # Prepare region list
    plot_regions = regions.copy()
    if include_whole_country:
        plot_regions = plot_regions + [None]

    # Print date range info if specified
    if start_date or end_date:
        date_range_str = f"Date range: {start_date or 'start'} to {end_date or 'end'}"
        print(f"\n{date_range_str}")

    # First pass: compute all data and filter out regions based on criteria
    print("Checking regions for data...")
    all_plot_data = []
    regions_with_data = []
    skipped_no_res = []
    skipped_no_electrolyzer = []

    for region in plot_regions:
        raw_df = compute_additionality_compliance_data(
            network,
            region=region,
            year=year,
            additionality=additionality,
            start_date=start_date,
            end_date=end_date,
            **kwargs,
        )
        plot_df = apply_nice_names_and_resample(raw_df, nice_names_power, nice_names)

        region_name = region if region else "whole_country"

        # Check if region has any RES generation
        electrolyzer_col = "Electrolyzer consumption"
        res_cols = [col for col in plot_df.columns if col != electrolyzer_col]
        total_res_generation = plot_df[res_cols].sum().sum()

        if total_res_generation <= 0:
            skipped_no_res.append(region_name)
            continue

        # Check if region has electrolyzer consumption (if filtering enabled)
        if skip_no_electrolyzer:
            avg_electrolyzer_mw = plot_df[electrolyzer_col].mean()
            if avg_electrolyzer_mw < min_electrolyzer_mw:
                skipped_no_electrolyzer.append(region_name)
                continue

        # Region passes all filters
        all_plot_data.append(plot_df)
        regions_with_data.append(region)

    # Update plot_regions to only include regions with data
    plot_regions = regions_with_data

    # Print diagnostic information
    if skipped_no_res:
        print(
            f"  Skipped {len(skipped_no_res)} region(s) with no RES generation: {', '.join(skipped_no_res)}"
        )
    if skipped_no_electrolyzer:
        print(
            f"  Skipped {len(skipped_no_electrolyzer)} region(s) with no/negligible electrolyzer consumption: {', '.join(skipped_no_electrolyzer)}"
        )
    if regions_with_data:
        region_names = [r if r else "whole_country" for r in regions_with_data]

    if len(plot_regions) == 0:
        print("\nWarning: No regions with data found after filtering!")
        return {}, None

    # Calculate subplot grid
    nplots = len(plot_regions)
    nrows = math.ceil(nplots / ncols)

    # Calculate max value across all regions for shared y-axis or individual axes check
    max_vals = []
    for plot_df in all_plot_data:
        electrolyzer_col = "Electrolyzer consumption"
        res_cols = [col for col in plot_df.columns if col != electrolyzer_col]
        # Get max of stacked RES sum
        max_res = plot_df[res_cols].sum(axis=1).max()
        # Get max of electrolyzer line
        max_electrolyzer = plot_df[electrolyzer_col].max()
        max_vals.append(max(max_res, max_electrolyzer))

    if shared_ylim and ylim is None:
        max_val = max(max_vals) if max_vals else 0
        # Add 10% padding
        ylim = (0, max_val * 1.1)

    # Create figure with subplots
    fig, axes = plt.subplots(
        nrows=nrows,
        ncols=ncols,
        figsize=(subplot_width * ncols, subplot_height * nrows),
        squeeze=False,
    )
    axes = axes.flatten()

    # Store dataframes
    dataframes = {}

    # Plot each region
    for idx, region in enumerate(plot_regions):
        ax = axes[idx]

        # Use pre-computed data (all_plot_data is always populated now)
        plot_df = all_plot_data[idx]

        # Store dataframe
        region_key = region if region else "whole_country"
        dataframes[region_key] = plot_df

        # Separate RES columns and electrolyzer
        electrolyzer_col = "Electrolyzer consumption"
        res_cols = [col for col in plot_df.columns if col != electrolyzer_col]

        # Create color list
        color_list = []
        for col in res_cols:
            color = tech_power_color.get(col, tech_colors.get(col, "gray"))
            color_list.append(color)

        # Plot stacked areas for RES carriers
        plot_df[res_cols].plot.area(
            ax=ax,
            stacked=True,
            alpha=0.7,
            color=color_list,
            legend=False,  # We'll add a single legend later
        )

        # Plot electrolyzer consumption as line
        plot_df[electrolyzer_col].plot(
            ax=ax,
            linewidth=2,
            color="black",
            linestyle="-",
            label=electrolyzer_col,
            legend=False,
        )

        # Format subplot
        ax.set_ylabel("(GW)", fontsize=10)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b"))
        ax.set_xlabel(None)
        ax.grid(alpha=0.3, zorder=0)

        # Apply y-axis limits if specified
        if ylim is not None:
            ax.set_ylim(ylim)
        else:
            # If not shared_ylim, ensure ylim covers the max value for this specific region
            region_max = max_vals[idx] if idx < len(max_vals) else 0
            ax.set_ylim(0, region_max * 1.1)

        # Build subplot title
        subplot_title = region if region else "Whole Country"
        ax.set_title(subplot_title, fontsize=11)

        # Remove x-axis labels for all but bottom row
        if idx < nplots - ncols:
            ax.set_xticklabels([])

    # Hide unused subplots
    for idx in range(nplots, len(axes)):
        axes[idx].set_visible(False)

    # Add overall title
    title_parts = [
        "Temporal matching and additionality-compliant electricity generation VS Electrolyzer consumption -"
    ]

    # Add scenario name and/or year based on flags
    subtitle = []
    if show_scenario_name and scenario_name:
        subtitle.append(scenario_name)
    if show_year and year:
        subtitle.append(year)

    if subtitle:
        title_parts.append(str(subtitle[0]))

    # if additionality:
    #     title_parts.append("[additionality-compliant]")

    fig.suptitle(" ".join(title_parts), fontsize=14, y=0.995)

    # Create unified legend
    # Get handles and labels from first subplot
    handles, labels = axes[0].get_legend_handles_labels()

    # # Add electrolyzer line to legend
    # from matplotlib.lines import Line2D
    # electrolyzer_handle = Line2D(
    #     [0], [0], color='black', linewidth=2, label='Electrolyzer consumption')
    # handles.append(electrolyzer_handle)
    # labels.append('Electrolyzer consumption')

    # Place legend outside the plot area
    fig.legend(
        handles,
        labels,
        loc="center left",
        bbox_to_anchor=(1.0, 0.5),
        frameon=True,
        fancybox=True,
        shadow=False,
        fontsize=10,
    )

    # Leave space for legend and title
    plt.tight_layout(rect=[0, 0, 0.95, 0.99])

    if show_plot:
        showfig()

    return dataframes, fig


def analyze_additionality_multiple_subplots(
    networks: Dict[str, any],
    regions: Optional[List[str]] = None,
    include_whole_country: bool = True,
    additionality: bool = True,
    skip_no_electrolyzer: bool = True,
    min_electrolyzer_mw: float = 1.0,
    nice_names_power: Optional[Dict[str, str]] = None,
    nice_names: Optional[Dict[str, str]] = None,
    tech_power_color: Optional[Dict[str, str]] = None,
    tech_colors: Optional[Dict[str, str]] = None,
    show_scenario_name: bool = True,
    show_year: bool = True,
    shared_ylim: bool = True,
    ylim: Optional[Tuple[float, float]] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip_years: Optional[List[int]] = None,
    show_plots: bool = True,
    ncols: int = 2,
    subplot_height: float = 4,
    subplot_width: float = 9,
    **kwargs,
) -> Dict[str, Tuple[Dict[str, pd.DataFrame], plt.Figure]]:
    """
    Analyze additionality compliance across multiple networks, creating one subplot figure per network showing all regions.

    This creates a single figure per scenario with subplots for each region (and optionally
    the whole country), making it easy to compare regions within each scenario.

    Parameters
    ----------
    networks : dict of {str: pypsa.Network}
        Dictionary mapping network names to network objects
    regions : list of str, optional
        List of grid regions to plot. If None, automatically detect from first network.
    include_whole_country : bool, default True
        Whether to include a subplot for the whole country
    additionality : bool, default True
        Apply additionality filtering
    skip_no_electrolyzer : bool, default True
        If True, skip regions with no or negligible electrolyzer consumption
    min_electrolyzer_mw : float, default 1.0
        Minimum average electrolyzer consumption (MW) for a region to be included.
        Only used if skip_no_electrolyzer=True.
    nice_names_power : dict, optional
        Power carrier name mapping
    nice_names : dict, optional
        Fallback carrier name mapping
    tech_power_color : dict, optional
        Color mapping for power technologies
    tech_colors : dict, optional
        Fallback color mapping
    show_scenario_name : bool, default True
        Whether to include scenario name in plot titles
    show_year : bool, default True
        Whether to include year in plot titles
    shared_ylim : bool, default True
        Whether to use the same y-axis scale across all subplots within each figure.
        If True, automatically calculates max value across all regions.
    ylim : tuple of (float, float), optional
        Explicitly set y-axis limits (min, max) for all figures. If provided, overrides shared_ylim.
    start_date : str, optional
        Start date for plotting (format: 'MM-DD', e.g., '06-01' for June 1st).
        Year is automatically taken from network snapshots. If None, plots from the beginning of the year.
    end_date : str, optional
        End date for plotting (format: 'MM-DD', e.g., '08-31' for August 31st).
        Year is automatically taken from network snapshots. If None, plots to the end of the year.
    skip_years : list of int, optional
        Years to skip (e.g., [2023] if hourly matching not implemented)
    show_plots : bool, default True
        Whether to display plots
    ncols : int, default 2
        Number of columns in subplot grid
    subplot_height : float, default 4
        Height of each subplot
    subplot_width : float, default 9
        Width of each subplot
    **kwargs
        Additional arguments passed to compute_additionality_compliance_data

    Returns
    -------
    dict
        Dictionary: {network_name: (dict_of_dataframes, figure)}
        where dict_of_dataframes maps region names to dataframes
    """

    if skip_years is None:
        skip_years = []

    # Auto-detect regions if not provided
    if regions is None:
        first_network = list(networks.values())[0]
        regions = get_all_regions(first_network)

    results = {}

    for network_name, network in networks.items():
        # Extract year
        match = re.search(r"(20\d{2})", network_name)
        year = int(match.group(1)) if match else None

        # Skip if in skip list
        if year and year in skip_years:
            continue

        # Create subplot figure for this network
        dataframes, fig = plot_additionality_regions_subplots(
            network=network,
            network_name=network_name,
            regions=regions,
            include_whole_country=include_whole_country,
            additionality=additionality,
            skip_no_electrolyzer=skip_no_electrolyzer,
            min_electrolyzer_mw=min_electrolyzer_mw,
            nice_names_power=nice_names_power,
            nice_names=nice_names,
            tech_power_color=tech_power_color,
            tech_colors=tech_colors,
            show_scenario_name=show_scenario_name,
            show_year=show_year,
            shared_ylim=shared_ylim,
            ylim=ylim,
            start_date=start_date,
            end_date=end_date,
            ncols=ncols,
            subplot_height=subplot_height,
            subplot_width=subplot_width,
            show_plot=show_plots,
            **kwargs,
        )

        results[network_name] = (dataframes, fig)

    return results


def plot_solar_cf_from_network(
    network, cmap="YlOrRd", figsize=(20, 10), plot_type="scatter"
):
    """
    Plot solar capacity factors from a PyPSA network.
    Supported plot types: scatter, nearest, cubic, weighted.
    """

    from matplotlib.lines import Line2D
    from scipy.interpolate import griddata
    from scipy.spatial import cKDTree

    # Filter solar generators and attach coordinates
    solar_gens = network.generators[network.generators.carrier == "solar"].copy()
    solar_gens = solar_gens.join(network.buses[["x", "y"]], on="bus")

    # Compute average capacity factor
    cf_df = network.generators_t.p_max_pu.mean(axis=0).rename("avg_cf").to_frame()
    solar_gens = solar_gens.join(cf_df).dropna(subset=["avg_cf", "x", "y"])
    solar_gens = solar_gens[np.isfinite(solar_gens["p_nom_max"])]

    # Grid for interpolated plot types
    x_grid = np.linspace(solar_gens["x"].min(), solar_gens["x"].max(), 300)
    y_grid = np.linspace(solar_gens["y"].min(), solar_gens["y"].max(), 200)
    Xg, Yg = np.meshgrid(x_grid, y_grid)

    fig, ax = plt.subplots(
        figsize=figsize, subplot_kw={"projection": ccrs.PlateCarree()}
    )

    # Map background
    ax.set_extent([-125, -66, 24, 50], crs=ccrs.PlateCarree())
    ax.add_feature(cfeature.COASTLINE, linewidth=0.5)
    ax.add_feature(cfeature.BORDERS, linewidth=0.5)
    ax.add_feature(cfeature.STATES, linewidth=0.3, edgecolor="gray")
    ax.add_feature(cfeature.LAND, facecolor="lightgray", alpha=0.3)
    ax.add_feature(cfeature.OCEAN, facecolor="lightblue", alpha=0.3)

    gl = ax.gridlines(draw_labels=True, linewidth=0.5, alpha=0.5, linestyle="--")
    gl.top_labels = False
    gl.right_labels = False

    # SCATTER PLOT
    if plot_type == "scatter":
        marker_sizes = solar_gens["p_nom_max"] / 500

        scatter = ax.scatter(
            solar_gens["x"],
            solar_gens["y"],
            c=solar_gens["avg_cf"],
            s=marker_sizes,
            cmap=cmap,
            alpha=0.6,
            edgecolors="black",
            linewidth=0.5,
            transform=ccrs.PlateCarree(),
        )

        plt.colorbar(
            scatter,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

    # NEAREST INTERPOLATION
    elif plot_type == "nearest":
        cf_grid = griddata(
            (solar_gens["x"], solar_gens["y"]),
            solar_gens["avg_cf"],
            (Xg, Yg),
            method="nearest",
        )

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            solar_gens["x"],
            solar_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    # CUBIC INTERPOLATION
    elif plot_type == "cubic":
        cf_grid = griddata(
            (solar_gens["x"], solar_gens["y"]),
            solar_gens["avg_cf"],
            (Xg, Yg),
            method="cubic",
        )

        cf_grid = np.clip(
            cf_grid, solar_gens["avg_cf"].min(), solar_gens["avg_cf"].max()
        )

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            solar_gens["x"],
            solar_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    # WEIGHTED INTERPOLATION
    elif plot_type == "weighted":
        tree = cKDTree(np.column_stack([solar_gens["x"], solar_gens["y"]]))
        pts = np.column_stack([Xg.ravel(), Yg.ravel()])

        d, idx = tree.query(pts, k=5)
        d = np.where(d == 0, 1e-10, d)
        w = 1 / d
        w /= w.sum(axis=1, keepdims=True)

        cf_grid = (solar_gens["avg_cf"].values[idx] * w).sum(axis=1).reshape(Xg.shape)

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            solar_gens["x"],
            solar_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    else:
        raise ValueError(f"Unknown plot_type '{plot_type}'")

    # LEGEND ONLY FOR SCATTER
    if plot_type == "scatter":
        legend_box = fig.add_axes([0.74, 0.30, 0.11, 0.18])

        legend_box.set_facecolor("white")
        legend_box.set_alpha(0.9)
        legend_box.set_xticks([])
        legend_box.set_yticks([])
        legend_box.set_xlim(0, 1)
        legend_box.set_ylim(0, 1)

        legend_box.text(0.05, 0.92, "Max. installable capacity", fontsize=11, va="top")

        caps = [200e3, 1000e3]  # MW
        labels = ["200 GW", "1000 GW"]
        sizes = [c / 500 for c in caps]
        y_pos = [0.60, 0.28]

        for y, s, lab in zip(y_pos, sizes, labels):
            legend_box.scatter(
                0.22, y, s=s, color="gray", alpha=0.6, edgecolors="black", linewidth=0.6
            )
            legend_box.text(0.45, y, lab, va="center", fontsize=11)

    ax.set_title("Solar capacity factor", fontsize=16, pad=20)

    return fig, ax, solar_gens


def plot_wind_cf_from_network(
    network, wind_type="onwind", cmap="viridis", figsize=(20, 10), plot_type="scatter"
):
    """
    Plot wind capacity factors from a PyPSA network.
    Supported plot types: scatter, nearest, cubic, weighted.
    Wind types: 'onwind', 'offwind', or 'both'.
    """

    from matplotlib.lines import Line2D
    from scipy.interpolate import griddata
    from scipy.spatial import cKDTree

    # Filter wind generators based on type
    if wind_type == "both":
        wind_gens = network.generators[
            (network.generators.carrier == "onwind")
            | (network.generators.carrier == "offwind")
        ].copy()
    else:
        wind_gens = network.generators[network.generators.carrier == wind_type].copy()

    # Attach coordinates
    wind_gens = wind_gens.join(network.buses[["x", "y"]], on="bus")

    # Compute average capacity factor
    cf_df = network.generators_t.p_max_pu.mean(axis=0).rename("avg_cf").to_frame()
    wind_gens = wind_gens.join(cf_df).dropna(subset=["avg_cf", "x", "y"])
    wind_gens = wind_gens[np.isfinite(wind_gens["p_nom_max"])]

    # Grid for interpolated plot types
    x_grid = np.linspace(wind_gens["x"].min(), wind_gens["x"].max(), 300)
    y_grid = np.linspace(wind_gens["y"].min(), wind_gens["y"].max(), 200)
    Xg, Yg = np.meshgrid(x_grid, y_grid)

    fig, ax = plt.subplots(
        figsize=figsize, subplot_kw={"projection": ccrs.PlateCarree()}
    )

    # Map background
    ax.set_extent([-125, -66, 24, 50], crs=ccrs.PlateCarree())
    ax.add_feature(cfeature.COASTLINE, linewidth=0.5)
    ax.add_feature(cfeature.BORDERS, linewidth=0.5)
    ax.add_feature(cfeature.STATES, linewidth=0.3, edgecolor="gray")
    ax.add_feature(cfeature.LAND, facecolor="lightgray", alpha=0.3)
    ax.add_feature(cfeature.OCEAN, facecolor="lightblue", alpha=0.3)

    gl = ax.gridlines(draw_labels=True, linewidth=0.5, alpha=0.5, linestyle="--")
    gl.top_labels = False
    gl.right_labels = False

    # SCATTER PLOT
    if plot_type == "scatter":
        marker_sizes = wind_gens["p_nom_max"] / 500

        scatter = ax.scatter(
            wind_gens["x"],
            wind_gens["y"],
            c=wind_gens["avg_cf"],
            s=marker_sizes,
            cmap=cmap,
            alpha=0.6,
            edgecolors="black",
            linewidth=0.5,
            transform=ccrs.PlateCarree(),
        )

        plt.colorbar(
            scatter,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

    # NEAREST INTERPOLATION
    elif plot_type == "nearest":
        cf_grid = griddata(
            (wind_gens["x"], wind_gens["y"]),
            wind_gens["avg_cf"],
            (Xg, Yg),
            method="nearest",
        )

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            wind_gens["x"],
            wind_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    # CUBIC INTERPOLATION
    elif plot_type == "cubic":
        cf_grid = griddata(
            (wind_gens["x"], wind_gens["y"]),
            wind_gens["avg_cf"],
            (Xg, Yg),
            method="cubic",
        )

        cf_grid = np.clip(cf_grid, wind_gens["avg_cf"].min(), wind_gens["avg_cf"].max())

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            wind_gens["x"],
            wind_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    # WEIGHTED INTERPOLATION
    elif plot_type == "weighted":
        tree = cKDTree(np.column_stack([wind_gens["x"], wind_gens["y"]]))
        pts = np.column_stack([Xg.ravel(), Yg.ravel()])

        d, idx = tree.query(pts, k=5)
        d = np.where(d == 0, 1e-10, d)
        w = 1 / d
        w /= w.sum(axis=1, keepdims=True)

        cf_grid = (wind_gens["avg_cf"].values[idx] * w).sum(axis=1).reshape(Xg.shape)

        img = ax.pcolormesh(
            Xg, Yg, cf_grid, cmap=cmap, shading="auto", transform=ccrs.PlateCarree()
        )

        plt.colorbar(
            img,
            ax=ax,
            orientation="horizontal",
            pad=0.05,
            label="Annual average capacity factor (-)",
        )

        ax.scatter(
            wind_gens["x"],
            wind_gens["y"],
            s=20,
            c="black",
            alpha=0.5,
            transform=ccrs.PlateCarree(),
        )

    else:
        raise ValueError(f"Unknown plot_type '{plot_type}'")

    # LEGEND ONLY FOR SCATTER
    if plot_type == "scatter":
        legend_box = fig.add_axes([0.74, 0.30, 0.11, 0.18])

        legend_box.set_facecolor("white")
        legend_box.set_alpha(0.9)
        legend_box.set_xticks([])
        legend_box.set_yticks([])
        legend_box.set_xlim(0, 1)
        legend_box.set_ylim(0, 1)

        legend_box.text(0.05, 0.92, "Max. installable capacity", fontsize=11, va="top")

        caps = [200e3, 1000e3]  # MW
        labels = ["200 GW", "1000 GW"]
        sizes = [c / 500 for c in caps]
        y_pos = [0.60, 0.28]

        for y, s, lab in zip(y_pos, sizes, labels):
            legend_box.scatter(
                0.22, y, s=s, color="gray", alpha=0.6, edgecolors="black", linewidth=0.6
            )
            legend_box.text(0.45, y, lab, va="center", fontsize=11)

    # Set title
    wind_label = wind_type.replace("onwind", "Onshore wind").replace(
        "offwind", "Offshore wind"
    )
    if wind_type == "both":
        wind_label = "Wind (onshore & offshore)"

    ax.set_title(f"{wind_label} capacity factor", fontsize=16, pad=20)

    return fig, ax, wind_gens


def summarize_h2_and_ekerosene_flows(networks):
    """
    Returns two tables (in TWh / year):

    1) Hydrogen flows
    2) e-kerosene flows

    Conventions:
    - Units: TWh / year
    - Positive values = flow INTO the listed process or sink
    - Storage = net annual change of state of charge
    """

    pd.options.display.float_format = "{:.3f}".format

    def get_year(name):
        m = re.search(r"(20\d{2})", name)
        return int(m.group(1)) if m else None

    h2_rows = {}
    eker_rows = {}

    for name, n in networks.items():
        year = get_year(name)
        if year is None:
            continue

        w = n.snapshot_weightings.generators
        scale = 1e-6  # MWh → TWh

        # HYDROGEN
        h2 = {}

        # H2 production (electrolysis)
        el = n.links[n.links.carrier.str.contains("electrolyzer", case=False, na=False)]
        if not el.empty:
            h2["H2 production (electrolysis)"] = -(
                n.links_t.p1[el.index].mul(w, axis=0).sum().sum() * scale
            )

        # H2 to links (by carrier, bus0 = H2)
        h2_links = n.links[n.links.bus0.str.contains("H2", na=False)]

        for carrier in sorted(h2_links.carrier.unique()):
            idx = h2_links.index[h2_links.carrier == carrier]
            val = n.links_t.p0[idx].abs().mul(w, axis=0).sum().sum() * scale

            label = (
                carrier.replace("Pem", "PEM")
                .replace("Soec", "SOEC")
                .replace("Smr Cc", "SMR CC")
                .replace("Smr", "SMR")
                .replace("Cc", "CC")
            )

            h2[f"H2 → {label}"] = val

        # H2 storage (net)
        h2_stores = n.stores[n.stores.carrier.str.contains("H2", na=False)]
        if not h2_stores.empty:
            h2["H2 storage (net)"] = (
                n.stores_t.e[h2_stores.index]
                .iloc[-1]
                .sub(n.stores_t.e[h2_stores.index].iloc[0])
                .sum()
                * scale
            )

        h2_rows[year] = h2

        # e-kerosene
        ek = {}

        # Production (Fischer–Tropsch output only)
        ft = n.links[n.links.carrier == "Fischer-Tropsch"]
        if not ft.empty:
            ek["e-kerosene production (FT)"] = -(
                n.links_t.p1[ft.index].mul(w, axis=0).sum().sum() * scale
            )

        # Dumping to oil
        dump = n.links[n.links.carrier == "e-kerosene-to-oil"]
        if not dump.empty:
            ek["e-kerosene → oil"] = (
                n.links_t.p0[dump.index].abs().mul(w, axis=0).sum().sum() * scale
            )

        # Storage (net)
        ek_stores = n.stores[n.stores.carrier == "e-kerosene"]
        if not ek_stores.empty:
            ek["e-kerosene storage (net)"] = (
                n.stores_t.e[ek_stores.index]
                .iloc[-1]
                .sub(n.stores_t.e[ek_stores.index].iloc[0])
                .sum()
                * scale
            )

        eker_rows[year] = ek

    h2_table = pd.DataFrame(h2_rows).T.sort_index().round(3)
    eker_table = pd.DataFrame(eker_rows).T.sort_index().round(3)

    h2_table.index.name = "Year"
    eker_table.index.name = "Year"

    return h2_table, eker_table


def print_h2_and_ekerosene_tables(h2_flows, ekerosene_flows):
    print("Hydrogen flows (TWh H2 / year)")
    print("Positive values indicate hydrogen flowing INTO the listed process.")
    display(h2_flows)

    print("e-kerosene flows (TWh e-kerosene / year)")
    print("Positive values indicate e-kerosene flowing INTO the listed bus.")
    display(ekerosene_flows)


def compute_marginal_h2_price_by_grid_region(
    networks,
    h2_carriers,
    regional_fees,
    emm_mapping,
    output_threshold=1.0,
    include_baseload=True,
    baseload_charge_path="./data/energy_charge_rate.csv",
    customer_charge_mw=400.0,
    demand_charge_rate=9.0,
    baseload_percentages=None,
    year_title=True,
):
    """
    Compute weighted average marginal H2 price by grid region (USD/kg H2),
    including transmission fees and baseload charges.

    Parameters
    ----------
    year_title : bool, default True
        If True, group results by extracted year (int, e.g. 2030).
        If False, group results by the full network key (scenario name).

    Returns
    -------
    pd.DataFrame with columns:
        - grid_region
        - year (int when year_title=True, str when year_title=False)
        - weighted_price (USD/kg H2)
        - total_h2_output (MWh)
    """

    conv = 1000.0 / 33.0  # kg H2 per MWh H2

    # --------------------------------------------------
    # Baseload charges (keyed by YEAR)
    # --------------------------------------------------
    baseload_charges = {}
    if include_baseload:
        baseload_charges = calculate_baseload_charge(
            networks=networks,
            h2_carriers=h2_carriers,
            emm_mapping=emm_mapping,
            energy_charge_path=baseload_charge_path,
            customer_charge_mw=customer_charge_mw,
            demand_charge_rate=demand_charge_rate,
            baseload_percentages=baseload_percentages,
            output_threshold=output_threshold,
            verbose=False,
            year_title=year_title,
        )

    all_results = []

    # --------------------------------------------------
    # Loop over networks (one per year / scenario)
    # --------------------------------------------------
    for year_key, net in networks.items():
        match = re.search(r"\d{4}", str(year_key))
        if not match:
            continue

        scen_year = int(match.group())

        # Skip base year explicitly if needed
        if scen_year == 2023:
            continue

        # Determine the key for results based on year_title
        key = scen_year if year_title else year_key

        links = net.links[net.links.carrier.isin(h2_carriers)]
        if links.empty:
            continue

        # --------------------------------------------------
        # Flows
        # --------------------------------------------------
        p0 = net.links_t.p0[links.index]
        p1 = net.links_t.p1[links.index]
        w = net.snapshot_weightings.generators

        cons = p0.clip(lower=0).multiply(w, axis=0)  # MWh_el
        h2 = (-p1).clip(lower=0).multiply(w, axis=0)  # MWh_H2
        h2_out = h2.sum()

        valid = h2_out > output_threshold
        if valid.sum() == 0:
            continue

        out_valid = h2_out[valid]

        # --------------------------------------------------
        # Marginal H2 price (from bus duals)
        # --------------------------------------------------
        h2_price = {}
        for l in valid.index[valid]:
            bus = links.at[l, "bus1"]
            h2_price[l] = (h2[l] * net.buses_t.marginal_price[bus]).sum()

        h2_price_val = pd.Series(h2_price) / out_valid / conv  # USD/kg H2

        df = pd.DataFrame(
            {
                "h2_price": h2_price_val,
                "h2_out": out_valid,
                "bus": links.loc[valid, "bus1"],
            }
        )

        df["grid_region"] = df["bus"].map(net.buses["grid_region"])
        df["EMM"] = df["grid_region"].map(emm_mapping)

        # --------------------------------------------------
        # Transmission fees
        # --------------------------------------------------
        fee_map = (
            regional_fees.loc[
                regional_fees["Year"] == scen_year,
                ["region", "Transmission nom USD/MWh"],
            ]
            .set_index("region")
            .squeeze()
        )

        elec_rate = cons.loc[:, valid].sum(axis=0) / out_valid
        df["transmission"] = df["EMM"].map(fee_map).fillna(0.0) * elec_rate / conv

        # --------------------------------------------------
        # Baseload charges (by key)
        # --------------------------------------------------
        if include_baseload and key in baseload_charges:
            bl = baseload_charges[key].set_index("grid_region")[
                "baseload_cost_per_mwh_h2"
            ]
            df["baseload"] = df["grid_region"].map(bl).fillna(0.0) / conv
        else:
            df["baseload"] = 0.0

        # --------------------------------------------------
        # Final price and year/scenario key
        # --------------------------------------------------
        df["price"] = df["h2_price"] + df["transmission"] + df["baseload"]
        df["year"] = key

        all_results.append(df)

    if not all_results:
        return None

    # --------------------------------------------------
    # Aggregate by grid region and year
    # --------------------------------------------------
    all_df = pd.concat(all_results, ignore_index=True)

    region_price = (
        all_df.groupby(["grid_region", "year"])
        .apply(
            lambda g: pd.Series(
                {
                    "weighted_price": (g["price"] * g["h2_out"]).sum()
                    / g["h2_out"].sum(),
                    "total_h2_output": g["h2_out"].sum(),
                }
            )
        )
        .reset_index()
    )

    # --------------------------------------------------
    # Final sanity check (fail fast)
    # --------------------------------------------------
    if year_title and not pd.api.types.is_numeric_dtype(region_price["year"]):
        raise TypeError("Internal error: 'year' must be numeric when year_title=True")

    return region_price


def plot_marginal_h2_price_maps(
    region_price: pd.DataFrame,
    grid_regions_shapes: gpd.GeoDataFrame,
):
    """
    Plot marginal hydrogen price maps by grid region, one map per year.

    Parameters
    ----------
    region_price : pd.DataFrame
        Must contain columns:
            - grid_region
            - year (int)
            - weighted_price (USD/kg H2)

    grid_regions_shapes : gpd.GeoDataFrame
        GeoDataFrame with grid region geometries.
        Must contain a column identifying the grid region.

    Notes
    -----
    - One figure is produced per year.
    - Color scale is shared across all years (5–95 percentile).
    - Scenario names are never used; only numeric years appear.
    """

    # -----------------------------
    # Normalize grid_region column
    # -----------------------------
    shapes = grid_regions_shapes.copy()

    for col in shapes.columns:
        if col.lower().startswith("grid"):
            shapes = shapes.rename(columns={col: "grid_region"})
            break
    else:
        raise KeyError("No grid_region column found in grid_regions_shapes")

    # -----------------------------
    # Sanity checks
    # -----------------------------
    required_cols = {"grid_region", "year", "weighted_price"}
    missing = required_cols - set(region_price.columns)
    if missing:
        raise KeyError(f"region_price missing required columns: {missing}")

    # -----------------------------
    # Merge shapes with data
    # -----------------------------
    plot_df = shapes.merge(region_price, on="grid_region", how="left")

    # -----------------------------
    # Shared color scale
    # -----------------------------
    vmin = plot_df["weighted_price"].quantile(0.05)
    vmax = plot_df["weighted_price"].quantile(0.95)

    # -----------------------------
    # Plot one map per year/scenario
    # -----------------------------
    for year in sorted(plot_df["year"].dropna().unique(), key=str):
        fig, ax = plt.subplots(
            figsize=(12, 10),
            subplot_kw={"projection": ccrs.PlateCarree()},
        )

        year_df = plot_df[plot_df["year"] == year]

        year_df.plot(
            column="weighted_price",
            cmap="RdYlGn_r",
            linewidth=0.8,
            edgecolor="0.8",
            legend=True,
            legend_kwds={"label": "Marginal H2 price (USD/kg H2)"},
            vmin=vmin,
            vmax=vmax,
            ax=ax,
        )

        ax.set_extent([-130, -65, 20, 55])
        ax.axis("off")

        # Use int formatting only for numeric years
        title_label = int(year) if isinstance(year, (int, float, np.integer)) else year
        ax.set_title(f"Hydrogen marginal price by grid region \u2013 {title_label}")

        showfig()


def build_marginal_h2_price_table(
    region_price,
):
    """
    Build one table per year/scenario for marginal H2 prices by grid region.

    Returns
    -------
    dict[int|str, pd.DataFrame]
        {year_or_scenario: DataFrame indexed by grid_region}
    """

    tables = {}

    for year in sorted(region_price["year"].unique(), key=str):
        df = region_price[region_price["year"] == year]

        if df.empty:
            continue

        table = (
            df.set_index("grid_region")[["weighted_price", "total_h2_output"]]
            .rename(
                columns={
                    "weighted_price": "Marginal H2 price (USD/kg H2)",
                    "total_h2_output": "H2 output (MWh)",
                }
            )
            .sort_index()
        )

        # Use int key for numeric years, string key for scenarios
        key = int(year) if isinstance(year, (int, float, np.integer)) else year
        tables[key] = table

    return tables


# Regional Dispatch Plot for Mid-Atlantic
def plot_regional_dispatch(
    network, tech_colors, nice_names, region="Mid-Atlantic", year_str=None
):
    """
    Plot electricity dispatch for Mid-Atlantic region with total demand and data center demand.

    Parameters:
    -----------
    network : pypsa.Network
        PyPSA network object
    tech_colors : dict
        Technology color mapping from plotting.yaml
    nice_names : dict
        Technology name mapping from plotting.yaml
    year_str : str, optional
        Year for title (if None, extracted from network)
    """

    # Filter for Mid-Atlantic region buses
    mid_atlantic_buses = network.buses[network.buses.grid_region == region].index

    if len(mid_atlantic_buses) == 0:
        print("No buses found in Mid-Atlantic region")
        return

    # Calculate dispatch for Mid-Atlantic region
    gen_and_sto_carriers = {
        "csp",
        "solar",
        "onwind",
        "offwind-dc",
        "offwind-ac",
        "nuclear",
        "geothermal",
        "ror",
        "hydro",
        "solar rooftop",
    }
    link_carriers = ["coal", "oil", "OCGT", "CCGT", "biomass", "biomass CHP", "gas CHP"]

    # Generators in Mid-Atlantic
    gen = network.generators[
        network.generators.bus.isin(mid_atlantic_buses)
        & network.generators.carrier.isin(gen_and_sto_carriers)
    ]
    gen_p = network.generators_t.p[gen.index].clip(lower=0)
    gen_dispatch = gen_p.groupby(gen["carrier"], axis=1).sum()

    # Storage units in Mid-Atlantic
    sto = network.storage_units[
        network.storage_units.bus.isin(mid_atlantic_buses)
        & network.storage_units.carrier.isin(gen_and_sto_carriers)
    ]
    sto_p = network.storage_units_t.p[sto.index].clip(lower=0)
    sto_dispatch = sto_p.groupby(sto["carrier"], axis=1).sum()

    # Links in Mid-Atlantic
    link_frames = []
    for carrier in link_carriers:
        links = network.links[
            (network.links.carrier == carrier)
            & (network.links.bus1.isin(mid_atlantic_buses))
        ]
        if links.empty:
            continue
        p1 = network.links_t.p1[links.index].clip(upper=0)
        p1_positive = -p1
        df = p1_positive.groupby(links["carrier"], axis=1).sum()
        link_frames.append(df)

    # Battery in Mid-Atlantic
    battery_links = network.links[
        (network.links.carrier == "battery discharger")
        & (network.links.bus1.isin(mid_atlantic_buses))
    ]
    if not battery_links.empty:
        p1 = network.links_t.p1[battery_links.index].clip(upper=0)
        battery_dispatch = -p1.groupby(battery_links["carrier"], axis=1).sum()
        battery_dispatch.columns = ["battery discharger"]
        link_frames.append(battery_dispatch)

    link_dispatch = (
        pd.concat(link_frames, axis=1)
        if link_frames
        else pd.DataFrame(index=network.snapshots)
    )

    # Combine all generation
    supply_gw = pd.concat([gen_dispatch, sto_dispatch, link_dispatch], axis=1)
    supply_gw = (
        supply_gw.groupby(supply_gw.columns, axis=1).sum().clip(lower=0) / 1000
    )  # Convert to GW

    # Calculate total demand for Mid-Atlantic
    mid_atlantic_loads = network.loads[network.loads.bus.isin(mid_atlantic_buses)]

    # AC loads
    ac_loads = mid_atlantic_loads[mid_atlantic_loads.carrier == "AC"]
    ac_demand = network.loads_t.p_set[
        ac_loads.index.intersection(network.loads_t.p_set.columns)
    ].sum(axis=1)

    # Services and EVs
    serv_idx = [
        i
        for i in mid_atlantic_loads[
            mid_atlantic_loads.carrier == "services electricity"
        ].index
        if i in network.loads_t.p_set.columns
    ]
    ev_idx = [
        i
        for i in mid_atlantic_loads[
            mid_atlantic_loads.carrier == "land transport EV"
        ].index
        if i in network.loads_t.p_set.columns
    ]
    other_idx = [
        i
        for i in mid_atlantic_loads[
            mid_atlantic_loads.carrier == "other electricity"
        ].index
        if i in network.loads_t.p_set.columns
    ]

    serv_demand = network.loads_t.p_set[serv_idx].sum(axis=1) if serv_idx else 0
    ev_demand = network.loads_t.p_set[ev_idx].sum(axis=1) if ev_idx else 0
    other_demand = network.loads_t.p_set[other_idx].sum(axis=1) if other_idx else 0

    # Data center demand
    dc_loads = mid_atlantic_loads[mid_atlantic_loads.carrier == "data center"]
    data_center_demand = dc_loads.p_set.sum()  # Constant profile
    dc_profile = pd.Series(data_center_demand, index=network.snapshots)

    # Static loads
    static_load_carriers = [
        "rail transport electricity",
        "agriculture electricity",
        "industry electricity",
    ]
    static_loads = mid_atlantic_loads[
        mid_atlantic_loads.carrier.isin(static_load_carriers)
    ]
    static_demand = static_loads.p_set.sum()
    static_profile = pd.Series(static_demand, index=network.snapshots)

    # Industrial AC links
    target_processes = [
        "SMR CC",
        "Haber-Bosch",
        "ethanol from starch",
        "ethanol from starch CC",
        "DRI",
        "DRI CC",
        "DRI H2",
        "BF-BOF",
        "BF-BOF CC",
        "EAF",
        "dry clinker",
        "cement finishing",
        "dry clinker CC",
    ]
    process_links = network.links[network.links.carrier.isin(target_processes)]
    ac_input_links = process_links[process_links.bus0.isin(mid_atlantic_buses)].index
    ind_ac_demand = (
        network.links_t.p0[ac_input_links].sum(axis=1) if len(ac_input_links) > 0 else 0
    )

    # Total demand (convert to GW)
    total_demand = (
        ac_demand
        + serv_demand
        + ev_demand
        + other_demand
        + dc_profile
        + static_profile
        + abs(ind_ac_demand)
    ) / 1000
    data_center_demand_gw = dc_profile / 1000

    # Resample to daily averages
    supply_gw.index = pd.to_datetime(supply_gw.index)
    total_demand.index = pd.to_datetime(total_demand.index)
    data_center_demand_gw.index = pd.to_datetime(data_center_demand_gw.index)

    supply_daily = supply_gw.resample("24H").mean()
    demand_daily = total_demand.resample("24H").mean()
    dc_demand_daily = data_center_demand_gw.resample("24H").mean()

    # Define technology order
    ordered_columns = [
        "nuclear",
        "coal",
        "biomass",
        "CCGT",
        "OCGT",
        "oil",
        "hydro",
        "ror",
        "geothermal",
        "gas CHP",
        "biomass CHP",
        "solar",
        "solar rooftop",
        "csp",
        "onwind",
        "offwind-ac",
        "offwind-dc",
        "battery discharger",
    ]

    # Filter and order columns
    supply_daily = supply_daily[
        [c for c in ordered_columns if c in supply_daily.columns]
    ]

    # Create plot
    fig, ax = plt.subplots(figsize=(22, 6))

    # Stacked area for generation
    supply_daily.plot.area(
        ax=ax,
        stacked=True,
        linewidth=0,
        color=[tech_colors.get(c, "gray") for c in supply_daily.columns],
        legend=False,
    )

    # Plot total demand line
    demand_daily.plot(
        ax=ax,
        color="red",
        linewidth=2.5,
        linestyle="-",
        label="Total Demand",
        alpha=0.9,
    )

    # Plot data center demand line
    dc_demand_daily.plot(
        ax=ax,
        color="purple",
        linewidth=2,
        linestyle="--",
        label="Data Center Demand",
        alpha=0.8,
    )

    # Add horizontal line at zero
    ax.axhline(y=0, color="black", linewidth=1.5, linestyle="-", alpha=0.8)

    # Set title
    if year_str is None:
        year_match = re.search(r"\d{4}", str(network))
        year_str = year_match.group() if year_match else "Unknown Year"

    ax.set_title(
        f"Mid-Atlantic Electricity Dispatch & Demand – {year_str}", fontsize=14
    )
    ax.set_ylabel("Power (GW)", fontsize=12)
    ax.set_ylim(0, max(supply_daily.sum(axis=1).max(), demand_daily.max()) * 1.05)
    ax.grid(True, alpha=0.3)

    # Set x-axis formatting
    start = supply_daily.index.min().replace(day=1)
    end = supply_daily.index.max()
    month_starts = pd.date_range(start=start, end=end, freq="MS")

    ax.set_xlim(start, end)
    ax.set_xticks(month_starts)
    ax.set_xticklabels(month_starts.strftime("%b"))
    ax.tick_params(axis="x", which="both", labelbottom=True)
    ax.set_xlabel("Time (months)", fontsize=12)

    # Create legend
    handles, labels = ax.get_legend_handles_labels()
    sums = supply_daily.sum()

    # Filter out zero generation technologies but keep demand lines
    filtered = [
        (h, l)
        for h, l in zip(handles, labels)
        if sums.get(l, 0) > 0 or l in ["Total Demand", "Data Center Demand"]
    ]

    if filtered:
        handles, labels = zip(*filtered)
        pretty_labels = [
            nice_names.get(label, label)
            if label not in ["Total Demand", "Data Center Demand"]
            else label
            for label in labels
        ]

        ax.legend(
            handles,
            pretty_labels,
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            title="Technology",
            fontsize=11,
            title_fontsize=12,
        )

    plt.tight_layout(rect=[0, 0.05, 0.85, 1])
    showfig()


def compute_regional_co2_production_capture_and_ft_price(
    networks: dict,
    ft_carrier: str = "Fischer-Tropsch",
    cc_carriers: set = None,
    year_title: bool = True,
    verbose: bool = True,
):
    """
    Regional CO2 accounting with marginal prices.

    Outputs by Grid Region (per network/year):
      - Captured CO2 (Mt)
      - CO2 used by FT (Mt)
      - Marginal CO2 price – captured (USD/tCO2)
      - Marginal CO2 price – FT (USD/tCO2)

    Marginal prices are computed as the mean shadow price of the CO2 bus
    over snapshots where the corresponding activity is strictly positive.
    """

    if cc_carriers is None:
        raise ValueError("cc_carriers must be provided")

    results = {}

    for name, net in networks.items():
        # -------------------------
        # Scenario year
        # -------------------------
        scen_year = int(re.search(r"\d{4}", str(name)).group())

        # Snapshot duration [h]
        if len(net.snapshots) > 1:
            dt_h = (net.snapshots[1] - net.snapshots[0]).total_seconds() / 3600.0
        else:
            dt_h = 1.0

        rows_cc = []
        rows_ft = []

        # -------------------------
        # CO2 capture (CCS, DAC, etc.)
        # -------------------------
        ccs = net.links[net.links.carrier.isin(cc_carriers)]

        for link in ccs.index:
            captured_t = 0.0
            mu_series = None
            flow_series = None

            for j in range(6):
                p_col = f"p{j}"
                if p_col not in net.links_t or link not in net.links_t[p_col]:
                    continue

                bus = net.links.at[link, f"bus{j}"]
                if not isinstance(bus, str) or "co2" not in bus.lower():
                    continue

                series = net.links_t[p_col][link]

                if series.sum() < 0:
                    captured_t += -series.sum() * dt_h
                    flow_series = -series.clip(upper=0) * dt_h
                    if bus in net.buses_t.marginal_price.columns:
                        mu_series = net.buses_t.marginal_price[bus]

            if captured_t <= 0 or mu_series is None or flow_series is None:
                continue

            active = flow_series > 0
            if not active.any():
                continue

            mu_captured = mu_series[active].mean()

            try:
                region = net.buses.at[ccs.at[link, "bus0"], "grid_region"]
            except KeyError:
                continue

            rows_cc.append(
                {
                    "Grid Region": region,
                    "Captured CO2 (Mt)": captured_t / 1e6,
                    "Marginal CO2 price – captured (USD/tCO2)": mu_captured,
                }
            )

        # -------------------------
        # CO2 used by Fischer–Tropsch
        # -------------------------
        ft_links = net.links.query("carrier == @ft_carrier").index

        for ft in ft_links:
            for i in range(8):
                p_col = f"p{i}"
                eff_col = f"efficiency{i}"
                bus_col = f"bus{i}"

                if (
                    p_col not in net.links_t
                    or ft not in net.links_t[p_col]
                    or eff_col not in net.links.columns
                ):
                    continue

                bus = net.links.at[ft, bus_col]
                eff = net.links.at[ft, eff_col]

                if not isinstance(bus, str) or pd.isna(eff) or "co2" not in bus.lower():
                    continue

                p = net.links_t[p_col][ft]
                co2_used_t = abs(eff) * p * dt_h

                if co2_used_t.sum() <= 0:
                    continue
                if bus not in net.buses_t.marginal_price.columns:
                    continue

                mu = net.buses_t.marginal_price[bus]
                active = co2_used_t > 0
                if not active.any():
                    continue

                mu_ft = mu[active].mean()

                try:
                    region = net.buses.at[net.links.at[ft, "bus0"], "grid_region"]
                except KeyError:
                    continue

                rows_ft.append(
                    {
                        "Grid Region": region,
                        "CO2 used by FT (Mt)": co2_used_t.sum() / 1e6,
                        "Marginal CO2 price – FT (USD/tCO2)": mu_ft,
                    }
                )

        # -------------------------
        # Aggregate by region
        # -------------------------
        df_cc = pd.DataFrame(rows_cc)
        df_ft = pd.DataFrame(rows_ft)

        if df_cc.empty and df_ft.empty:
            continue

        g = (
            df_cc.groupby("Grid Region")
            .mean()
            .join(df_ft.groupby("Grid Region").mean(), how="outer")
            .fillna(0.0)
            .sort_index()
        )

        results[scen_year if year_title else str(name)] = g

    # -------------------------
    # Unified display (safe for Papermill)
    # -------------------------
    if verbose:
        for year, df in results.items():
            print(f"\nYear: {year}")

            df_disp = df.reset_index()

            num_cols = df_disp.select_dtypes(include="number").columns
            fmt = {col: "{:.2f}" for col in num_cols}

            display(df_disp.style.format(fmt).hide(axis="index"))

    return results
