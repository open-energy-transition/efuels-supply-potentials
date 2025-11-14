# SPDX-FileCopyrightText:  Open Energy Transition gGmbH
#
# SPDX-License-Identifier: AGPL-3.0-or-later

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import re
import requests
from io import StringIO

# Define paths
BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "data/consolidated_costs"
SCRIPTS_DIRS = [
    BASE_DIR / "scripts",
    BASE_DIR / "submodules/pypsa-earth/scripts",
]

# GitHub URLs for cost data (raw file links)
COST_URLS = {
    2030: "https://github.com/open-energy-transition/technology-data/raw/refs/heads/master/outputs/US/costs_2030.csv",
    2035: "https://github.com/open-energy-transition/technology-data/raw/refs/heads/master/outputs/US/costs_2035.csv",
    2040: "https://github.com/open-energy-transition/technology-data/raw/refs/heads/master/outputs/US/costs_2040.csv",
}


def extract_used_technologies():
    """
    Parse Python scripts to find all technologies referenced via costs.at[...] calls.
    Returns a set of technology names actually used in the model.
    """
    technologies = set()

    for script_dir in SCRIPTS_DIRS:
        if not script_dir.exists():
            continue

        for py_file in script_dir.rglob("*.py"):
            try:
                with open(py_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # Pattern to match costs.at['technology', ...]
                    pattern = r'costs\.at\[[\"\']([^\"\']+)[\"\']\s*,'
                    matches = re.findall(pattern, content)
                    if matches:
                        technologies.update(matches)
            except Exception:
                pass

    # Add some common technologies that might be referenced differently
    additional_techs = {
        "solar", "solar-rooftop", "solar-utility",
        "onwind", "offwind-ac", "offwind-dc",
        "hydro", "ror", "PHS",
        "battery", "battery storage", "battery inverter",
        "OCGT", "CCGT", "coal", "lignite", "oil", "biomass", "nuclear", "geothermal",
        "H2", "hydrogen storage tank", "hydrogen storage underground",
        "electrolysis", "fuel cell",
        "HVAC overhead", "HVDC overhead", "HVDC submarine", "HVDC inverter pair",
        "csp",
    }
    technologies.update(additional_techs)
    return technologies


def download_costs_data(year):
    """Download costs data for a specific year from GitHub."""
    url = COST_URLS[year]
    print(f"  Downloading {year} data...")
    print(f"    URL: {url}")

    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        df = pd.read_csv(StringIO(response.text))
        return df
    except Exception as e:
        return None


def load_all_years_costs_data():
    """Load and combine costs data from all years."""
    print("\nLoading costs data from multiple years...")

    all_dfs = []
    for year in sorted(COST_URLS.keys()):
        df = download_costs_data(year)
        if df is not None:
            df['year'] = year
            all_dfs.append(df)

    if not all_dfs:
        raise ValueError("No cost data could be loaded!")

    df_combined = pd.concat(all_dfs, ignore_index=True)
    return df_combined


def filter_to_long_format(df, used_technologies, scenario='Moderate', financial_case='Market'):
    """
    Filter costs data to only used technologies and convert to long format.
    Long format: technology | parameter | value | unit | source | currency_year | year
    """

    # Filter by scenario and financial_case if those columns exist
    df_filtered = df.copy()
    if 'scenario' in df.columns and 'financial_case' in df.columns:
        mask = (
            ((df['scenario'] == scenario) | df['scenario'].isna()) &
            ((df['financial_case'] == financial_case)
             | df['financial_case'].isna())
        )
        df_filtered = df[mask].copy()

    # Filter to only used technologies
    df_filtered = df_filtered[df_filtered['technology'].isin(
        used_technologies)].copy()

    # Select columns for long format (including year)
    df_long = df_filtered[['technology', 'parameter', 'value',
                           'unit', 'source', 'currency_year', 'year']].copy()

    # Clean up
    df_long['currency_year'] = df_long['currency_year'].apply(
        lambda x: str(int(x)) if pd.notna(x) and isinstance(
            x, (int, float)) else str(x) if pd.notna(x) else ''
    )
    df_long['source'] = df_long['source'].fillna('').astype(str)
    df_long['year'] = df_long['year'].astype(int)

    return df_long


def categorize_technologies_long(df_long):
    """
    Categorize technologies into three groups for long format data.
    Returns three DataFrames.
    """

    # Define technology categories
    electricity_techs = {
        'solar', 'solar-rooftop', 'solar-utility', 'solar rooftop',
        'onwind', 'offwind', 'offwind-ac', 'offwind-dc',
        'offwind-ac-station', 'offwind-dc-station',
        'offwind-ac-connection-submarine', 'offwind-ac-connection-underground',
        'offwind-dc-connection-submarine', 'offwind-dc-connection-underground',
        'hydro', 'ror', 'PHS',
        'csp', 'csp-tower', 'csp-tower TES', 'csp-tower power block',
        'nuclear',
        'CCGT', 'OCGT',
        'coal', 'lignite',
        'biomass', 'solid biomass',
        'oil',
        'geothermal',
        'battery inverter', 'battery', 'battery storage',
        'HVAC overhead', 'HVDC overhead', 'HVDC submarine', 'HVDC inverter pair',
        'home battery',
    }

    hydrogen_ft_techs = {
        'electrolysis', 'Alkaline electrolyzer large size', 'Alkaline electrolyzer medium size',
        'Alkaline electrolyzer small size', 'PEM electrolyzer', 'SOEC',
        'fuel cell', 'H2 Fuel Cell',
        'hydrogen storage tank', 'hydrogen storage underground',
        'H2 pipeline', 'H2 (g) pipeline', 'H2 liquefaction',
        'methanation', 'helmeth',
        'Fischer-Tropsch',
        'SMR',
        'H2', 'H2 Electrolysis',
    }

    # Keywords for fallback categorization
    industrial_keywords = ['ethanol', 'starch', 'dri', 'bf-bof', 'blast furnace',
                           'clinker', 'cement', 'steel', 'ammonia', 'haber-bosch',
                           'dac', 'direct air capture', 'industry', 'process emissions',
                           'iron', 'ore', 'scrap', 'eaf', 'bioethanol']

    h2_keywords = ['h2', 'hydrogen',
                   'electrolyzer', 'electrolysis', 'fuel cell']

    # Categorize based on explicit sets
    df_elec = df_long[df_long['technology'].isin(electricity_techs)].copy()
    df_h2 = df_long[df_long['technology'].isin(hydrogen_ft_techs)].copy()

    # Get remaining technologies
    remaining = df_long[~df_long['technology'].isin(electricity_techs) &
                        ~df_long['technology'].isin(hydrogen_ft_techs)].copy()

    # Use keywords for remaining
    industrial_mask = remaining['technology'].str.lower().apply(
        lambda x: any(
            kw in x for kw in industrial_keywords) or 'cc' in x.lower()
    )
    df_industrial = remaining[industrial_mask].copy()

    # Assign remaining h2-related to hydrogen category
    remaining = remaining[~industrial_mask]
    h2_mask = remaining['technology'].str.lower().apply(
        lambda x: any(kw in x for kw in h2_keywords)
    )
    df_h2 = pd.concat([df_h2, remaining[h2_mask]], ignore_index=True)

    # Everything else goes to industrial
    df_industrial = pd.concat(
        [df_industrial, remaining[~h2_mask]], ignore_index=True)

    return df_elec, df_h2, df_industrial


def get_technology_subcategories(df_long):
    """
    Determine sub-categories for technologies to enable Excel grouping.
    Returns a dict mapping technology -> subcategory string.
    """
    subcats = {}

    for tech in df_long['technology'].unique():
        tech_lower = tech.lower()

        # Solar
        if 'solar' in tech_lower:
            if 'rooftop' in tech_lower:
                subcats[tech] = 'Solar - Rooftop'
            elif 'utility' in tech_lower:
                subcats[tech] = 'Solar - Utility Scale'
            elif 'csp' in tech_lower or 'tower' in tech_lower:
                subcats[tech] = 'Solar - CSP'
            else:
                subcats[tech] = 'Solar - Other'

        # Wind
        elif 'wind' in tech_lower:
            if 'onwind' in tech_lower or 'onshore' in tech_lower:
                subcats[tech] = 'Wind - Onshore'
            elif 'offwind' in tech_lower or 'offshore' in tech_lower:
                subcats[tech] = 'Wind - Offshore'
            else:
                subcats[tech] = 'Wind - Other'

        # Hydro
        elif 'hydro' in tech_lower or 'ror' in tech_lower or 'phs' in tech_lower:
            if 'phs' in tech_lower or 'pumped' in tech_lower:
                subcats[tech] = 'Hydro - Pumped Storage'
            elif 'ror' in tech_lower or 'river' in tech_lower:
                subcats[tech] = 'Hydro - Run-of-River'
            else:
                subcats[tech] = 'Hydro - Conventional'

        # Battery
        elif 'battery' in tech_lower:
            subcats[tech] = 'Storage - Battery'

        # Hydrogen
        elif 'h2' in tech_lower or 'hydrogen' in tech_lower:
            if 'storage' in tech_lower:
                subcats[tech] = 'Storage - Hydrogen'
            elif 'electrolyz' in tech_lower or 'electroly' in tech_lower:
                subcats[tech] = 'Hydrogen - Electrolyzers'
            elif 'fuel cell' in tech_lower:
                subcats[tech] = 'Hydrogen - Fuel Cells'
            elif 'pipeline' in tech_lower:
                subcats[tech] = 'Hydrogen - Transport'
            else:
                subcats[tech] = 'Hydrogen - Other'

        # Fossil fuels
        elif any(x in tech_lower for x in ['ccgt', 'ocgt', 'coal', 'gas', 'oil', 'lignite']):
            subcats[tech] = 'Fossil Fuel Generation'

        # Other generation
        elif any(x in tech_lower for x in ['biomass', 'geothermal', 'nuclear']):
            if 'biomass' in tech_lower:
                subcats[tech] = 'Biomass'
            elif 'geothermal' in tech_lower:
                subcats[tech] = 'Geothermal'
            elif 'nuclear' in tech_lower:
                subcats[tech] = 'Nuclear'

        # Industrial
        elif any(x in tech_lower for x in ['ethanol', 'ammonia', 'haber', 'cement', 'steel',
                                           'dri', 'bf-bof', 'clinker', 'dac']):
            if 'ethanol' in tech_lower:
                subcats[tech] = 'Industrial - Ethanol'
            elif 'ammonia' in tech_lower or 'haber' in tech_lower:
                subcats[tech] = 'Industrial - Ammonia'
            elif 'cement' in tech_lower or 'clinker' in tech_lower:
                subcats[tech] = 'Industrial - Cement'
            elif 'steel' in tech_lower or 'dri' in tech_lower or 'bf-bof' in tech_lower or 'eaf' in tech_lower:
                subcats[tech] = 'Industrial - Steel'
            elif 'dac' in tech_lower:
                subcats[tech] = 'Industrial - DAC'
            else:
                subcats[tech] = 'Industrial - Other'

        # Transmission
        elif any(x in tech_lower for x in ['hvac', 'hvdc', 'line', 'transmission']):
            subcats[tech] = 'Transmission'

        else:
            subcats[tech] = 'Other'

    return subcats


def apply_excel_formatting_to_sheet(ws, df_long):
    """
    Apply professional Excel formatting to a worksheet.
    Includes headers, column widths, freeze panes, filters, conditional formatting, and grouping.
    """
    # 1. Format header row
    header_fill = PatternFill(start_color="366092",
                              end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    # 2. Auto-adjust column widths
    column_widths = {
        'A': 35,  # technology
        'B': 25,  # parameter
        'C': 12,  # value
        'D': 20,  # unit
        'E': 60,  # source
        'F': 15,  # currency_year
        'G': 10,  # year
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 3. Freeze panes (freeze header row)
    ws.freeze_panes = "A2"

    # 4. Enable AutoFilter
    ws.auto_filter.ref = ws.dimensions

    # 5. Conditional formatting for missing values (NaN)
    ws.conditional_formatting.add(
        f'C2:C{ws.max_row}',
        CellIsRule(
            operator='equal',
            formula=['"nan"'],
            fill=PatternFill(start_color="FFC7CE",
                             end_color="FFC7CE", fill_type="solid"),
            font=Font(color="9C0006")
        )
    )

    # 6. Conditional formatting for zero values
    ws.conditional_formatting.add(
        f'C2:C{ws.max_row}',
        CellIsRule(
            operator='equal',
            formula=['0'],
            fill=PatternFill(start_color="FFEB9C",
                             end_color="FFEB9C", fill_type="solid")
        )
    )

    # 7. Row grouping by subcategory
    subcats = get_technology_subcategories(df_long)
    df_with_subcat = df_long.copy()
    df_with_subcat['subcategory'] = df_with_subcat['technology'].map(subcats)
    df_with_subcat = df_with_subcat.sort_values(
        ['subcategory', 'technology', 'parameter', 'year'])

    current_subcat = None
    start_row = 2

    for idx, row in enumerate(df_with_subcat.itertuples(), start=2):
        subcat = row.subcategory
        if current_subcat != subcat:
            if current_subcat is not None and idx > start_row + 1:
                ws.row_dimensions.group(start_row, idx - 1, hidden=False)
            current_subcat = subcat
            start_row = idx

    if start_row < ws.max_row:
        ws.row_dimensions.group(start_row, ws.max_row, hidden=False)

    # 8. Add light borders to all cells
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=False)


def write_excel_file_single(df_elec, df_h2, df_industrial, filename):
    """
    Write all three DataFrames to a single Excel file with 3 sheets.
    """
    filepath = OUTPUT_DIR / filename

    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Write each DataFrame to a separate sheet
        datasets = [
            (df_elec, "Electricity Generation", "1_Electricity_Generation"),
            (df_h2, "Hydrogen & Fischer-Tropsch", "2_Hydrogen_FischerTropsch"),
            (df_industrial, "Industrial & DAC", "3_Industrial_DAC")
        ]

        for df_long, title, sheet_name in datasets:

            # Sort by technology, parameter, and year
            df_sorted = df_long.sort_values(
                ['technology', 'parameter', 'year']).reset_index(drop=True)

            # Write to sheet
            df_sorted.to_excel(writer, index=False, sheet_name=sheet_name)

    # Load the workbook and apply formatting to each sheet
    wb = openpyxl.load_workbook(filepath)
    sheet_names = ["1_Electricity_Generation",
                   "2_Hydrogen_FischerTropsch", "3_Industrial_DAC"]
    dfs = [df_elec, df_h2, df_industrial]

    for sheet_name, df_long in zip(sheet_names, dfs):
        ws = wb[sheet_name]
        apply_excel_formatting_to_sheet(ws, df_long)

    wb.save(filepath)

    return filepath


def main():
    """Main execution function."""

    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

    # 1. Extract technologies used in the model
    used_technologies = extract_used_technologies()

    # 2. Load data from multiple years (download from GitHub)
    df_costs = load_all_years_costs_data()

    # Process both scenarios
    scenarios = ['Moderate', 'Advanced']
    output_files = []

    for scenario in scenarios:

        # 3. Filter and convert to long format for this scenario
        df_long = filter_to_long_format(
            df_costs, used_technologies, scenario=scenario, financial_case='Market')

        # 4. Categorize technologies
        df_elec, df_h2, df_industrial = categorize_technologies_long(df_long)

        # 5. Write single Excel file with 3 sheets
        filename = f"model_input_costs_multiyear_{scenario.lower()}.xlsx"
        output_file = write_excel_file_single(
            df_elec, df_h2, df_industrial, filename)

        output_files.append(
            (scenario, output_file, df_elec, df_h2, df_industrial))

    # Final summary
    for scenario, filepath, df_elec, df_h2, df_industrial in output_files:
        print(f"\n{scenario} Scenario:")
        print(f"  File: model_input_costs_multiyear_{scenario.lower()}.xlsx")
        print(
            f"  - Electricity Generation: {df_elec['technology'].nunique()} technologies, {len(df_elec)} rows")
        print(
            f"  - Hydrogen & FT: {df_h2['technology'].nunique()} technologies, {len(df_h2)} rows")
        print(
            f"  - Industrial & DAC: {df_industrial['technology'].nunique()} technologies, {len(df_industrial)} rows")
        print(f"  Path: {filepath}")

    print(f"\nAll files saved in: {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
